#!/usr/bin/env python3
"""
Generate remaining Excel files - Part 3
Purchase Orders, Timesheets, Defects
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# =============================================================================
# FILE 4: Purchase_Orders_Master.xlsx
# =============================================================================
def generate_purchase_orders():
    """Generate purchase orders master file"""
    filename = os.path.join(BASE_DIR, "dummy_data/06_PURCHASE_ORDERS_INVOICES/Purchase_Orders_Master.xlsx")

    wb = Workbook()

    # SHEET 1: PO Register
    ws1 = wb.active
    ws1.title = "PO Register"

    ws1['A1'] = "PURCHASE ORDER REGISTER"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:H1')

    headers = ["PO#", "Date", "Supplier", "Description", "Amount", "Delivery Date", "Invoice Received", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # Generate 30+ purchase orders
    suppliers = [
        ("Bob's Hardware", "Timber & fixings"),
        ("ReadyMix Concrete", "Concrete supply"),
        ("Spark Electrical", "Electrical materials"),
        ("Aqua Plumbing", "Plumbing supplies"),
        ("Timber Supplies Co", "Structural timber"),
        ("Pacific Paint", "Paint & materials"),
        ("BuildMart", "General supplies"),
        ("Steel Direct", "Steel beams"),
        ("Window World", "Windows & doors"),
        ("Tile Palace", "Floor & wall tiles"),
    ]

    start_date = datetime(2024, 7, 1)
    row = 4
    for i in range(30):
        po_num = f"PO-{2024}-{1000+i}"
        po_date = start_date + timedelta(days=i*3)
        supplier, desc = random.choice(suppliers)
        amount = random.randint(1000, 25000)
        delivery_date = po_date + timedelta(days=random.randint(7, 21))
        invoice_received = "YES" if random.random() < 0.7 else "NO"
        status = random.choice(["COMPLETE", "COMPLETE", "COMPLETE", "PENDING", "DELIVERED"])

        ws1.cell(row=row, column=1).value = po_num
        ws1.cell(row=row, column=2).value = po_date.strftime("%d/%m/%Y")
        ws1.cell(row=row, column=3).value = supplier
        ws1.cell(row=row, column=4).value = desc
        ws1.cell(row=row, column=5).value = amount
        ws1.cell(row=row, column=5).number_format = '$#,##0'
        ws1.cell(row=row, column=6).value = delivery_date.strftime("%d/%m/%Y")
        ws1.cell(row=row, column=7).value = invoice_received
        ws1.cell(row=row, column=8).value = status

        # Highlight pending invoices
        if invoice_received == "NO":
            ws1.cell(row=row, column=7).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        row += 1

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 12

    print(f"✓ Sheet 1: PO Register created (30 orders)")

    # SHEET 2: Invoice Matching
    ws2 = wb.create_sheet("Invoice Matching")
    ws2['A1'] = "PO TO INVOICE MATCHING"
    ws2['A1'].font = Font(size=14, bold=True)

    headers2 = ["PO#", "Invoice#", "PO Amount", "Invoice Amount", "Variance", "Notes"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Sample matching data
    matches = [
        ["PO-2024-1000", "BH-2024-0847", 2600, 2868, -268, "Incl. GST difference"],
        ["PO-2024-1001", "RM-2024-8845", 7200, 7872, -672, "Extra pump fees"],
        ["PO-2024-1002", "SES-2024-3421", 2400, 2399, 1, "OK"],
        ["PO-2024-1003", "APS-2024-8912", 3800, 3872, -72, "OK"],
        ["PO-2024-1004", "TSC-INV-4421", 5700, 5759, -59, "OK"],
    ]

    for row_idx, match in enumerate(matches, 4):
        for col_idx, value in enumerate(match, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = value

            if col_idx in [3, 4, 5]:
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0'

            # Highlight variances
            if col_idx == 5 and isinstance(value, (int, float)) and value != 0:
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    print(f"✓ Sheet 2: Invoice Matching created")

    wb.save(filename)
    print(f"✓✓ Generated: {filename}\n")

# =============================================================================
# FILE 5: Timesheets_September_2024.xlsx
# =============================================================================
def generate_timesheets():
    """Generate timesheet file"""
    filename = os.path.join(BASE_DIR, "dummy_data/08_LABOUR_TIMESHEETS/Timesheets_September_2024.xlsx")

    wb = Workbook()

    # SHEET 1: Site Supervisor
    ws1 = wb.active
    ws1.title = "Site Supervisor"

    ws1['A1'] = "TIMESHEET - SITE SUPERVISOR - SEPTEMBER 2024"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:F1')

    ws1['A2'] = "Name: Tom Richards"
    ws1['A2'].font = Font(bold=True)
    ws1['A3'] = "Rate: $75/hour"

    headers = ["Date", "Day", "Start Time", "End Time", "Hours", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Generate September days (20 working days)
    start = datetime(2024, 9, 2)  # First Monday of Sept
    row = 6
    total_hours = 0

    for i in range(20):
        date = start + timedelta(days=i)
        if date.weekday() < 5:  # Monday-Friday
            hours = random.choice([8, 8.5, 9, 10])
            total_hours += hours

            ws1.cell(row=row, column=1).value = date.strftime("%d/%m/%Y")
            ws1.cell(row=row, column=2).value = date.strftime("%A")
            ws1.cell(row=row, column=3).value = "7:00 AM"
            ws1.cell(row=row, column=4).value = f"{int(7 + hours)}:{'00' if hours % 1 == 0 else '30'} PM"
            ws1.cell(row=row, column=5).value = hours
            ws1.cell(row=row, column=6).value = random.choice(["Site supervision", "Inspections", "Subbie coordination", "Client meeting", "Council inspection"])

            row += 1

    # Totals
    row += 1
    ws1.cell(row=row, column=4).value = "TOTAL HOURS:"
    ws1.cell(row=row, column=4).font = Font(bold=True)
    ws1.cell(row=row, column=5).value = total_hours
    ws1.cell(row=row, column=5).font = Font(bold=True, size=12)
    ws1.cell(row=row, column=5).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    row += 1
    ws1.cell(row=row, column=4).value = "TOTAL PAY:"
    ws1.cell(row=row, column=4).font = Font(bold=True)
    ws1.cell(row=row, column=5).value = total_hours * 75
    ws1.cell(row=row, column=5).number_format = '$#,##0.00'
    ws1.cell(row=row, column=5).font = Font(bold=True, size=12)
    ws1.cell(row=row, column=5).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    print(f"✓ Sheet 1: Site Supervisor timesheet created ({total_hours} hours)")

    # SHEET 2: Labourers
    ws2 = wb.create_sheet("Labourers")
    ws2['A1'] = "LABOURER TIMESHEETS - SEPTEMBER 2024"
    ws2['A1'].font = Font(size=14, bold=True)

    headers2 = ["Name", "Date", "Hours", "Rate", "Amount", "Task"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    labourers = ["Mark Wilson", "Steve Brown", "Peter Lee", "Andrew Chen"]
    row = 4

    for labourer in labourers:
        for i in range(10):  # 10 days each
            date = start + timedelta(days=i)
            hours = random.choice([8, 8, 8.5, 9])
            rate = 35

            ws2.cell(row=row, column=1).value = labourer
            ws2.cell(row=row, column=2).value = date.strftime("%d/%m/%Y")
            ws2.cell(row=row, column=3).value = hours
            ws2.cell(row=row, column=4).value = rate
            ws2.cell(row=row, column=4).number_format = '$#,##0'
            ws2.cell(row=row, column=5).value = hours * rate
            ws2.cell(row=row, column=5).number_format = '$#,##0.00'
            ws2.cell(row=row, column=6).value = random.choice(["General labour", "Cleanup", "Material handling", "Site prep"])

            row += 1

    print(f"✓ Sheet 2: Labourers created (4 workers)")

    # SHEET 3: Weekly Summary
    ws3 = wb.create_sheet("Weekly Summary")
    ws3['A1'] = "WEEKLY LABOUR SUMMARY - SEPTEMBER 2024"
    ws3['A1'].font = Font(size=14, bold=True)

    headers3 = ["Week Ending", "Total Hours", "Total Cost", "Notes"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    weeks = [
        ["08/09/2024", 248, 10920, "Normal week"],
        ["15/09/2024", 232, 10220, "2 days lost to rain"],
        ["22/09/2024", 256, 11280, "Extra hours for roof completion"],
        ["29/09/2024", 240, 10560, "Normal week"],
    ]

    for row_idx, week in enumerate(weeks, 4):
        for col_idx, value in enumerate(week, 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.value = value

            if col_idx == 3:
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0'

    print(f"✓ Sheet 3: Weekly Summary created")

    wb.save(filename)
    print(f"✓✓ Generated: {filename}\n")

# =============================================================================
# FILE 6: Defects_And_Snagging.xlsx
# =============================================================================
def generate_defects():
    """Generate defects and snagging list"""
    filename = os.path.join(BASE_DIR, "dummy_data/15_DEFECTS_SNAGGING/Defects_And_Snagging.xlsx")

    wb = Workbook()

    ws = wb.active
    ws.title = "Defects List"

    ws['A1'] = "DEFECTS & SNAGGING LIST - PROJECT A"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:I1')

    headers = ["ID", "Location", "Trade", "Description", "Severity", "Reported Date", "Due Date", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    defects = [
        ["D001", "Main Bathroom", "Plumbing", "Shower leak - needs waterproofing repair", "CRITICAL", "2024-09-05", "2024-09-12", "OVERDUE", "Blocking handover!"],
        ["D002", "Kitchen", "Electrical", "Power point not working", "HIGH", "2024-09-10", "2024-09-17", "IN PROGRESS", "Electrician booked"],
        ["D003", "Bedroom 2", "Painting", "Touch up wall paint", "LOW", "2024-09-15", "2024-09-22", "PENDING", ""],
        ["D004", "Living Room", "Flooring", "Timber floor scratch", "MEDIUM", "2024-09-12", "2024-09-19", "PENDING", "Needs re-sanding"],
        ["D005", "Front Entry", "Carpentry", "Door not closing properly", "MEDIUM", "2024-09-08", "2024-09-15", "FIXED", "Adjusted hinges"],
        ["D006", "Laundry", "Plumbing", "Tap dripping", "LOW", "2024-09-18", "2024-09-25", "PENDING", ""],
        ["D007", "Garage", "Electrical", "Light switch faulty", "LOW", "2024-09-16", "2024-09-23", "PENDING", ""],
        ["D008", "Master Bedroom", "Plastering", "Ceiling crack", "MEDIUM", "2024-09-14", "2024-09-21", "FIXED", "Patched and painted"],
        ["D009", "Bathroom 2", "Tiling", "Grout needs cleaning", "LOW", "2024-09-17", "2024-09-24", "PENDING", ""],
        ["D010", "External", "Brickwork", "Mortar spillage on paving", "LOW", "2024-09-11", "2024-09-18", "FIXED", "Cleaned"],
        ["D011", "Staircase", "Carpentry", "Handrail loose", "HIGH", "2024-09-13", "2024-09-20", "IN PROGRESS", "Safety issue"],
        ["D012", "Kitchen", "Cabinets", "Drawer not aligned", "LOW", "2024-09-19", "2024-09-26", "PENDING", ""],
        ["D013", "Ensuite", "Plumbing", "Basin drain slow", "MEDIUM", "2024-09-10", "2024-09-17", "FIXED", "Cleared"],
        ["D014", "Living Room", "Electrical", "Downlight not centered", "LOW", "2024-09-18", "2024-09-25", "PENDING", ""],
        ["D015", "External", "Landscaping", "Pavers uneven", "MEDIUM", "2024-09-15", "2024-09-22", "PENDING", "Landscaper to fix"],
        ["D016", "Bedroom 3", "Windows", "Window hard to open", "MEDIUM", "2024-09-12", "2024-09-19", "FIXED", "Adjusted"],
        ["D017", "Kitchen", "Benchtop", "Small chip in stone", "LOW", "2024-09-16", "2024-09-23", "PENDING", ""],
        ["D018", "Garage", "Painting", "Ceiling needs touch up", "LOW", "2024-09-17", "2024-09-24", "PENDING", ""],
        ["D019", "External", "Roof", "Tile alignment issue", "MEDIUM", "2024-09-09", "2024-09-16", "FIXED", "Re-laid tiles"],
        ["D020", "Laundry", "Cabinets", "Door hinge squeaks", "LOW", "2024-09-18", "2024-09-25", "PENDING", "Needs oil"],
    ]

    for row_idx, defect in enumerate(defects, 4):
        for col_idx, value in enumerate(defect, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Color code severity
            if col_idx == 5:
                if value == "CRITICAL":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif value == "HIGH":
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(bold=True)
                elif value == "MEDIUM":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Color code status
            if col_idx == 8:
                if value == "OVERDUE":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif value == "IN PROGRESS":
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif value == "FIXED":
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25

    print(f"✓ Defects list created (20 defects - 1 CRITICAL OVERDUE)")

    wb.save(filename)
    print(f"✓✓ Generated: {filename}\n")

if __name__ == "__main__":
    print("\n=== Part 3: Final Excel Files ===")
    generate_purchase_orders()
    generate_timesheets()
    generate_defects()
    print("\n✓✓✓ All Excel files generated successfully!")
