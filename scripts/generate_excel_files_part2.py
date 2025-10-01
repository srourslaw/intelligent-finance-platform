#!/usr/bin/env python3
"""
Generate remaining Excel files - Part 2
Subcontractors, Client Payments, Purchase Orders, Timesheets, Defects
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# =============================================================================
# FILE 2: Subcontractor_Register.xlsx
# =============================================================================
def generate_subcontractor_register():
    """Generate subcontractor register with messy data"""
    filename = os.path.join(BASE_DIR, "dummy_data/07_SUBCONTRACTORS/Subcontractor_Register.xlsx")

    wb = Workbook()

    # SHEET 1: Active Subbies
    ws1 = wb.active
    ws1.title = "Active Subbies"

    ws1['A1'] = "SUBCONTRACTOR REGISTER - PROJECT A"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:J1')

    headers = ["ID", "Company Name", "Contact", "Phone", "Email", "ABN", "License", "Insurance Expiry", "Contract Value", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    subbies = [
        ["SUB-001", "BigDig Excavations", "Tony M", "0412 345 678", "tony@bigdig.com.au", "51 234 567 890", "EXC-12345", "2025-06-30", 12120, "COMPLETE"],
        ["SUB-002", "Precision Concrete Solutions", "James Chen", "(02) 5555-1234", "james@precisioncon.com", "22445667889", "123456C", "2024-12-31", 49300, "COMPLETE"],
        ["SUB-003", "FastFrame Carpentry", "Tom Williams", "04 1234 5678", "tom@fastframe.com.au", "33 556 778 990", "234567C", "2025-08-15", 46500, "IN PROGRESS"],
        ["SUB-004", "TopRoof Tilers", "Sarah Brown", "0423456789", "", "44 667 889 001", "345678C", "2025-04-20", 23650, "IN PROGRESS"],
        ["SUB-005", "Premier Bricklaying", "David Lee", "(02) 9876-5432", "david@premierbrick.com", "55778889991", "456789C", "2025-07-10", 42500, "COMPLETE"],
        ["SUB-006", "John's Plumbing Services", "Robert Chen", "0434567890", "rob@johnsplumbing.com.au", "66 889 990 112", "PL-234567", "2025-05-25", 19150, "IN PROGRESS"],
        ["SUB-007", "Bright Spark Electrical", "David Patterson", "04 2345 6789", "david@brightspark.com", "44223445667", "EL-123456", "2025-09-30", 28540, "IN PROGRESS"],
        ["SUB-008", "AirMaster HVAC", "Michael Tang", "(02) 8888-1234", "", "77 990 112 233", "AC-567890", "2025-03-15", 12500, "COMPLETE"],
        ["SUB-009", "Complete Plastering Co", "Steve Wilson", "0445678901", "steve@completeplaster.com", "88112233445", "345678C", "2025-06-01", 20350, "NOT STARTED"],
        ["SUB-010", "PremierTile", "Lisa Wang", "04 3456 7890", "lisa@premiertile.com.au", "99 223 344 556", "234567C", "2025-07-20", 12300, "NOT STARTED"],
        ["SUB-011", "ColorPro Painting", "John Smith", "(02) 7777-1111", "john@colorpro.com", "11 334 455 667", "234567C", "2025-08-10", 14500, "NOT STARTED"],
        ["SUB-012", "Green Oasis Landscaping", "Emma Green", "0456789012", "", "22 445 556 778", "LAND-12345", "2025-05-15", 6800, "NOT STARTED"],
        ["SUB-013", "SkyHigh Scaffolding", "Peter Brown", "04 5678 9012", "peter@skyhigh.com.au", "33 556 667 889", "SCAF-67890", "2025-04-30", 4200, "COMPLETE"],
        ["SUB-014", "SealTight Waterproofing", "Andrew Kim", "0467890123", "andrew@sealtight.com", "44 667 778 990", "234567C", "2023-12-31", 2800, "COMPLETE"],
        ["SUB-015", "StrongSteel Reinforcement", "Mark Johnson", "(02) 6666-2222", "", "55 778 889 001", "345678C", "2025-06-20", 6000, "COMPLETE"],
    ]

    for row_idx, subbie in enumerate(subbies, 4):
        for col_idx, value in enumerate(subbie, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Format currency
            if col_idx == 9:
                cell.number_format = '$#,##0'

            # Highlight expired insurance
            if col_idx == 8 and isinstance(value, str):
                try:
                    exp_date = datetime.strptime(value, "%Y-%m-%d")
                    if exp_date < datetime.now():
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                except:
                    pass

    # Set column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 30
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 15

    print(f"✓ Sheet 1: Active Subbies created (15 contractors)")

    # SHEET 2: Payment Schedule
    ws2 = wb.create_sheet("Payment Schedule")
    ws2['A1'] = "SUBCONTRACTOR PAYMENT SCHEDULE"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:I1')

    headers2 = ["Payment ID", "Subcontractor", "Invoice #", "Description", "Amount", "GST", "Total", "Due Date", "Status"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    payments = [
        ["PAY-001", "BigDig Excavations", "BD-2024-001", "Excavation complete", 11018, 1102, 12120, "2024-08-15", "PAID"],
        ["PAY-002", "Precision Concrete", "PC-2024-045", "Footing pour", 18000, 1800, 19800, "2024-08-20", "PAID"],
        ["PAY-003", "Precision Concrete", "PC-2024-067", "Slab pour", 27545, 2755, 30300, "2024-09-10", "PAID"],
        ["PAY-004", "FastFrame Carpentry", "FC-PC-001", "Progress Claim 1", 9091, 909, 10000, "2024-09-15", "PAID"],
        ["PAY-005", "FastFrame Carpentry", "FC-PC-002", "Progress Claim 2", 13636, 1364, 15000, "2024-10-05", "PAID"],
        ["PAY-006", "TopRoof Tilers", "TR-2024-156", "Roof tiling progress", 21500, 2150, 23650, "2024-10-15", "OUTSTANDING"],
        ["PAY-007", "Premier Bricklaying", "PB-INV-089", "Brickwork complete", 38636, 3864, 42500, "2024-09-20", "OVERDUE"],
        ["PAY-008", "John's Plumbing", "JP-PC-001", "Rough-in complete", 8682, 868, 9550, "2024-09-25", "PAID"],
        ["PAY-009", "John's Plumbing", "JP-PC-002", "Second fix", 8682, 868, 9550, "2024-10-20", "OUTSTANDING"],
        ["PAY-010", "Bright Spark Electrical", "BSE-PC-001", "Rough-in", 12300, 1230, 13530, "2024-09-30", "PAID"],
        ["PAY-011", "Bright Spark Electrical", "BSE-PC-002", "Second fix progress", 13727, 1373, 15100, "2024-10-25", "NOT DUE"],
        ["PAY-012", "AirMaster HVAC", "AM-2024-456", "HVAC install complete", 11364, 1136, 12500, "2024-09-15", "PAID"],
        ["PAY-013", "SkyHigh Scaffolding", "SH-2024-789", "Scaffolding hire 4 weeks", 3818, 382, 4200, "2024-09-05", "PAID"],
        ["PAY-014", "SealTight Waterproofing", "ST-2024-234", "Waterproofing", 2545, 255, 2800, "2024-08-30", "PAID"],
        ["PAY-015", "StrongSteel", "SS-2024-567", "Steel fixing", 5455, 545, 6000, "2024-08-25", "PAID"],
    ]

    for row_idx, payment in enumerate(payments, 4):
        for col_idx, value in enumerate(payment, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Format currency
            if col_idx in [5, 6, 7]:
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'

            # Highlight overdue
            if col_idx == 9 and value == "OVERDUE":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif col_idx == 9 and value == "OUTSTANDING":
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)

    print(f"✓ Sheet 2: Payment Schedule created (15 payments)")

    # SHEET 3: Contact Log
    ws3 = wb.create_sheet("Contact Log")
    ws3['A1'] = "SUBCONTRACTOR COMMUNICATION LOG"
    ws3['A1'].font = Font(size=14, bold=True)
    ws3.merge_cells('A1:E1')

    headers3 = ["Date", "Subcontractor", "Contact", "Subject", "Notes"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    contacts = [
        ["2024-08-15", "BigDig", "Tony", "Start date", "Confirmed excavation start Mon 19/8"],
        ["2024-08-25", "Precision Concrete", "James", "Slab booking", "Booked for 5/9 - weather permitting"],
        ["2024-09-01", "FastFrame", "Tom", "Timber delivery", "Delayed 2 days - supplier issue"],
        ["2024-09-10", "TopRoof", "Sarah", "Tile selection", "Client to confirm color by Fri"],
        ["2024-09-15", "Premier Brick", "David", "Payment query", "Chasing payment - URGENT"],
        ["2024-09-20", "John's Plumbing", "Robert", "Fixture delivery", "Delayed - arriving 25/9"],
    ]

    for row_idx, contact in enumerate(contacts, 4):
        for col_idx, value in enumerate(contact, 1):
            ws3.cell(row=row_idx, column=col_idx).value = value

    ws3.column_dimensions['E'].width = 40

    print(f"✓ Sheet 3: Contact Log created")

    wb.save(filename)
    print(f"✓✓ Generated: {filename}\n")

# =============================================================================
# FILE 3: Client_Payment_Tracker.xlsx
# =============================================================================
def generate_client_payments():
    """Generate client payment tracker"""
    filename = os.path.join(BASE_DIR, "dummy_data/11_CLIENT_BILLING/Client_Payment_Tracker.xlsx")

    wb = Workbook()

    # SHEET 1: Payment Schedule
    ws1 = wb.active
    ws1.title = "Payment Schedule"

    ws1['A1'] = "CLIENT PAYMENT SCHEDULE - PROJECT A"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:G1')

    headers = ["Milestone", "Invoice#", "Description", "Amount", "Due Date", "Paid Date", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    milestones = [
        ["Deposit (10%)", "INV-C-001", "Deposit on contract signing", 65000, "2024-01-20", "2024-01-18", "PAID"],
        ["Base/Slab (15%)", "INV-C-002", "Completion of slab", 97500, "2024-08-15", "2024-08-12", "PAID"],
        ["Frame (20%)", "INV-C-003", "Frame complete and roof on", 130000, "2024-09-25", "", "OVERDUE"],
        ["Lock-up (20%)", "INV-C-004", "Brickwork, windows, roof tiles", 130000, "TBC", "", "NOT INVOICED"],
        ["Fixing (15%)", "INV-C-005", "Internal fixing, plaster, paint", 97500, "TBC", "", "NOT INVOICED"],
        ["Practical (15%)", "INV-C-006", "Practical completion", 97500, "TBC", "", "NOT INVOICED"],
        ["Final/Defects (5%)", "INV-C-007", "Final completion + defects", 32500, "TBC", "", "NOT INVOICED"],
    ]

    for row_idx, milestone in enumerate(milestones, 4):
        for col_idx, value in enumerate(milestone, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = value

            if col_idx == 4:  # Amount column
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0'

            # Highlight status
            if col_idx == 7:
                if value == "PAID":
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    cell.font = Font(bold=True)
                elif value == "OVERDUE":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif value == "NOT INVOICED":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Totals
    row = len(milestones) + 5
    ws1.cell(row=row, column=3).value = "TOTAL CONTRACT:"
    ws1.cell(row=row, column=3).font = Font(bold=True)
    ws1.cell(row=row, column=4).value = 650000
    ws1.cell(row=row, column=4).number_format = '$#,##0'
    ws1.cell(row=row, column=4).font = Font(bold=True, size=12)

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 35
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 15

    print(f"✓ Sheet 1: Payment Schedule created (7 milestones)")

    # SHEET 2: Variations Register
    ws2 = wb.create_sheet("Variations")
    ws2['A1'] = "CLIENT VARIATIONS REGISTER"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:G1')

    headers2 = ["VO#", "Date", "Description", "Cost", "Client Price", "Status", "Invoiced"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    variations = [
        ["VO-001", "2024-08-15", "Kitchen benchtop upgrade to stone", 3682, 4500, "APPROVED", "YES"],
        ["VO-002", "2024-08-20", "Additional bedroom window", 982, 1200, "APPROVED", "NO"],
        ["VO-003", "2024-09-01", "Bathroom tile upgrade", 2291, 2800, "APPROVED", "NO"],
        ["VO-004", "2024-09-05", "Extra balcony waterproofing", 1227, 1500, "APPROVED", "NO"],
        ["VO-005", "2024-09-10", "Ducted AC upgrade from split", 2618, 3200, "PENDING", "NO"],
        ["VO-006", "2024-09-15", "Extra power points x8", 785, 960, "APPROVED", "NO"],
    ]

    for row_idx, var in enumerate(variations, 4):
        for col_idx, value in enumerate(var, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = value

            if col_idx in [4, 5]:  # Cost columns
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0'

            # Highlight not invoiced
            if col_idx == 7 and value == "NO":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(bold=True)

    # Revenue leakage note
    row = len(variations) + 5
    ws2.cell(row=row, column=1).value = "REVENUE LEAKAGE:"
    ws2.cell(row=row, column=1).font = Font(bold=True, color="FF0000", size=12)
    ws2.cell(row=row, column=5).value = 6460
    ws2.cell(row=row, column=5).number_format = '$#,##0'
    ws2.cell(row=row, column=5).font = Font(bold=True, color="FF0000", size=12)

    print(f"✓ Sheet 2: Variations created (6 variations, $6,460 not invoiced)")

    # SHEET 3: Communications
    ws3 = wb.create_sheet("Communications")
    ws3['A1'] = "CLIENT COMMUNICATIONS"
    ws3['A1'].font = Font(size=14, bold=True)

    headers3 = ["Date", "Type", "Subject", "Notes"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    comms = [
        ["2024-08-15", "Email", "Payment reminder", "Sent invoice INV-C-002"],
        ["2024-09-01", "Site Meeting", "Variation requests", "Discussed kitchen and bathroom upgrades"],
        ["2024-09-25", "Email", "Payment overdue", "INV-C-003 now 5 days overdue - called client"],
        ["2024-09-30", "Phone", "Progress update", "Frame complete, moving to lock-up phase"],
    ]

    for row_idx, comm in enumerate(comms, 4):
        for col_idx, value in enumerate(comm, 1):
            ws3.cell(row=row_idx, column=col_idx).value = value

    ws3.column_dimensions['D'].width = 50

    print(f"✓ Sheet 3: Communications created")

    wb.save(filename)
    print(f"✓✓ Generated: {filename}\n")

# Continue in next script...

if __name__ == "__main__":
    print("\n=== Part 2: Additional Excel Files ===")
    generate_subcontractor_register()
    generate_client_payments()
