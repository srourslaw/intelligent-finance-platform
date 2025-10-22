"""
Part 2: Recreate remaining Excel spreadsheets (Purchase Orders, Subcontractors, Labour, Site Reports, etc.)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os

# Color schemes
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

BASE_PATH = "/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data"

def apply_header_style(ws, row=1):
    """Apply professional header styling"""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def apply_borders_to_range(ws, start_row, end_row, start_col, end_col):
    """Apply borders to a range of cells"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = THIN_BORDER

def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================================
# 06_PURCHASE_ORDERS_INVOICES (5 files)
# ============================================================================

def create_steel_supplier_po():
    """06_PURCHASE_ORDERS_INVOICES/Steel_Supplier_PO_2024_001.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    # PO Header
    ws['A1'] = "PURCHASE ORDER"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = "PO Number: PO-2024-001"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = "Date: 2024-07-15"
    ws['A4'] = "Supplier: Ace Steel Reinforcement Pty Ltd"
    ws['A5'] = "Delivery Date: 2024-07-25"

    # Line Items
    headers = ['Item Description', 'Specification', 'Quantity', 'Unit', 'Unit Price', 'Subtotal']
    row = 7
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT
        ws.cell(row, col).border = THIN_BORDER

    items = [
        ['N12 Reinforcement Bar', 'AS/NZS 4671 - 12mm diameter', 850, 'kg', 2.45, 2082.50],
        ['N16 Reinforcement Bar', 'AS/NZS 4671 - 16mm diameter', 1200, 'kg', 2.52, 3024.00],
        ['N20 Reinforcement Bar', 'AS/NZS 4671 - 20mm diameter', 680, 'kg', 2.65, 1802.00],
        ['Steel Mesh SL82', '6.0m x 2.4m sheets', 45, 'sheets', 68.50, 3082.50],
        ['Steel Mesh SL92', '6.0m x 2.4m sheets', 30, 'sheets', 78.20, 2346.00],
        ['Bar Ties & Accessories', 'Wire ties, spacers, chairs', 1, 'lot', 450.00, 450.00],
        ['Delivery & Crane Unload', 'To 123 Sunset Blvd', 1, 'service', 850.00, 850.00],
    ]

    row = 8
    for item in items:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value).border = THIN_BORDER
            if col in [5, 6]:
                ws.cell(row, col).number_format = '$#,##0.00'
        row += 1

    # Totals
    total_row = row + 1
    ws.cell(total_row, 5, "SUBTOTAL:").font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F8:F{row-1})").number_format = '$#,##0.00'
    ws.cell(total_row, 6).font = TOTAL_FONT

    ws.cell(total_row + 1, 5, "GST (10%):").font = TOTAL_FONT
    ws.cell(total_row + 1, 6, f"=F{total_row}*0.1").number_format = '$#,##0.00'
    ws.cell(total_row + 1, 6).font = TOTAL_FONT

    ws.cell(total_row + 2, 5, "TOTAL:").font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 6, f"=F{total_row}+F{total_row+1}").number_format = '$#,##0.00'
    ws.cell(total_row + 2, 6).font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 6).fill = TOTAL_FILL

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "06_PURCHASE_ORDERS_INVOICES", "Steel_Supplier_PO_2024_001.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_timber_supplier_invoice():
    """06_PURCHASE_ORDERS_INVOICES/Timber_Supplier_Invoice_2024_078.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws['A1'] = "TAX INVOICE"
    ws['A1'].font = Font(bold=True, size=16, color="8B4513")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Timberland Frames Pty Ltd"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = "ABN: 12 345 678 901"
    ws['A4'] = "Invoice: INV-2024-078"
    ws['A5'] = "Date: 2024-08-05"
    ws['A6'] = "Customer: Sunset Blvd Development"

    headers = ['Item Description', 'Specification', 'Quantity', 'Unit', 'Unit Price', 'Total']
    row = 8
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT
        ws.cell(row, col).border = THIN_BORDER

    items = [
        ['F5 Treated Pine Studs', '90x45mm H3 LOSP - 2.4m', 280, 'ea', 12.85, 3598.00],
        ['F7 Treated Pine Plates', '90x45mm H3 LOSP - 3.6m', 95, 'ea', 18.50, 1757.50],
        ['LVL Beams', '240x45mm - 6.0m lengths', 18, 'ea', 185.00, 3330.00],
        ['Roof Trusses', 'Engineered trusses - 8.5m span', 24, 'ea', 425.00, 10200.00],
        ['Structural Plywood', 'F17 Structural - 2400x1200x17mm', 65, 'sheets', 78.50, 5102.50],
        ['Noggins & Blocking', 'Various sizes', 1, 'lot', 850.00, 850.00],
        ['Fasteners & Brackets', 'Nails, screws, joist hangers', 1, 'lot', 1250.00, 1250.00],
        ['Delivery', 'Crane truck delivery', 1, 'service', 650.00, 650.00],
    ]

    row = 9
    for item in items:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value).border = THIN_BORDER
            if col in [5, 6]:
                ws.cell(row, col).number_format = '$#,##0.00'
        row += 1

    total_row = row + 1
    ws.cell(total_row, 5, "SUBTOTAL:").font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F9:F{row-1})").number_format = '$#,##0.00'
    ws.cell(total_row, 6).font = TOTAL_FONT

    ws.cell(total_row + 1, 5, "GST (10%):").font = TOTAL_FONT
    ws.cell(total_row + 1, 6, f"=F{total_row}*0.1").number_format = '$#,##0.00'

    ws.cell(total_row + 2, 5, "TOTAL DUE:").font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 6, f"=F{total_row}+F{total_row+1}").number_format = '$#,##0.00'
    ws.cell(total_row + 2, 6).font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 6).fill = TOTAL_FILL

    ws.cell(total_row + 4, 1, "Payment Terms: 30 days from invoice date")
    ws.cell(total_row + 5, 1, "Payment Method: Direct Deposit to BSB 062-000 Acc 1234-5678")

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "06_PURCHASE_ORDERS_INVOICES", "Timber_Supplier_Invoice_2024_078.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_plumbing_materials_po():
    """06_PURCHASE_ORDERS_INVOICES/Plumbing_Materials_PO_2024_012.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    ws['A1'] = "PURCHASE ORDER - PLUMBING MATERIALS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = "PO Number: PO-2024-012"
    ws['A3'] = "Date: 2024-08-20"
    ws['A4'] = "Supplier: Reece Plumbing Supplies"
    ws['A5'] = "Delivery: 2024-08-30"

    headers = ['Description', 'Part Number', 'Qty', 'Unit Price', 'Total']
    row = 7
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT

    items = [
        ['PEX Piping 20mm', 'PEX-20-100', 100, 4.50, 450.00],
        ['PEX Piping 25mm', 'PEX-25-50', 50, 6.80, 340.00],
        ['Copper Pipe 15mm', 'CU-15-25', 25, 18.50, 462.50],
        ['PVC Drainage 90mm', 'PVC-90-6M', 12, 32.00, 384.00],
        ['Toilet Suite - Caroma', 'CAR-TS-001', 3, 485.00, 1455.00],
        ['Bathroom Vanity Tap Set', 'VAN-TAP-CH', 3, 165.00, 495.00],
        ['Kitchen Sink Mixer', 'KIT-MIX-SS', 1, 285.00, 285.00],
        ['Hot Water Service 250L', 'HWS-250-RHEEM', 1, 1850.00, 1850.00],
        ['Pipe Fittings & Valves', 'FITTINGS-LOT', 1, 650.00, 650.00],
    ]

    row = 8
    for item in items:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value).border = THIN_BORDER
            if col in [4, 5]:
                ws.cell(row, col).number_format = '$#,##0.00'
        row += 1

    total_row = row + 1
    ws.cell(total_row, 4, "TOTAL:").font = TOTAL_FONT
    ws.cell(total_row, 5, f"=SUM(E8:E{row-1})").number_format = '$#,##0.00'
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 5).fill = TOTAL_FILL

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "06_PURCHASE_ORDERS_INVOICES", "Plumbing_Materials_PO_2024_012.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_electrical_supplies_invoice():
    """06_PURCHASE_ORDERS_INVOICES/Electrical_Supplies_Invoice_2024_156.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws['A1'] = "TAX INVOICE - ELECTRICAL SUPPLIES"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Lawrence & Hanson Electrical"
    ws['A3'] = "Invoice: INV-2024-156 | Date: 2024-09-05"

    headers = ['Description', 'Code', 'Qty', 'Unit Price', 'Total']
    row = 5
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT

    items = [
        ['Distribution Board 24-way', 'DB-24W', 1, 485.00, 485.00],
        ['Circuit Breakers 20A', 'CB-20A', 12, 28.50, 342.00],
        ['RCD Safety Switch 40A', 'RCD-40A', 4, 95.00, 380.00],
        ['Cable 2.5mm TPS', 'TPS-2.5-100M', 200, 3.80, 760.00],
        ['Cable 6mm TPS', 'TPS-6-50M', 50, 8.50, 425.00],
        ['LED Downlights', 'LED-DL-90', 24, 32.00, 768.00],
        ['Power Points Double', 'GPO-2G', 18, 12.50, 225.00],
        ['Light Switches', 'SW-1G', 15, 8.80, 132.00],
        ['Conduit & Accessories', 'COND-LOT', 1, 450.00, 450.00],
    ]

    row = 6
    for item in items:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value).border = THIN_BORDER
            if col in [4, 5]:
                ws.cell(row, col).number_format = '$#,##0.00'
        row += 1

    total_row = row + 1
    ws.cell(total_row, 4, "SUBTOTAL:").font = TOTAL_FONT
    ws.cell(total_row, 5, f"=SUM(E6:E{row-1})").number_format = '$#,##0.00'

    ws.cell(total_row + 1, 4, "GST:").font = TOTAL_FONT
    ws.cell(total_row + 1, 5, f"=E{total_row}*0.1").number_format = '$#,##0.00'

    ws.cell(total_row + 2, 4, "TOTAL:").font = TOTAL_FONT
    ws.cell(total_row + 2, 5, f"=E{total_row}+E{total_row+1}").number_format = '$#,##0.00'
    ws.cell(total_row + 2, 5).fill = TOTAL_FILL

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "06_PURCHASE_ORDERS_INVOICES", "Electrical_Supplies_Invoice_2024_156.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_window_door_invoice():
    """06_PURCHASE_ORDERS_INVOICES/Window_Door_Invoice_2024_089.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws['A1'] = "INVOICE - WINDOWS & DOORS"
    ws['A1'].font = Font(bold=True, size=14, color="0066CC")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Premium Windows & Doors Pty Ltd"
    ws['A3'] = "Invoice: WD-2024-089 | Date: 2024-08-25"

    headers = ['Item', 'Specification', 'Qty', 'Unit Price', 'Install', 'Total']
    row = 5
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT

    items = [
        ['Front Entry Door', 'Pivot door 2400x1200 timber', 1, 3850.00, 450.00, 4300.00],
        ['Sliding Door 3-panel', 'Alum 3000x2100 Low-E glass', 2, 2850.00, 380.00, 6460.00],
        ['Awning Window 1200x900', 'Alum powder coat with flyscreens', 8, 485.00, 95.00, 4640.00],
        ['Fixed Window 1800x1200', 'Alum with Low-E glass', 4, 680.00, 120.00, 3200.00],
        ['Bedroom Doors', 'Hollow core 2040x820', 5, 285.00, 85.00, 1850.00],
        ['Bathroom Doors', 'Solid core 2040x720', 3, 320.00, 85.00, 1215.00],
        ['Hardware & Locks', 'Handles, hinges, locks', 1, 1250.00, 0, 1250.00],
    ]

    row = 6
    for item in items:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value).border = THIN_BORDER
            if col in [4, 5, 6]:
                ws.cell(row, col).number_format = '$#,##0.00'
        row += 1

    total_row = row + 1
    ws.cell(total_row, 5, "SUBTOTAL:").font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F6:F{row-1})").number_format = '$#,##0.00'

    ws.cell(total_row + 1, 5, "GST:").font = TOTAL_FONT
    ws.cell(total_row + 1, 6, f"=F{total_row}*0.1").number_format = '$#,##0.00'

    ws.cell(total_row + 2, 5, "TOTAL DUE:").font = TOTAL_FONT
    ws.cell(total_row + 2, 6, f"=F{total_row}+F{total_row+1}").number_format = '$#,##0.00'
    ws.cell(total_row + 2, 6).fill = TOTAL_FILL

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "06_PURCHASE_ORDERS_INVOICES", "Window_Door_Invoice_2024_089.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 07_SUBCONTRACTORS (2 files)
# ============================================================================

def create_subcontractor_agreements():
    """07_SUBCONTRACTORS/Subcontractor_Agreements_Register.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subcontractor Register"

    headers = ['Subcontractor Name', 'Trade', 'ABN', 'Contact Person', 'Phone',
               'Contract Value', 'Start Date', 'Est Completion', 'Insurance Expiry', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Metro Demolition', 'Demolition', '45 678 901 234', 'John Smith', '0412 345 678', 15000, '2024-06-20', '2024-07-05', '2025-06-01', 'COMPLETE'],
        ['BetaMix Concrete', 'Concrete Supply', '56 789 012 345', 'Sarah Lee', '0423 456 789', 45000, '2024-07-15', '2024-08-10', '2025-12-31', 'COMPLETE'],
        ['Timberland Frames', 'Framing', '67 890 123 456', 'Mike Brown', '0434 567 890', 66000, '2024-07-28', '2024-08-25', '2025-07-15', 'COMPLETE'],
        ['Metro Roofing', 'Roofing', '78 901 234 567', 'David Wilson', '0445 678 901', 28500, '2024-08-20', '2024-09-10', '2025-08-30', 'IN PROGRESS'],
        ['ProPlumb Services', 'Plumbing', '89 012 345 678', 'Lisa Chen', '0456 789 012', 35000, '2024-08-25', '2024-09-20', '2025-09-01', 'IN PROGRESS'],
        ['Spark Electrical', 'Electrical', '90 123 456 789', 'Tom Anderson', '0467 890 123', 32000, '2024-09-01', '2024-10-01', '2025-10-15', 'IN PROGRESS'],
        ['AirFlow HVAC', 'Air Conditioning', '01 234 567 890', 'Emma Taylor', '0478 901 234', 18500, '2024-09-15', '2024-10-05', '2025-11-20', 'SCHEDULED'],
        ['Premium Tiling', 'Tiling', '12 345 678 901', 'Mark Roberts', '0489 012 345', 22000, '2024-09-20', '2024-10-15', '2025-09-30', 'SCHEDULED'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency column
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 6).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL CONTRACT VALUE:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 6).font = TOTAL_FONT
    ws.cell(total_row, 6).fill = TOTAL_FILL
    ws.cell(total_row, 6).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 10)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "07_SUBCONTRACTORS", "Subcontractor_Agreements_Register.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_subcontractor_payment_schedule():
    """07_SUBCONTRACTORS/Subcontractor_Payment_Schedule.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Schedule"

    headers = ['Subcontractor', 'Trade', 'Invoice Number', 'Invoice Date', 'Description',
               'Amount (Ex GST)', 'GST', 'Total', 'Due Date', 'Payment Date', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Metro Demolition', 'Demolition', 'MD-2024-001', '2024-07-05', 'Full demolition complete', 13636.36, 1363.64, 15000.00, '2024-07-19', '2024-07-18', 'PAID'],
        ['BetaMix Concrete', 'Concrete', 'BM-2024-078', '2024-07-25', 'Slab pour - 45m³', 36818.18, 3681.82, 40500.00, '2024-08-08', '2024-08-05', 'PAID'],
        ['BetaMix Concrete', 'Concrete', 'BM-2024-089', '2024-08-05', 'Additional concrete 12m³', 4090.91, 409.09, 4500.00, '2024-08-19', '2024-08-20', 'PAID'],
        ['Timberland Frames', 'Framing', 'TF-2024-156', '2024-08-25', 'Framing complete', 60000.00, 6000.00, 66000.00, '2024-09-08', '2024-09-05', 'PAID'],
        ['Metro Roofing', 'Roofing', 'MR-2024-034', '2024-09-10', '50% progress payment', 12954.55, 1295.45, 14250.00, '2024-09-24', 'PENDING', 'INVOICED'],
        ['ProPlumb Services', 'Plumbing', 'PP-2024-189', '2024-09-15', 'Rough-in complete', 15909.09, 1590.91, 17500.00, '2024-09-29', 'PENDING', 'INVOICED'],
        ['Spark Electrical', 'Electrical', 'SE-2024-267', '2024-09-18', 'First fix electrical', 14545.45, 1454.55, 16000.00, '2024-10-02', 'PENDING', 'INVOICED'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [6, 7, 8]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL PAYMENTS:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 7, f"=SUM(G2:G{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H2:H{ws.max_row})")
    for col in [6, 7, 8]:
        ws.cell(total_row, col).font = TOTAL_FONT
        ws.cell(total_row, col).fill = TOTAL_FILL
        ws.cell(total_row, col).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 11)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "07_SUBCONTRACTORS", "Subcontractor_Payment_Schedule.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("RECREATING EXCEL SPREADSHEETS - PART 2")
    print("=" * 80)

    # 06_PURCHASE_ORDERS_INVOICES (5 files)
    print("\n[06_PURCHASE_ORDERS_INVOICES]")
    create_steel_supplier_po()
    create_timber_supplier_invoice()
    create_plumbing_materials_po()
    create_electrical_supplies_invoice()
    create_window_door_invoice()

    # 07_SUBCONTRACTORS (2 files)
    print("\n[07_SUBCONTRACTORS]")
    create_subcontractor_agreements()
    create_subcontractor_payment_schedule()

    print("\n" + "=" * 80)
    print("PART 2 COMPLETE: 7 files created")
    print("=" * 80)
