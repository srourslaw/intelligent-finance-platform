"""
Phase 3: Remaining folders 10-18 + ALL missing PDFs
"""
import os, random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont

BASE_DIR = Path(__file__).parent / "dummy_data"

def create_excel(file_path, sheet_data):
    wb = openpyxl.Workbook()
    for sheet_name, data in sheet_data.items():
        ws = wb.active if sheet_name == 'Sheet' else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        for col_idx, header in enumerate(data['headers'], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for row_idx, row_data in enumerate(data['rows'], 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(file_path)
    print(f"✓ {file_path.name}")

def create_pdf(file_path, title, lines):
    c = canvas.Canvas(str(file_path), pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1*inch, height - 1*inch, title)
    c.setFont("Helvetica", 10)
    y = height - 1.5*inch
    for line in lines:
        c.drawString(1*inch, y, str(line))
        y -= 0.25*inch
        if y < 1*inch: c.showPage(); y = height - 1*inch
    c.save()
    print(f"✓ {file_path.name}")

def create_image(file_path, text):
    img = Image.new('RGB', (800, 600), color=(220, 220, 220))
    draw = ImageDraw.Draw(img)
    try: font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 36)
    except: font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), text, font=font)
    w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(((800-w)//2, (600-h)//2), text, fill=(60, 60, 60), font=font)
    img.save(file_path)
    print(f"✓ {file_path.name}")

print("\\n" + "="*60)
print("PHASE 3: FOLDERS 10-18 + ALL PDFs")
print("="*60)

# 10_VARIATIONS_CHANGES
print("\\n[10_VARIATIONS_CHANGES]")
folder = BASE_DIR / "10_VARIATIONS_CHANGES"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Variation_Order_Register.xlsx", {
    "Variations": {
        "headers": ["VO#", "Date", "Description", "Cost", "Client Price", "Status", "Invoiced"],
        "rows": [
            ["VO-001", "2024-08-15", "Upgrade kitchen to premium", 4200, 6800, "Approved", "YES"],
            ["VO-002", "2024-09-01", "Add 2 extra windows", 1800, 3200, "Approved", "YES"],
            ["VO-003", "2024-09-10", "Upgrade bathroom tiles", 2100, 3600, "Approved", "NO"],
            ["VO-004", "2024-09-20", "Add outdoor patio", 8500, 12000, "Pending", "NO"],
        ]
    }
})

create_excel(folder / "VO_Impact_on_Budget.xlsx", {
    "Impact": {
        "headers": ["Variation", "Cost", "Price", "Margin", "Budget Impact"],
        "rows": [
            ["VO-001 Kitchen", 4200, 6800, 2600, "Positive"],
            ["VO-002 Windows", 1800, 3200, 1400, "Positive"],
            ["VO-003 Tiles", 2100, 3600, 1500, "Positive"],
            ["VO-004 Patio", 8500, 12000, 3500, "Pending"],
            ["TOTAL", 16600, 25600, 9000, ""],
        ]
    }
})

# 11_CLIENT_BILLING
print("\\n[11_CLIENT_BILLING]")
folder = BASE_DIR / "11_CLIENT_BILLING"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Payment_Schedule.xlsx", {
    "Schedule": {
        "headers": ["Milestone", "% Complete", "Amount", "Due Date", "Status"],
        "rows": [
            ["Deposit", "0%", 65000, "2024-06-01", "PAID"],
            ["Slab Complete", "30%", 130000, "2024-08-25", "PAID"],
            ["Frame & Roof", "50%", 130000, "2024-09-30", "PAID"],
            ["Lockup", "70%", 90000, "2024-10-31", "Pending"],
            ["Fixing Complete", "90%", 100000, "2024-11-30", "Pending"],
            ["Practical Completion", "100%", 135000, "2024-12-20", "Pending"],
        ]
    }
})

create_excel(folder / "Outstanding_Client_Invoices.xlsx", {
    "Outstanding": {
        "headers": ["Invoice #", "Milestone", "Amount", "Due Date", "Days Outstanding"],
        "rows": [
            ["INV-004", "Lockup", 90000, "2024-10-31", "Not Due"],
            ["INV-005", "Fixing", 100000, "2024-11-30", "Not Due"],
            ["INV-006", "Final", 135000, "2024-12-20", "Not Due"],
        ]
    }
})

# 12_BUDGET_TRACKING
print("\\n[12_BUDGET_TRACKING]")
folder = BASE_DIR / "12_BUDGET_TRACKING"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Project_Budget_v1.xlsx", {
    "Budget v1": {
        "headers": ["Category", "Original Budget"],
        "rows": [
            ["Land", 260000],
            ["Design", 18000],
            ["Permits", 10000],
            ["Site Prep", 24000],
            ["Foundation", 42000],
            ["Frame", 60000],
            ["External", 100000],
            ["Services", 65000],
            ["Finishes", 70000],
            ["Contingency", 30000],
            ["TOTAL", 679000],
        ]
    }
})

create_excel(folder / "Project_Budget_v2_updated.xlsx", {
    "Budget v2 UPDATED": {
        "headers": ["Category", "v1 Budget", "v2 Updated", "Change"],
        "rows": [
            ["Land", 260000, 265000, 5000],
            ["Design", 18000, 19500, 1500],
            ["Frame", 60000, 62000, 2000],
            ["External", 100000, 103000, 3000],
            ["Finishes", 70000, 72000, 2000],
            ["TOTAL", 679000, 695000, 16000],
        ]
    }
})

create_excel(folder / "Budget_vs_Actual.xlsx", {
    "Comparison": {
        "headers": ["Category", "Budget", "Actual", "Variance", "% Variance"],
        "rows": [
            ["Land", 265000, 265000, 0, "0%"],
            ["Design", 19500, 19740, -240, "-1.2%"],
            ["Permits", 11000, 10950, 50, "0.5%"],
            ["Site Prep", 25000, 26670, -1670, "-6.7%"],
            ["Foundation", 46000, 47090, -1090, "-2.4%"],
            ["Frame", 62000, 61900, 100, "0.2%"],
            ["External", 103000, 79650, 23350, "22.7%"],
        ]
    }
})

create_excel(folder / "Cost_Breakdown_by_Phase.xlsx", {
    "By Phase": {
        "headers": ["Phase", "Budget", "Actual", "% Complete"],
        "rows": [
            ["Pre-Construction", 295000, 295690, "100%"],
            ["Groundworks", 71000, 73760, "100%"],
            ["Structure", 108000, 108990, "100%"],
            ["Envelope", 103000, 79650, "75%"],
            ["Services", 68000, 42800, "60%"],
            ["Finishes", 72000, 12340, "15%"],
        ]
    }
})

create_excel(folder / "Weekly_Cost_Report.xlsx", {
    "Week Ending 2024-09-27": {
        "headers": ["Item", "This Week", "Month to Date", "Project to Date"],
        "rows": [
            ["Labour", 4200, 18900, 72330],
            ["Materials", 8900, 32100, 156780],
            ["Subcontractors", 12000, 48500, 234600],
            ["TOTAL", 25100, 99500, 463710],
        ]
    }
})

create_excel(folder / "Monthly_Financial_Summary_Aug.xlsx", {
    "August 2024": {
        "headers": ["Category", "Budget", "Spent", "Remaining"],
        "rows": [
            ["Labour", 9680, 9680, 0],
            ["Materials", 45000, 42890, 2110],
            ["Subcontractors", 80000, 78950, 1050],
            ["TOTAL AUGUST", 134680, 131520, 3160],
        ]
    }
})

create_excel(folder / "Monthly_Financial_Summary_Sept.xlsx", {
    "September 2024": {
        "headers": ["Category", "Budget", "Spent", "Remaining"],
        "rows": [
            ["Labour", 10120, 10120, 0],
            ["Materials", 52000, 48900, 3100],
            ["Subcontractors", 95000, 89600, 5400],
            ["TOTAL SEPTEMBER", 157120, 148620, 8500],
        ]
    }
})

create_excel(folder / "Cashflow_Forecast.xlsx", {
    "Forecast": {
        "headers": ["Month", "Income", "Expenses", "Net Cashflow", "Cumulative"],
        "rows": [
            ["Oct 2024", 90000, 125000, -35000, -35000],
            ["Nov 2024", 100000, 95000, 5000, -30000],
            ["Dec 2024", 135000, 45000, 90000, 60000],
        ]
    }
})

create_excel(folder / "Cashflow_Actual.xlsx", {
    "Actual": {
        "headers": ["Month", "Income", "Expenses", "Net Cashflow", "Cumulative"],
        "rows": [
            ["Jun 2024", 65000, 268000, -203000, -203000],
            ["Jul 2024", 0, 45000, -45000, -248000],
            ["Aug 2024", 130000, 131520, -1520, -249520],
            ["Sept 2024", 130000, 148620, -18620, -268140],
        ]
    }
})

create_excel(folder / "Profit_Margin_Calculation.xlsx", {
    "Profit": {
        "headers": ["Item", "Amount"],
        "rows": [
            ["Contract Value", 650000],
            ["Total Costs (Forecast)", 829820],
            ["Projected Profit/Loss", -179820],
            ["Margin %", "-27.7%"],
            ["", ""],
            ["Status", "OVER BUDGET"],
        ]
    }
})

# 13_SCHEDULE_TIMELINE
print("\\n[13_SCHEDULE_TIMELINE]")
folder = BASE_DIR / "13_SCHEDULE_TIMELINE"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Project_Schedule_Gantt.xlsx", {
    "Schedule": {
        "headers": ["Task", "Start", "End", "Duration", "Status"],
        "rows": [
            ["Site Prep", "2024-06-01", "2024-08-01", "60 days", "Complete"],
            ["Foundation", "2024-08-02", "2024-08-25", "23 days", "Complete"],
            ["Frame & Roof", "2024-08-26", "2024-09-25", "30 days", "Complete"],
            ["External Walls", "2024-09-10", "2024-10-20", "40 days", "In Progress"],
            ["Services Rough-in", "2024-09-25", "2024-10-30", "35 days", "In Progress"],
            ["Plasterboard", "2024-10-20", "2024-11-10", "21 days", "Not Started"],
            ["Fixing & Finishes", "2024-11-01", "2024-12-10", "40 days", "Not Started"],
            ["Final Inspection", "2024-12-15", "2024-12-20", "5 days", "Not Started"],
        ]
    }
})

create_excel(folder / "Milestone_Tracker.xlsx", {
    "Milestones": {
        "headers": ["Milestone", "Planned Date", "Actual Date", "Variance (days)", "Status"],
        "rows": [
            ["Slab Down", "2024-08-20", "2024-08-25", 5, "Complete"],
            ["Frame Complete", "2024-09-20", "2024-09-25", 5, "Complete"],
            ["Lockup", "2024-10-25", "", "", "In Progress"],
            ["Fixing Complete", "2024-11-25", "", "", "Not Started"],
            ["Practical Completion", "2024-12-15", "", "", "Not Started"],
        ]
    }
})

create_excel(folder / "Critical_Path_Analysis.xlsx", {
    "Critical Path": {
        "headers": ["Activity", "Duration", "Dependencies", "Float", "Critical?"],
        "rows": [
            ["Foundation", "23 days", "Site Prep", "0 days", "YES"],
            ["Frame", "30 days", "Foundation", "0 days", "YES"],
            ["Roof", "15 days", "Frame", "0 days", "YES"],
            ["External Walls", "40 days", "Frame", "5 days", "NO"],
            ["Plumbing Rough-in", "20 days", "Frame", "10 days", "NO"],
        ]
    }
})

create_excel(folder / "Delays_Log.xlsx", {
    "Delays": {
        "headers": ["Date", "Reason", "Days Lost", "Impact", "Mitigation"],
        "rows": [
            ["2024-08-05", "Heavy Rain", 2, "Foundation delayed", "Overtime"],
            ["2024-09-11", "Wind - Roof", 1, "Roof install stopped", "Extra crew"],
            ["2024-09-23", "Material shortage", 3, "Electrical delayed", "Alternative supplier"],
        ]
    }
})

# 14_COMPLIANCE_CERTIFICATES
print("\\n[14_COMPLIANCE_CERTIFICATES]")
folder = BASE_DIR / "14_COMPLIANCE_CERTIFICATES"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Compliance_Checklist.xlsx", {
    "Checklist": {
        "headers": ["Item", "Required", "Received", "Date", "Status"],
        "rows": [
            ["Building Permit", "YES", "YES", "2024-07-25", "Complete"],
            ["DA Approval", "YES", "YES", "2024-07-10", "Complete"],
            ["Waterproofing Cert", "YES", "NO", "", "Pending"],
            ["Termite Treatment", "YES", "NO", "", "Pending"],
            ["Electrical Compliance", "YES", "YES", "2024-09-28", "Complete"],
            ["Plumbing Compliance", "YES", "YES", "2024-09-30", "Complete"],
            ["Structural Warranty", "YES", "NO", "", "Pending"],
            ["Occupancy Certificate", "YES", "NO", "", "Pending"],
        ]
    }
})

# 15_DEFECTS_SNAGGING
print("\\n[15_DEFECTS_SNAGGING]")
folder = BASE_DIR / "15_DEFECTS_SNAGGING"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Snagging_Items.xlsx", {
    "Snagging": {
        "headers": ["Item", "Location", "Trade", "Priority", "Status"],
        "rows": [
            ["Paint touch-up required", "Living Room", "Painter", "Low", "Open"],
            ["Door alignment", "Bedroom 2", "Carpenter", "Medium", "Open"],
            ["Tile crack", "Bathroom 1", "Tiler", "High", "Open"],
            ["Window seal", "Kitchen", "Window installer", "Medium", "Open"],
        ]
    }
})

create_excel(folder / "Defect_Rectification_Log.xlsx", {
    "Rectification": {
        "headers": ["Defect ID", "Description", "Reported", "Fixed", "Days to Fix", "Status"],
        "rows": [
            ["D-001", "Paint touch-up", "2024-09-28", "", "", "Open"],
            ["D-002", "Door alignment", "2024-09-29", "", "", "Open"],
            ["D-003", "Tile crack", "2024-09-30", "", "", "Open"],
        ]
    }
})

# Create defect images
print("\\n[15_DEFECTS_SNAGGING - Images]")
defect_folder = folder / "Defect_Photos"
defect_folder.mkdir(parents=True, exist_ok=True)

create_image(defect_folder / "Defect_01_Paint_Touch_Up.jpg", "Paint Touch Up\\nLiving Room")
create_image(defect_folder / "Defect_02_Door_Alignment.jpg", "Door Alignment\\nBedroom 2")
create_image(defect_folder / "Defect_03_Tile_Crack.jpg", "Tile Crack\\nBathroom 1")

# 16_HANDOVER
print("\\n[16_HANDOVER]")
folder = BASE_DIR / "16_HANDOVER"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Handover_Checklist.xlsx", {
    "Checklist": {
        "headers": ["Item", "Status", "Notes"],
        "rows": [
            ["All defects rectified", "Pending", "3 items outstanding"],
            ["Occupancy Certificate", "Pending", ""],
            ["Keys provided", "Pending", ""],
            ["Warranties collected", "Pending", ""],
            ["As-built drawings", "Pending", ""],
            ["Maintenance schedule", "Pending", ""],
            ["Client walkthrough", "Pending", "To be scheduled"],
        ]
    }
})

create_excel(folder / "Keys_Register.xlsx", {
    "Keys": {
        "headers": ["Key Type", "Quantity", "Given to Client", "Date"],
        "rows": [
            ["Front Door", 3, "No", ""],
            ["Back Door", 2, "No", ""],
            ["Garage", 2, "No", ""],
            ["Mailbox", 2, "No", ""],
        ]
    }
})

# 17_CORRESPONDENCE
print("\\n[17_CORRESPONDENCE]")

# 18_MISC_RANDOM
print("\\n[18_MISC_RANDOM]")
folder = BASE_DIR / "18_MISC_RANDOM"
folder.mkdir(parents=True, exist_ok=True)

create_excel(folder / "Todo_List.xlsx", {
    "Todo": {
        "headers": ["Task", "Priority", "Due", "Status"],
        "rows": [
            ["Chase outstanding invoices", "HIGH", "ASAP", "TODO"],
            ["Book final inspection", "HIGH", "Next week", "TODO"],
            ["Order bathroom fixtures", "MEDIUM", "This week", "DONE"],
            ["Fix defects", "HIGH", "Before handover", "IN PROGRESS"],
        ]
    }
})

create_excel(folder / "Phone_Numbers_Contacts.xlsx", {
    "Contacts": {
        "headers": ["Name", "Company", "Phone", "Email", "Role"],
        "rows": [
            ["John Smith", "Self", "0412 345 678", "john@email.com", "Builder"],
            ["Mike Brown", "Spark Electric", "0423 456 789", "mike@spark.com", "Electrician"],
            ["Sarah Lee", "AquaFlow", "0434 567 890", "sarah@aqua.com", "Plumber"],
            ["Council", "Sydney Council", "02 9999 1234", "council@sydney.nsw", "Authority"],
        ]
    }
})

create_excel(folder / "Old_Version_Budget_DELETE.xlsx", {
    "OLD BUDGET": {
        "headers": ["Category", "Amount"],
        "rows": [
            ["THIS FILE IS OLD", ""],
            ["DO NOT USE", ""],
            ["", ""],
            ["Land", 250000],
            ["Construction", 400000],
            ["TOTAL", 650000],
        ]
    }
})

print("\\n" + "="*60)
print("PHASE 3 COMPLETE - EXCEL FILES DONE!")
print("="*60)
print(f"\\n✓ All Excel files for folders 10-18 created")
print(f"\\n>>> Next: Run generate_pdfs.py for ALL missing PDFs")
