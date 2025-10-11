"""
Excel Population Service

Populates Financial Model Excel spreadsheet with categorized and aggregated data.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelPopulator:
    """Service for populating Excel financial models."""

    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize Excel populator.

        Args:
            template_path: Path to Excel template file (optional)
        """
        self.template_path = template_path

    def create_financial_model(
        self,
        categorized_data: Dict[str, Any],
        output_path: str,
        project_name: str = "Construction Project"
    ) -> str:
        """
        Create populated financial model Excel file.

        Args:
            categorized_data: Dictionary with categorized transactions
            output_path: Path where to save the Excel file
            project_name: Name of the project

        Returns:
            Path to created Excel file
        """
        try:
            # Create new workbook
            wb = openpyxl.Workbook()

            # Create sheets
            self._create_summary_sheet(wb, categorized_data, project_name)
            self._create_revenue_sheet(wb, categorized_data)
            self._create_direct_costs_sheet(wb, categorized_data)
            self._create_indirect_costs_sheet(wb, categorized_data)
            self._create_transactions_sheet(wb, categorized_data)

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save workbook
            wb.save(output_path)
            logger.info(f"Financial model created: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error creating financial model: {e}")
            raise

    def _create_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        data: Dict[str, Any],
        project_name: str
    ):
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        # Header styling
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Title
        ws['A1'] = project_name
        ws['A1'].font = Font(name='Arial', size=18, bold=True)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(name='Arial', size=10, color='666666')

        # Key metrics
        row = 4
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 2

        # Calculate totals
        total_revenue = self._calculate_category_total(data, 'Revenue')
        total_direct_costs = self._calculate_category_total(data, 'Direct Costs')
        total_indirect_costs = self._calculate_category_total(data, 'Indirect Costs')
        gross_profit = total_revenue - total_direct_costs
        net_profit = gross_profit - total_indirect_costs
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Metrics
        metrics = [
            ("Total Revenue", total_revenue, '0.00'),
            ("Total Direct Costs", total_direct_costs, '0.00'),
            ("Gross Profit", gross_profit, '0.00'),
            ("Gross Margin", gross_margin, '0.00'),
            ("Total Indirect Costs", total_indirect_costs, '0.00'),
            ("Net Profit", net_profit, '0.00'),
            ("Net Margin", net_margin, '0.00'),
        ]

        for label, value, color_code in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value if '%' not in label else value / 100
            ws[f'B{row}'].number_format = '#,##0.00' if '%' not in label else '0.00%'

            # Color code profits
            if 'Profit' in label or 'Margin' in label:
                if value >= 0:
                    ws[f'B{row}'].font = Font(color='006100')
                else:
                    ws[f'B{row}'].font = Font(color='9C0006')

            row += 1

        # Data quality metrics
        row += 2
        ws[f'A{row}'] = "DATA QUALITY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 2
        summary = data.get('summary', {})
        quality_metrics = [
            ("Total Transactions", summary.get('total_transactions', 0)),
            ("Successfully Categorized", summary.get('total_transactions', 0) - summary.get('uncategorized_count', 0)),
            ("Uncategorized", summary.get('uncategorized_count', 0)),
            ("Average Confidence", summary.get('average_confidence', 0)),
            ("Low Confidence Count", summary.get('low_confidence_count', 0)),
        ]

        for label, value in quality_metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if 'Confidence' in label:
                ws[f'B{row}'].number_format = '0.00%'
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_revenue_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create revenue breakdown sheet."""
        ws = wb.create_sheet("Revenue")
        self._populate_category_sheet(ws, data, "Revenue", "Revenue Categories")

    def _create_direct_costs_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create direct costs breakdown sheet."""
        ws = wb.create_sheet("Direct Costs")
        self._populate_category_sheet(ws, data, "Direct Costs", "Direct Cost Categories")

    def _create_indirect_costs_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create indirect costs breakdown sheet."""
        ws = wb.create_sheet("Indirect Costs")
        self._populate_category_sheet(ws, data, "Indirect Costs", "Indirect Cost Categories")

    def _populate_category_sheet(
        self,
        ws,
        data: Dict[str, Any],
        category_type: str,
        title: str
    ):
        """Populate a category breakdown sheet."""
        # Header
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        ws['A1'] = title
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')

        # Column headers
        headers = ['Category', 'Transaction Count', 'Total Amount', 'Avg Confidence', '% of Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Get transactions for this category type
        transactions = data.get('categorized_transactions', [])
        categories = data.get('summary', {}).get('categories', {})

        # Filter categories by type
        relevant_categories = self._get_categories_by_type(categories, category_type)

        # Calculate total for percentage
        total_amount = sum(cat_data['total_amount'] for cat_data in relevant_categories.values())

        # Populate data
        row = 4
        for cat_name, cat_data in sorted(relevant_categories.items(), key=lambda x: x[1]['total_amount'], reverse=True):
            ws.cell(row=row, column=1).value = cat_name
            ws.cell(row=row, column=2).value = cat_data['count']
            ws.cell(row=row, column=3).value = cat_data['total_amount']
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).value = cat_data['avg_confidence']
            ws.cell(row=row, column=4).number_format = '0.00%'

            percentage = (cat_data['total_amount'] / total_amount) if total_amount > 0 else 0
            ws.cell(row=row, column=5).value = percentage
            ws.cell(row=row, column=5).number_format = '0.00%'

            row += 1

        # Total row
        if row > 4:
            ws.cell(row=row, column=1).value = "TOTAL"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3).value = total_amount
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15

    def _create_transactions_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create detailed transactions sheet."""
        ws = wb.create_sheet("Transactions")

        # Header
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        ws['A1'] = "All Transactions"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:G1')

        # Column headers
        headers = ['Date', 'Description', 'Vendor', 'Amount', 'Category', 'Confidence', 'Source File']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Populate transactions
        transactions = data.get('categorized_transactions', [])
        row = 4
        for txn in transactions:
            ws.cell(row=row, column=1).value = txn.get('date', '')
            ws.cell(row=row, column=2).value = txn.get('description', '')
            ws.cell(row=row, column=3).value = txn.get('vendor', '')
            ws.cell(row=row, column=4).value = txn.get('amount', 0)
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).value = txn.get('category', 'Uncategorized')
            ws.cell(row=row, column=6).value = txn.get('confidence', 0)
            ws.cell(row=row, column=6).number_format = '0.00%'
            ws.cell(row=row, column=7).value = txn.get('source_file', '')

            # Color code by confidence
            confidence = txn.get('confidence', 0)
            if confidence < 0.6:
                ws.cell(row=row, column=6).font = Font(color='9C0006')
            elif confidence < 0.8:
                ws.cell(row=row, column=6).font = Font(color='9C5700')
            else:
                ws.cell(row=row, column=6).font = Font(color='006100')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30

    def _calculate_category_total(self, data: Dict[str, Any], category_type: str) -> float:
        """Calculate total for a category type."""
        categories = data.get('summary', {}).get('categories', {})
        relevant_categories = self._get_categories_by_type(categories, category_type)
        return sum(cat_data['total_amount'] for cat_data in relevant_categories.values())

    def _get_categories_by_type(
        self,
        categories: Dict[str, Any],
        category_type: str
    ) -> Dict[str, Any]:
        """
        Filter categories by type (Revenue, Direct Costs, Indirect Costs).

        Args:
            categories: Dictionary of all categories
            category_type: Type to filter by

        Returns:
            Filtered dictionary of categories
        """
        # Define which categories belong to which type
        revenue_keywords = ['revenue', 'income', 'payment', 'billing', 'change order']
        direct_cost_keywords = [
            'materials', 'labor', 'equipment', 'subcontractor',
            'concrete', 'steel', 'lumber', 'carpenter', 'electrician', 'plumber'
        ]

        filtered = {}
        for cat_name, cat_data in categories.items():
            cat_lower = cat_name.lower()

            if category_type == "Revenue":
                if any(kw in cat_lower for kw in revenue_keywords):
                    filtered[cat_name] = cat_data

            elif category_type == "Direct Costs":
                if any(kw in cat_lower for kw in direct_cost_keywords):
                    filtered[cat_name] = cat_data

            elif category_type == "Indirect Costs":
                # Indirect costs are everything that's not revenue or direct costs
                is_revenue = any(kw in cat_lower for kw in revenue_keywords)
                is_direct = any(kw in cat_lower for kw in direct_cost_keywords)
                if not is_revenue and not is_direct and cat_name != 'Uncategorized':
                    filtered[cat_name] = cat_data

        return filtered


def create_excel_populator(template_path: Optional[str] = None) -> ExcelPopulator:
    """
    Factory function to create Excel populator instance.

    Args:
        template_path: Path to Excel template file

    Returns:
        ExcelPopulator instance
    """
    return ExcelPopulator(template_path)
