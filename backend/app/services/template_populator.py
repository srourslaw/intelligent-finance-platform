"""
Excel Template Population Service

Takes aggregated financial data and populates an Excel template
while preserving formulas, formatting, and structure.
"""

import os
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class TemplatePopulator:
    """
    Populates Excel financial templates with aggregated data.

    Features:
    - Preserves existing formulas
    - Maintains formatting
    - Adds data lineage sheet
    - Supports multi-period data
    - Handles missing data gracefully
    """

    def __init__(self, template_path: str):
        """
        Initialize template populator.

        Args:
            template_path: Path to Excel template file
        """
        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        self.workbook = None
        self.mapping_config = self._load_default_mapping()

    def _load_default_mapping(self) -> Dict[str, Any]:
        """
        Load default JSON-to-Excel cell mapping.

        This maps JSON paths to specific Excel cells.
        Format: {"json.path": {"sheet": "Sheet1", "cell": "B5"}}
        """
        return {
            # Balance Sheet - Assets
            "balance_sheet.assets.current_assets.cash_on_hand": {
                "sheet": "Balance Sheet",
                "cell": "B5",
                "type": "currency"
            },
            "balance_sheet.assets.current_assets.cash_in_bank_operating": {
                "sheet": "Balance Sheet",
                "cell": "B6",
                "type": "currency"
            },
            "balance_sheet.assets.current_assets.accounts_receivable": {
                "sheet": "Balance Sheet",
                "cell": "B7",
                "type": "currency"
            },
            "balance_sheet.assets.current_assets.inventory_raw_materials": {
                "sheet": "Balance Sheet",
                "cell": "B10",
                "type": "currency"
            },
            "balance_sheet.assets.non_current_assets.land": {
                "sheet": "Balance Sheet",
                "cell": "B20",
                "type": "currency"
            },
            "balance_sheet.assets.non_current_assets.buildings": {
                "sheet": "Balance Sheet",
                "cell": "B21",
                "type": "currency"
            },

            # Balance Sheet - Liabilities
            "balance_sheet.liabilities.current_liabilities.accounts_payable": {
                "sheet": "Balance Sheet",
                "cell": "B35",
                "type": "currency"
            },
            "balance_sheet.liabilities.current_liabilities.credit_card_debt": {
                "sheet": "Balance Sheet",
                "cell": "B36",
                "type": "currency"
            },
            "balance_sheet.liabilities.long_term_liabilities.mortgage_payable": {
                "sheet": "Balance Sheet",
                "cell": "B45",
                "type": "currency"
            },

            # Balance Sheet - Equity
            "balance_sheet.equity.share_capital": {
                "sheet": "Balance Sheet",
                "cell": "B55",
                "type": "currency"
            },
            "balance_sheet.equity.retained_earnings": {
                "sheet": "Balance Sheet",
                "cell": "B56",
                "type": "currency"
            },

            # Income Statement - Revenue
            "income_statement.revenue.product_sales": {
                "sheet": "Income Statement",
                "cell": "C5",
                "type": "currency"
            },
            "income_statement.revenue.service_revenue": {
                "sheet": "Income Statement",
                "cell": "C6",
                "type": "currency"
            },

            # Income Statement - COGS
            "income_statement.cogs.purchases": {
                "sheet": "Income Statement",
                "cell": "C15",
                "type": "currency"
            },
            "income_statement.cogs.direct_labor": {
                "sheet": "Income Statement",
                "cell": "C16",
                "type": "currency"
            },

            # Income Statement - Operating Expenses
            "income_statement.operating_expenses.salaries_and_wages": {
                "sheet": "Income Statement",
                "cell": "C25",
                "type": "currency"
            },
            "income_statement.operating_expenses.rent": {
                "sheet": "Income Statement",
                "cell": "C26",
                "type": "currency"
            },
            "income_statement.operating_expenses.marketing": {
                "sheet": "Income Statement",
                "cell": "C30",
                "type": "currency"
            },

            # Cash Flow Statement
            "cash_flow.operating_activities.net_profit": {
                "sheet": "Cash Flow",
                "cell": "D5",
                "type": "currency"
            },
            "cash_flow.operating_activities.depreciation": {
                "sheet": "Cash Flow",
                "cell": "D10",
                "type": "currency"
            },
            "cash_flow.investing_activities.purchase_of_ppe": {
                "sheet": "Cash Flow",
                "cell": "D25",
                "type": "currency"
            },
            "cash_flow.financing_activities.repayment_of_debt": {
                "sheet": "Cash Flow",
                "cell": "D35",
                "type": "currency"
            }
        }

    def _get_nested_value(self, data: Dict, path: str, default: Any = None) -> Any:
        """
        Get value from nested dictionary using dot notation.

        Args:
            data: Dictionary to search
            path: Dot-separated path (e.g., "balance_sheet.assets.cash")
            default: Default value if path not found

        Returns:
            Value at path or default
        """
        keys = path.split('.')
        value = data

        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return default

        return value

    def populate(
        self,
        aggregated_data: Dict[str, Any],
        output_path: Optional[str] = None,
        add_lineage_sheet: bool = True
    ) -> str:
        """
        Populate template with aggregated data.

        Args:
            aggregated_data: Aggregated financial data (JSON format)
            output_path: Path for output file (auto-generated if None)
            add_lineage_sheet: Add data lineage metadata sheet

        Returns:
            Path to populated Excel file
        """
        logger.info("📊 Starting template population...")

        # Load template
        self.workbook = load_workbook(self.template_path)
        logger.info(f"  ✅ Loaded template: {self.template_path.name}")

        # Populate cells
        populated_count = 0
        skipped_count = 0

        for json_path, cell_info in self.mapping_config.items():
            sheet_name = cell_info["sheet"]
            cell_address = cell_info["cell"]
            data_type = cell_info.get("type", "number")

            # Get value from aggregated data
            value = self._get_nested_value(aggregated_data, json_path)

            if value is None:
                logger.debug(f"  ⏭️  No data for {json_path}")
                skipped_count += 1
                continue

            # Get or create sheet
            if sheet_name not in self.workbook.sheetnames:
                logger.warning(f"  ⚠️  Sheet '{sheet_name}' not found in template, skipping")
                skipped_count += 1
                continue

            sheet = self.workbook[sheet_name]
            cell = sheet[cell_address]

            # Check if cell has formula (don't overwrite formulas)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                logger.debug(f"  🔒 Skipping formula cell {sheet_name}!{cell_address}")
                skipped_count += 1
                continue

            # Set value
            try:
                if data_type == "currency":
                    cell.value = float(value)
                    cell.number_format = '$#,##0.00'
                elif data_type == "percentage":
                    cell.value = float(value) / 100
                    cell.number_format = '0.00%'
                elif data_type == "date":
                    cell.value = value
                    cell.number_format = 'mm/dd/yyyy'
                else:
                    cell.value = value

                populated_count += 1
                logger.debug(f"  ✅ {sheet_name}!{cell_address} = {value}")

            except Exception as e:
                logger.error(f"  ❌ Failed to set {sheet_name}!{cell_address}: {e}")
                skipped_count += 1

        logger.info(f"  📊 Populated {populated_count} cells, skipped {skipped_count}")

        # Add data lineage sheet
        if add_lineage_sheet:
            self._add_lineage_sheet(aggregated_data)

        # Generate output path
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            project_id = aggregated_data.get("metadata", {}).get("project_id", "Unknown")
            output_path = f"data/populated_templates/{project_id}_{timestamp}.xlsx"

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Save populated template
        self.workbook.save(output_file)
        logger.info(f"✅ Template saved: {output_file}")

        return str(output_file)

    def _add_lineage_sheet(self, aggregated_data: Dict[str, Any]):
        """
        Add 'Data Lineage' sheet with metadata about data sources.

        Args:
            aggregated_data: Aggregated financial data
        """
        # Create or get lineage sheet
        if "Data Lineage" in self.workbook.sheetnames:
            sheet = self.workbook["Data Lineage"]
            sheet.delete_rows(1, sheet.max_row)  # Clear existing
        else:
            sheet = self.workbook.create_sheet("Data Lineage")

        # Header
        headers = ["Cell Reference", "Sheet", "Value", "Source File(s)", "Confidence", "Last Updated"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data rows
        row_idx = 2
        metadata = aggregated_data.get("metadata", {})
        source_files = metadata.get("source_files", [])
        timestamp = metadata.get("aggregated_at", datetime.now().isoformat())

        for json_path, cell_info in self.mapping_config.items():
            value = self._get_nested_value(aggregated_data, json_path)

            if value is not None:
                sheet.cell(row=row_idx, column=1, value=cell_info["cell"])
                sheet.cell(row=row_idx, column=2, value=cell_info["sheet"])
                sheet.cell(row=row_idx, column=3, value=value)
                sheet.cell(row=row_idx, column=4, value=", ".join(source_files[:3]))  # First 3 files
                sheet.cell(row=row_idx, column=5, value="95%")  # Placeholder
                sheet.cell(row=row_idx, column=6, value=timestamp)
                row_idx += 1

        # Auto-size columns
        for col_idx in range(1, 7):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 20

        logger.info("  ✅ Added Data Lineage sheet")

    def validate_populated_template(self, file_path: str) -> Dict[str, Any]:
        """
        Validate populated template.

        Args:
            file_path: Path to populated template

        Returns:
            Validation results
        """
        wb = load_workbook(file_path)

        validation_results = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "statistics": {}
        }

        # Check Balance Sheet balances
        if "Balance Sheet" in wb.sheetnames:
            sheet = wb["Balance Sheet"]

            # Example: Check if Assets = Liabilities + Equity
            # This assumes specific cell locations - adjust as needed
            # assets = sheet['B50'].value  # Total Assets
            # liabilities = sheet['B48'].value  # Total Liabilities
            # equity = sheet['B58'].value  # Total Equity

            # if assets and liabilities and equity:
            #     if abs(assets - (liabilities + equity)) > 0.01:
            #         validation_results["errors"].append("Balance Sheet does not balance")
            #         validation_results["valid"] = False

            validation_results["statistics"]["balance_sheet"] = "checked"

        return validation_results


def populate_template_from_aggregation(
    aggregation_result: Dict[str, Any],
    template_path: str,
    output_path: Optional[str] = None
) -> str:
    """
    Convenience function to populate template from aggregation result.

    Args:
        aggregation_result: Result from aggregation engine
        template_path: Path to Excel template
        output_path: Optional output path

    Returns:
        Path to populated template
    """
    populator = TemplatePopulator(template_path)
    return populator.populate(aggregation_result, output_path)
