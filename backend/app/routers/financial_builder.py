"""
Financial Builder API Endpoints
Handles full pipeline: extraction → categorization → aggregation → Excel population
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Dict, Any, Optional
from pydantic import BaseModel
from datetime import datetime

from app.database import get_db
from app.services.template_parser import parse_template, TemplateParser
from app.services.file_extractor import start_extraction, get_extraction_status, FileExtractor
from app.services.ai_categorizer import create_categorizer
from app.services.excel_populator import create_excel_populator
from app.models.extraction import ExtractionJob, JobStatus, ExtractedData
import json
import os
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/financial-builder", tags=["financial-builder"])


class ParseTemplateResponse(BaseModel):
    success: bool
    project_id: str
    total_categories: int
    message: str


class StartExtractionResponse(BaseModel):
    success: bool
    job_id: str
    project_id: str
    total_files: int
    message: str


class JobStatusResponse(BaseModel):
    job_id: str
    project_id: str
    status: str
    total_files: int
    processed_files: int
    failed_files: int
    progress_percent: float
    error_message: Optional[str]
    metadata: Optional[str]  # JSON string with pipeline results


@router.post("/{project_id}/parse-template", response_model=ParseTemplateResponse)
async def parse_project_template(project_id: str):
    """
    Parse MASTER FINANCIAL STATEMENT TEMPLATE.md into structured JSON dictionary.

    Args:
        project_id: Project identifier

    Returns:
        Parsed template summary
    """
    try:
        parser = TemplateParser(project_id)
        template_dict = parser.parse()

        # Save to JSON file
        output_path = parser.save_to_json()

        return ParseTemplateResponse(
            success=True,
            project_id=project_id,
            total_categories=template_dict['metadata']['total_categories'],
            message=f"Template parsed successfully. Dictionary saved to {output_path}"
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Template parsing failed: {str(e)}")


@router.post("/{project_id}/extract-all", response_model=StartExtractionResponse)
async def extract_all_files(
    project_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Start bulk extraction of all files in project data folder.

    This endpoint:
    1. Scans project data folder for all PDFs, Excel, CSV files
    2. Creates extraction job in database
    3. Processes files in background (PDFs with MinerU, Excel with pandas)
    4. Stores extracted data in database

    Args:
        project_id: Project identifier
        background_tasks: FastAPI background tasks
        db: Database session

    Returns:
        Job ID and initial status
    """
    try:
        # Create extractor
        extractor = FileExtractor(project_id, db)

        # Scan files first
        files = extractor.scan_files()
        if not files:
            raise HTTPException(
                status_code=404,
                detail=f"No supported files found in project {project_id}/data folder"
            )

        # Create job
        job_id = extractor.create_extraction_job()

        # Run extraction in background
        background_tasks.add_task(extractor.run_extraction, job_id)

        return StartExtractionResponse(
            success=True,
            job_id=job_id,
            project_id=project_id,
            total_files=len(files),
            message=f"Extraction started for {len(files)} files. Use job_id to track progress."
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start extraction: {str(e)}")


@router.get("/jobs/{job_id}/status", response_model=JobStatusResponse)
async def get_job_status(job_id: str, db: Session = Depends(get_db)):
    """
    Get current status of an extraction job.

    Args:
        job_id: Job identifier
        db: Database session

    Returns:
        Job status with progress information
    """
    status = get_extraction_status(job_id, db)

    if not status:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    return JobStatusResponse(**status)


@router.post("/{project_id}/run-full-pipeline")
async def run_full_pipeline(
    project_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Run complete pipeline: parse template → extract files → categorize → aggregate → populate Excel.

    This is the main endpoint for the Financial Builder feature.

    Args:
        project_id: Project identifier
        background_tasks: FastAPI background tasks
        db: Database session

    Returns:
        Job ID for tracking overall pipeline progress
    """
    try:
        # Step 1: Parse template
        parser = TemplateParser(project_id)
        template_dict = parser.parse()
        parser.save_to_json()

        # Step 2: Start extraction
        extractor = FileExtractor(project_id, db)
        files = extractor.scan_files()

        if not files:
            raise HTTPException(
                status_code=404,
                detail=f"No files found in project {project_id}"
            )

        job_id = extractor.create_extraction_job()

        # Run full pipeline in background
        background_tasks.add_task(
            run_pipeline_phases,
            project_id,
            job_id,
            template_dict,
            db
        )

        return {
            "success": True,
            "job_id": job_id,
            "project_id": project_id,
            "total_files": len(files),
            "total_categories": template_dict['metadata']['total_categories'],
            "message": "Full pipeline started. Phases: 1) Parse Template ✓ 2) Extract Files (starting) 3) Categorize (pending) 4) Populate Excel (pending)"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline failed: {str(e)}")


@router.get("/{project_id}/extracted-data")
async def get_extracted_data(project_id: str, db: Session = Depends(get_db)):
    """
    Get all extracted data for a project.

    Args:
        project_id: Project identifier
        db: Database session

    Returns:
        List of extracted data records
    """
    from app.models.extraction import ExtractedData

    data = db.query(ExtractedData).filter(
        ExtractedData.project_id == project_id
    ).all()

    return {
        "project_id": project_id,
        "total_files": len(data),
        "files": [
            {
                "id": d.id,
                "file_name": d.file_name,
                "file_type": d.file_type,
                "extraction_status": d.extraction_status,
                "extraction_method": d.extraction_method,
                "text_preview": d.raw_text[:200] if d.raw_text else None,
                "created_at": d.created_at.isoformat() if d.created_at else None
            }
            for d in data
        ]
    }


def run_pipeline_phases(
    project_id: str,
    job_id: str,
    template_dict: Dict[str, Any],
    db: Session
):
    """
    Background task to run all pipeline phases sequentially with normalized data layer.

    Phases:
    1. Extract all files (PDFs, Excel, CSV) → raw text
    2. Parse into normalized data points → structured table
    3. Process data points (deduplicate, detect conflicts, validate)
    4. Retrieve validated data points for categorization
    5. Categorize transactions using AI
    6. Populate Excel financial model

    This implements the friend's feedback: each file processed once,
    data stored in normalized table, conflicts can be manually resolved,
    consistent results from same input data.

    Args:
        project_id: Project identifier
        job_id: Extraction job ID
        template_dict: Parsed template dictionary
        db: Database session
    """
    try:
        logger.info(f"[{project_id}] Pipeline started - Job ID: {job_id}")

        # Phase 1: Extract files
        logger.info(f"[{project_id}] Phase 1: Extracting files...")
        try:
            extractor = FileExtractor(project_id, db)
            extractor.run_extraction(job_id)

            # Wait for extraction to complete
            job = db.query(ExtractionJob).filter(ExtractionJob.job_id == job_id).first()
            if not job or job.status != JobStatus.COMPLETED:
                logger.error(f"[{project_id}] Extraction phase failed or incomplete")
                return

            logger.info(f"[{project_id}] Phase 1 complete: {job.processed_files} files extracted")
        except Exception as e:
            logger.error(f"[{project_id}] Phase 1 ERROR: {e}", exc_info=True)
            raise Exception(f"File extraction failed: {str(e)}")

        # Phase 2: Parse extracted data into normalized data points
        logger.info(f"[{project_id}] Phase 2: Parsing extracted data into normalized data points...")
        from app.services.transaction_parser import create_transaction_parser
        from app.services.data_point_mapper import create_data_point_mapper

        extracted_data = db.query(ExtractedData).filter(
            ExtractedData.project_id == project_id,
            ExtractedData.extraction_status == "success"
        ).all()

        # Parse each extracted file into data points
        parser = create_transaction_parser(project_id)
        all_data_points = []

        for data_record in extracted_data:
            try:
                data_points = parser.parse_extracted_data(data_record, data_record.id)
                all_data_points.extend(data_points)
            except Exception as e:
                logger.error(f"Error parsing {data_record.file_name}: {e}", exc_info=True)

        logger.info(f"[{project_id}] Phase 2 complete: {len(all_data_points)} data points parsed")

        # Phase 3: Process data points (deduplicate, detect conflicts, validate)
        logger.info(f"[{project_id}] Phase 3: Processing data points (dedup, conflicts, validation)...")
        mapper = create_data_point_mapper(db, project_id)
        processed_points, conflicts = mapper.process_data_points(all_data_points)

        logger.info(f"[{project_id}] Phase 3 complete: {len(processed_points)} data points processed, "
                   f"{len(conflicts)} conflicts detected")

        # Phase 4: Get validated data points for categorization
        logger.info(f"[{project_id}] Phase 4: Retrieving validated data points...")
        valid_data_points = mapper.get_data_points_for_processing()

        # Convert data points to transaction format for categorization
        transactions = []
        for dp in valid_data_points:
            transactions.append({
                'description': dp.description,
                'amount': dp.amount,
                'vendor': dp.vendor or '',
                'date': dp.transaction_date.isoformat() if dp.transaction_date else '',
                'source_file': dp.source_file_name
            })

        logger.info(f"[{project_id}] Phase 4 complete: {len(transactions)} transactions ready for categorization")

        # Phase 5: Categorize transactions
        logger.info(f"[{project_id}] Phase 5: Categorizing {len(transactions)} transactions...")
        categorizer = create_categorizer(template_dict)
        categorized_transactions = categorizer.categorize_batch(transactions, use_llm=False)

        # Get summary statistics
        summary = categorizer.get_category_summary(categorized_transactions)
        logger.info(f"[{project_id}] Phase 5 complete: {summary['total_transactions']} categorized, "
                   f"{summary['uncategorized_count']} uncategorized, "
                   f"avg confidence {summary['average_confidence']:.2%}")

        # Phase 6: Populate Excel
        logger.info(f"[{project_id}] Phase 6: Populating Excel financial model...")

        # Prepare data for Excel
        excel_data = {
            'categorized_transactions': categorized_transactions,
            'summary': summary
        }

        # Generate output path
        # Use /tmp for cloud deployments (Render, etc.) where file system is ephemeral
        base_dir = Path("/tmp") if os.getenv("RENDER") else Path("data")
        output_dir = base_dir / "projects" / project_id / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"Financial_Model_{timestamp}.xlsx"

        # Create Excel file with error handling
        try:
            populator = create_excel_populator()
            excel_path = populator.create_financial_model(
                excel_data,
                str(output_path),
                project_name=project_id
            )
            logger.info(f"[{project_id}] Phase 6 complete: Excel generated at {excel_path}")
        except Exception as excel_error:
            # Log the error but continue - we still have the categorized data
            logger.error(f"[{project_id}] Excel generation error: {excel_error}", exc_info=True)
            # Create a minimal Excel file as fallback
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws['A1'] = "Financial Model Generation Partially Failed"
            ws['A2'] = f"Error: {str(excel_error)}"
            ws['A3'] = f"Total Transactions: {len(transactions)}"
            ws['A4'] = f"Categorized: {len(categorized_transactions)}"
            wb.save(str(output_path))
            excel_path = str(output_path)
            logger.info(f"[{project_id}] Fallback Excel created at {excel_path}")

        # Update job with final results
        job.status = JobStatus.COMPLETED
        job.job_metadata = json.dumps({
            'total_transactions': len(transactions),
            'categorized': len(categorized_transactions),
            'uncategorized_count': summary['uncategorized_count'],
            'average_confidence': summary['average_confidence'],
            'excel_path': excel_path,
            'pipeline_complete': True
        })
        db.commit()

        logger.info(f"[{project_id}] Full pipeline COMPLETE!")

    except Exception as e:
        logger.error(f"[{project_id}] Pipeline failed: {str(e)}", exc_info=True)
        # Update job status to failed
        try:
            job = db.query(ExtractionJob).filter(ExtractionJob.job_id == job_id).first()
            if job:
                job.status = JobStatus.FAILED
                job.error_message = f"Pipeline failed: {str(e)}"
                db.commit()
        except:
            pass


@router.get("/{project_id}/excel-data")
async def get_excel_data(project_id: str):
    """
    Get Excel data as JSON for dashboard display.

    Args:
        project_id: Project identifier

    Returns:
        JSON with Excel sheet data
    """
    try:
        from app.database import SessionLocal
        import openpyxl

        db = SessionLocal()

        job = db.query(ExtractionJob).filter(
            ExtractionJob.project_id == project_id,
            ExtractionJob.status == JobStatus.COMPLETED
        ).order_by(ExtractionJob.created_at.desc()).first()

        db.close()

        if not job or not job.job_metadata:
            raise HTTPException(status_code=404, detail="No completed pipeline found for this project")

        # Parse metadata to get excel_path
        metadata = json.loads(job.job_metadata)
        excel_path = metadata.get('excel_path')

        if not excel_path:
            raise HTTPException(status_code=404, detail="Excel file path not found in job metadata")

        # Construct full path
        file_path = Path(excel_path)
        if not file_path.is_absolute():
            backend_dir = Path(__file__).parent.parent.parent
            file_path = backend_dir / excel_path

        if not file_path.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_path}")

        # Read Excel file and convert to JSON with formatting
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        excel_data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Auto-detect sheet format: Check if it's key-value format or table format
            # Look at multiple rows to determine the format

            # Check first 5 rows to find a header row pattern
            has_header_row = False
            for row_idx in range(1, min(6, sheet.max_row + 1)):
                row = sheet[row_idx]
                non_empty_count = sum(1 for cell in row if cell.value and str(cell.value).strip() and str(cell.value) != 'None')
                bold_count = sum(1 for cell in row if cell.value and str(cell.value) != 'None' and cell.font and cell.font.bold)

                # If we find a row with multiple bold cells and multiple values, it's a table header
                if bold_count >= 2 and non_empty_count >= 3:
                    has_header_row = True
                    break

            # If no header row found, it's key-value format
            is_key_value_format = not has_header_row

            if is_key_value_format:
                summary_sections = []
                current_section = None

                for row_idx in range(1, sheet.max_row + 1):
                    # Check first cell for section header or key
                    cell_a = sheet.cell(row=row_idx, column=1)
                    cell_b = sheet.cell(row=row_idx, column=2)

                    # Skip empty rows
                    if not cell_a.value or cell_a.value == 'None':
                        continue

                    # Section headers are in UPPERCASE and have no value in column B
                    cell_a_value = str(cell_a.value).strip()
                    is_section_header = (
                        cell_a_value.isupper() and
                        (not cell_b.value or cell_b.value == 'None') and
                        cell_a_value not in ['NONE']
                    )

                    if is_section_header:
                        # Start new section
                        if current_section:
                            summary_sections.append(current_section)
                        current_section = {
                            'title': cell_a_value,
                            'items': []
                        }
                    elif current_section is not None and cell_b.value and cell_b.value != 'None':
                        # Data row (key-value pair)
                        value = cell_b.value
                        if isinstance(value, (int, float)):
                            value = round(value, 2) if isinstance(value, float) else value
                        current_section['items'].append({
                            'label': cell_a_value,
                            'value': value
                        })

                if current_section:
                    summary_sections.append(current_section)

                excel_data[sheet_name] = {
                    'type': 'summary',
                    'sections': summary_sections,
                    'total_rows': sheet.max_row
                }
            else:
                # Standard table format for other sheets
                # Find the header row (look for row with multiple non-empty bold cells)
                header_row_idx = 1
                for row_idx in range(1, min(10, sheet.max_row + 1)):
                    row = sheet[row_idx]
                    bold_count = sum(1 for cell in row if cell.value and cell.value != 'None' and cell.font and cell.font.bold)
                    non_empty_count = sum(1 for cell in row if cell.value and cell.value != 'None')

                    # Header row typically has multiple bold cells
                    if bold_count >= 2 and non_empty_count >= 3:
                        header_row_idx = row_idx
                        break

                sheet_data = []
                header_styles = []

                # Get headers from detected header row
                headers = []
                for cell in sheet[header_row_idx]:
                    headers.append(str(cell.value) if cell.value is not None and cell.value != 'None' else "")
                    # Capture header cell styling
                    header_styles.append({
                        'bg_color': cell.fill.start_color.index if cell.fill and cell.fill.start_color else None,
                        'font_color': cell.font.color.index if cell.font and cell.font.color else None,
                        'bold': cell.font.bold if cell.font else False
                    })

                # Get data rows starting after header (limit to first 100 rows for performance)
                for row_idx in range(header_row_idx + 1, min(header_row_idx + 101, sheet.max_row + 1)):
                    row_data = []
                    row_dict = {}
                    has_data = False

                    for col_idx, header in enumerate(headers, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        value = cell.value

                        # Convert value to string, handle None
                        if value is not None:
                            has_data = True
                            if isinstance(value, (int, float)):
                                # Format numbers properly
                                if isinstance(value, float):
                                    row_dict[header] = round(value, 2)
                                else:
                                    row_dict[header] = value
                            else:
                                row_dict[header] = str(value)
                        else:
                            row_dict[header] = ""

                        # Capture cell styling
                        row_data.append({
                            'value': row_dict[header],
                            'bg_color': cell.fill.start_color.index if cell.fill and cell.fill.start_color else None,
                            'font_color': cell.font.color.index if cell.font and cell.font.color else None,
                            'bold': cell.font.bold if cell.font else False
                        })

                    if has_data:
                        sheet_data.append({
                            'data': row_dict,
                            'styles': row_data
                        })

                excel_data[sheet_name] = {
                    'type': 'table',
                    'headers': headers,
                    'header_styles': header_styles,
                    'data': sheet_data,
                    'total_rows': sheet.max_row - 1  # Exclude header row
                }

        return {
            'success': True,
            'project_id': project_id,
            'sheets': excel_data
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel data fetch error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch Excel data: {str(e)}")


@router.get("/{project_id}/download")
async def download_excel(project_id: str):
    """
    Download the generated Excel financial model.

    Args:
        project_id: Project identifier

    Returns:
        Excel file download
    """
    try:
        # Get the most recent completed job for this project
        from app.database import SessionLocal
        db = SessionLocal()

        job = db.query(ExtractionJob).filter(
            ExtractionJob.project_id == project_id,
            ExtractionJob.status == JobStatus.COMPLETED
        ).order_by(ExtractionJob.created_at.desc()).first()

        db.close()

        if not job or not job.job_metadata:
            raise HTTPException(status_code=404, detail="No completed pipeline found for this project")

        # Parse metadata to get excel_path
        metadata = json.loads(job.job_metadata)
        excel_path = metadata.get('excel_path')

        if not excel_path:
            raise HTTPException(status_code=404, detail="Excel file path not found in job metadata")

        # Construct full path - handle both absolute and relative paths
        file_path = Path(excel_path)
        if not file_path.is_absolute():
            # If relative, make it relative to backend directory
            backend_dir = Path(__file__).parent.parent.parent
            file_path = backend_dir / excel_path

        # Check if file exists
        if not file_path.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_path}")

        # Return file
        return FileResponse(
            path=str(file_path),
            filename=f"Financial_Model_{project_id}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")
