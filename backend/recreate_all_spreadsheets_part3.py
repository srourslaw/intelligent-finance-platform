"""
Part 3: Labour timesheets, Site Reports, Variations, Client Billing - 20 files
This is a HUGE script covering folders 08-11
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import os

# Color schemes
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

BASE_PATH = "/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data"

def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def apply_borders_to_range(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = THIN_BORDER

def auto_adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================================
# 08_LABOUR_TIMESHEETS (5 files)
# ============================================================================

def create_carpenter_timesheet_week1():
    """08_LABOUR_TIMESHEETS/Carpenter_Timesheet_Week_2024_31.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet Week 31"

    ws['A1'] = "CARPENTER TIMESHEET - WEEK 31 (July 29 - Aug 4, 2024)"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    ws['A2'] = "Employee: Mark Stevens | Trade: Carpenter | Rate: $48.00/hr"
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A2:H2')

    headers = ['Date', 'Day', 'Start Time', 'Finish Time', 'Break (hrs)', 'Hours Worked', 'Rate', 'Daily Total']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['2024-07-29', 'Monday', '7:00 AM', '4:30 PM', 0.5, 9.0, 48.00, 432.00],
        ['2024-07-30', 'Tuesday', '7:00 AM', '4:30 PM', 0.5, 9.0, 48.00, 432.00],
        ['2024-07-31', 'Wednesday', '7:00 AM', '4:30 PM', 0.5, 9.0, 48.00, 432.00],
        ['2024-08-01', 'Thursday', '7:00 AM', '4:30 PM', 0.5, 9.0, 48.00, 432.00],
        ['2024-08-02', 'Friday', '7:00 AM', '3:30 PM', 0.5, 8.0, 48.00, 384.00],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(5, ws.max_row + 1):
        ws.cell(row, 7).number_format = '$#,##0.00'
        ws.cell(row, 8).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL HOURS:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F5:F{ws.max_row})")
    ws.cell(total_row, 6).font = TOTAL_FONT
    ws.cell(total_row, 6).fill = TOTAL_FILL

    ws.cell(total_row, 7, "TOTAL PAY:")
    ws.cell(total_row, 7).font = TOTAL_FONT
    ws.cell(total_row, 8, f"=SUM(H5:H{ws.max_row})")
    ws.cell(total_row, 8).font = TOTAL_FONT
    ws.cell(total_row, 8).fill = TOTAL_FILL
    ws.cell(total_row, 8).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 4, ws.max_row, 1, 8)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "08_LABOUR_TIMESHEETS", "Carpenter_Timesheet_Week_2024_31.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_labourer_timesheet():
    """08_LABOUR_TIMESHEETS/General_Labourer_Timesheet_Week_2024_32.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet Week 32"

    ws['A1'] = "GENERAL LABOURER TIMESHEET - WEEK 32 (Aug 5-11, 2024)"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    ws['A2'] = "Employee: David Nguyen | Trade: General Labourer | Rate: $35.00/hr"
    ws.merge_cells('A2:H2')

    headers = ['Date', 'Day', 'Start', 'Finish', 'Break', 'Hours', 'Rate', 'Total']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['2024-08-05', 'Mon', '7:00', '16:30', 0.5, 9.0, 35.00, 315.00],
        ['2024-08-06', 'Tue', '7:00', '16:30', 0.5, 9.0, 35.00, 315.00],
        ['2024-08-07', 'Wed', '7:00', '16:30', 0.5, 9.0, 35.00, 315.00],
        ['2024-08-08', 'Thu', '7:00', '16:30', 0.5, 9.0, 35.00, 315.00],
        ['2024-08-09', 'Fri', '7:00', '15:30', 0.5, 8.0, 35.00, 280.00],
        ['2024-08-10', 'Sat', '8:00', '12:00', 0, 4.0, 52.50, 210.00],  # Overtime 1.5x
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(5, ws.max_row + 1):
        ws.cell(row, 7).number_format = '$#,##0.00'
        ws.cell(row, 8).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F5:F{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H5:H{ws.max_row})")
    ws.cell(total_row, 6).fill = TOTAL_FILL
    ws.cell(total_row, 8).fill = TOTAL_FILL
    ws.cell(total_row, 8).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 4, ws.max_row, 1, 8)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "08_LABOUR_TIMESHEETS", "General_Labourer_Timesheet_Week_2024_32.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_supervisor_timesheet():
    """08_LABOUR_TIMESHEETS/Site_Supervisor_Timesheet_Aug_2024.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "August 2024"

    ws['A1'] = "SITE SUPERVISOR MONTHLY TIMESHEET - AUGUST 2024"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:F1')

    ws['A2'] = "Employee: Robert Chen | Position: Site Supervisor | Salary: $95,000/year"
    ws.merge_cells('A2:F2')

    headers = ['Week', 'Week Ending', 'Days Worked', 'Hours Worked', 'Notes', 'Status']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['Week 31', '2024-08-04', 5, 45, 'Framing inspection & coordination', 'APPROVED'],
        ['Week 32', '2024-08-11', 5, 48, 'Roofing commencement oversight', 'APPROVED'],
        ['Week 33', '2024-08-18', 5, 46, 'Plumbing rough-in supervision', 'APPROVED'],
        ['Week 34', '2024-08-25', 5, 47, 'Electrical coordination', 'APPROVED'],
        ['Week 35', '2024-09-01', 3, 27, 'August end close-out (3 days)', 'APPROVED'],
    ]

    for row_data in data:
        ws.append(row_data)

    total_row = ws.max_row + 2
    ws.cell(total_row, 2, "TOTAL HOURS:")
    ws.cell(total_row, 2).font = TOTAL_FONT
    ws.cell(total_row, 4, f"=SUM(D5:D{ws.max_row})")
    ws.cell(total_row, 4).font = TOTAL_FONT
    ws.cell(total_row, 4).fill = TOTAL_FILL

    apply_borders_to_range(ws, 4, ws.max_row, 1, 6)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "08_LABOUR_TIMESHEETS", "Site_Supervisor_Timesheet_Aug_2024.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_plumber_timesheet():
    """08_LABOUR_TIMESHEETS/Plumber_Timesheet_Week_2024_35.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Week 35"

    ws['A1'] = "PLUMBER TIMESHEET - WEEK 35 (Aug 26 - Sep 1, 2024)"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    ws['A2'] = "Employee: Lisa Wong | Trade: Plumber | Rate: $55.00/hr"
    ws.merge_cells('A2:H2')

    headers = ['Date', 'Day', 'Start', 'Finish', 'Break', 'Hours', 'Rate', 'Total']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['2024-08-26', 'Mon', '7:30', '16:30', 0.5, 8.5, 55.00, 467.50],
        ['2024-08-27', 'Tue', '7:30', '16:30', 0.5, 8.5, 55.00, 467.50],
        ['2024-08-28', 'Wed', '7:30', '17:00', 0.5, 9.0, 55.00, 495.00],
        ['2024-08-29', 'Thu', '7:30', '16:30', 0.5, 8.5, 55.00, 467.50],
        ['2024-08-30', 'Fri', '7:30', '15:30', 0.5, 7.5, 55.00, 412.50],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(5, ws.max_row + 1):
        ws.cell(row, 7).number_format = '$#,##0.00'
        ws.cell(row, 8).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL:")
    ws.cell(total_row, 6, f"=SUM(F5:F{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H5:H{ws.max_row})")
    ws.cell(total_row, 6).fill = TOTAL_FILL
    ws.cell(total_row, 8).fill = TOTAL_FILL
    ws.cell(total_row, 8).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 4, ws.max_row, 1, 8)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "08_LABOUR_TIMESHEETS", "Plumber_Timesheet_Week_2024_35.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_electrician_timesheet():
    """08_LABOUR_TIMESHEETS/Electrician_Timesheet_Week_2024_36.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Week 36"

    ws['A1'] = "ELECTRICIAN TIMESHEET - WEEK 36 (Sep 2-8, 2024)"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    ws['A2'] = "Employee: Tom Anderson | Trade: Electrician | Rate: $58.00/hr"
    ws.merge_cells('A2:H2')

    headers = ['Date', 'Day', 'Start', 'Finish', 'Break', 'Hours', 'Rate', 'Total']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['2024-09-02', 'Mon', '7:00', '16:00', 0.5, 8.5, 58.00, 493.00],
        ['2024-09-03', 'Tue', '7:00', '16:00', 0.5, 8.5, 58.00, 493.00],
        ['2024-09-04', 'Wed', '7:00', '16:00', 0.5, 8.5, 58.00, 493.00],
        ['2024-09-05', 'Thu', '7:00', '17:00', 0.5, 9.5, 58.00, 551.00],
        ['2024-09-06', 'Fri', '7:00', '16:00', 0.5, 8.5, 58.00, 493.00],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(5, ws.max_row + 1):
        ws.cell(row, 7).number_format = '$#,##0.00'
        ws.cell(row, 8).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL:")
    ws.cell(total_row, 6, f"=SUM(F5:F{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H5:H{ws.max_row})")
    ws.cell(total_row, 6).fill = TOTAL_FILL
    ws.cell(total_row, 8).fill = TOTAL_FILL
    ws.cell(total_row, 8).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 4, ws.max_row, 1, 8)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "08_LABOUR_TIMESHEETS", "Electrician_Timesheet_Week_2024_36.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 09_SITE_REPORTS_PHOTOS (3 files)
# ============================================================================

def create_daily_site_report_july():
    """09_SITE_REPORTS_PHOTOS/Daily_Site_Report_2024-07-25.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Site Report"

    ws['A1'] = "DAILY SITE REPORT - July 25, 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A2'] = "Project: 123 Sunset Blvd | Supervisor: Robert Chen"
    ws.merge_cells('A2:D2')

    row = 4
    ws.cell(row, 1, "Weather Conditions:").font = Font(bold=True)
    ws.cell(row, 2, "Sunny, 18°C, Light winds")
    ws.merge_cells(f'B{row}:D{row}')

    row += 1
    ws.cell(row, 1, "Trades On Site:").font = Font(bold=True)
    ws.cell(row, 2, "Framers (4), Labourers (2), Crane Operator (1)")
    ws.merge_cells(f'B{row}:D{row}')

    row += 2
    ws.cell(row, 1, "WORK COMPLETED TODAY").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:D{row}')

    headers = ['Task', 'Progress', 'Workers', 'Notes']
    row += 1
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT

    tasks = [
        ['Timber frame - South wall', '100%', 2, 'Complete and braced'],
        ['Timber frame - North wall', '75%', 2, 'To finish tomorrow'],
        ['LVL beam installation', '100%', 2, 'All beams installed'],
        ['Material delivery', '100%', 1, 'Roof trusses delivered'],
    ]

    row += 1
    for task in tasks:
        for col, value in enumerate(task, 1):
            ws.cell(row, col, value).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row, 1, "ISSUES / DELAYS:").font = Font(bold=True)
    ws.cell(row + 1, 1, "- None reported")

    row += 3
    ws.cell(row, 1, "PLANNED FOR TOMORROW:").font = Font(bold=True)
    ws.cell(row + 1, 1, "- Complete north wall framing")
    ws.cell(row + 2, 1, "- Install roof trusses (crane scheduled)")

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "09_SITE_REPORTS_PHOTOS", "Daily_Site_Report_2024-07-25.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_weekly_site_report():
    """09_SITE_REPORTS_PHOTOS/Weekly_Site_Report_Week_31.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Week 31 Summary"

    ws['A1'] = "WEEKLY SITE REPORT - WEEK 31 (July 29 - Aug 4, 2024)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A3'] = "Overall Progress: 62% Complete"
    ws['A3'].font = Font(bold=True, size=12)

    headers = ['Milestone', 'Planned %', 'Actual %', 'Variance', 'Status']
    ws.append([])
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=5)

    data = [
        ['Demolition', 100, 100, 0, 'COMPLETE'],
        ['Foundation & Slab', 100, 100, 0, 'COMPLETE'],
        ['Framing', 80, 85, 5, 'AHEAD'],
        ['Roofing', 40, 35, -5, 'SLIGHTLY BEHIND'],
        ['Rough-in Plumbing', 30, 25, -5, 'SLIGHTLY BEHIND'],
        ['Rough-in Electrical', 20, 15, -5, 'SLIGHTLY BEHIND'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(6, ws.max_row + 1):
        for col in [2, 3, 4]:
            ws.cell(row, col).number_format = '0"%"'

    ws.cell(ws.max_row + 2, 1, "KEY ACHIEVEMENTS:").font = Font(bold=True)
    ws.cell(ws.max_row + 1, 1, "- Timber framing 85% complete (ahead of schedule)")
    ws.cell(ws.max_row + 1, 1, "- All structural beams installed")
    ws.cell(ws.max_row + 1, 1, "- Roof trusses delivered")

    ws.cell(ws.max_row + 2, 1, "ISSUES:").font = Font(bold=True)
    ws.cell(ws.max_row + 1, 1, "- Plumbing materials delayed 3 days (supplier issue)")

    apply_borders_to_range(ws, 5, 11, 1, 5)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "09_SITE_REPORTS_PHOTOS", "Weekly_Site_Report_Week_31.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_safety_inspection_log():
    """09_SITE_REPORTS_PHOTOS/Safety_Inspection_Log.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Safety Inspections"

    ws['A1'] = "SITE SAFETY INSPECTION LOG"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    headers = ['Inspection Date', 'Inspector', 'Area Inspected', 'Issues Found', 'Corrective Action', 'Status']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=3)

    data = [
        ['2024-07-01', 'SafeWork NSW', 'Scaffolding Setup', 'None', 'N/A', 'PASSED'],
        ['2024-07-15', 'R. Chen', 'Site Perimeter Fencing', 'Gap in fence near gate', 'Repaired same day', 'CLOSED'],
        ['2024-07-22', 'R. Chen', 'Electrical Cords & Tools', 'Damaged extension cord', 'Cord replaced', 'CLOSED'],
        ['2024-08-05', 'R. Chen', 'Fall Protection', 'None', 'N/A', 'PASSED'],
        ['2024-08-12', 'SafeWork NSW', 'General Site Inspection', 'None', 'N/A', 'PASSED'],
        ['2024-08-19', 'R. Chen', 'Ladder Safety', 'Ladder not tagged', 'Tag applied', 'CLOSED'],
        ['2024-08-26', 'R. Chen', 'PPE Compliance', 'All workers compliant', 'N/A', 'PASSED'],
    ]

    for row_data in data:
        ws.append(row_data)

    apply_borders_to_range(ws, 3, ws.max_row, 1, 6)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "09_SITE_REPORTS_PHOTOS", "Safety_Inspection_Log.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 10_VARIATIONS_CHANGES (2 files)
# ============================================================================

def create_variation_order_register():
    """10_VARIATIONS_CHANGES/Variation_Order_Register.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variation Orders"

    headers = ['VO Number', 'Date Raised', 'Description', 'Requested By', 'Cost Impact',
               'Time Impact (days)', 'Status', 'Approval Date', 'Notes']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['VO-001', '2024-07-10', 'Upgrade kitchen benchtop to stone', 'Client', 3200, 0, 'APPROVED', '2024-07-12', 'Change from laminate'],
        ['VO-002', '2024-07-18', 'Add skylights in living area (2 units)', 'Architect', 4500, 2, 'APPROVED', '2024-07-20', 'Design enhancement'],
        ['VO-003', '2024-08-05', 'Additional power points in bedrooms (6 units)', 'Client', 850, 0, 'APPROVED', '2024-08-06', 'Client request'],
        ['VO-004', '2024-08-15', 'Change from carpet to timber flooring', 'Client', 8500, 3, 'APPROVED', '2024-08-17', 'Upgrade to engineered oak'],
        ['VO-005', '2024-08-22', 'Extra bathroom exhaust fan', 'Building Inspector', 650, 0, 'APPROVED', '2024-08-22', 'Compliance requirement'],
        ['VO-006', '2024-09-01', 'Upgrade to designer light fixtures', 'Client', 2800, 0, 'PENDING', '-', 'Quote provided'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 5).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 4, "TOTAL VARIATIONS:")
    ws.cell(total_row, 4).font = TOTAL_FONT
    ws.cell(total_row, 5, f"=SUM(E2:E{ws.max_row})")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 5).fill = TOTAL_FILL
    ws.cell(total_row, 5).number_format = '$#,##0.00'

    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 6).font = TOTAL_FONT
    ws.cell(total_row, 6).fill = TOTAL_FILL

    apply_borders_to_range(ws, 1, ws.max_row, 1, 9)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "10_VARIATIONS_CHANGES", "Variation_Order_Register.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_change_order_log():
    """10_VARIATIONS_CHANGES/Change_Order_Log.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Change Orders"

    ws['A1'] = "CHANGE ORDER LOG - 123 Sunset Blvd Project"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:G1')

    headers = ['CO Number', 'Related VO', 'Description', 'Original Cost', 'Revised Cost', 'Difference', 'Status']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=3)

    data = [
        ['CO-001', 'VO-001', 'Kitchen benchtop stone upgrade', 1850, 5050, 3200, 'INVOICED'],
        ['CO-002', 'VO-002', 'Skylight installation', 0, 4500, 4500, 'INVOICED'],
        ['CO-003', 'VO-003', 'Additional power points', 0, 850, 850, 'INVOICED'],
        ['CO-004', 'VO-004', 'Timber flooring upgrade', 4200, 12700, 8500, 'APPROVED'],
        ['CO-005', 'VO-005', 'Extra exhaust fan', 0, 650, 650, 'APPROVED'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(4, ws.max_row + 1):
        for col in [4, 5, 6]:
            ws.cell(row, col).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL CHANGE ORDERS:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F4:F{ws.max_row})")
    ws.cell(total_row, 6).font = TOTAL_FONT
    ws.cell(total_row, 6).fill = TOTAL_FILL
    ws.cell(total_row, 6).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 3, ws.max_row, 1, 7)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "10_VARIATIONS_CHANGES", "Change_Order_Log.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 11_CLIENT_BILLING (5 files)
# ============================================================================

def create_progress_claim_1():
    """11_CLIENT_BILLING/Progress_Claim_01_July_2024.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress Claim 01"

    ws['A1'] = "PROGRESS CLAIM #01 - JULY 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Project: 123 Sunset Blvd | Client: Smith Family Trust"
    ws.merge_cells('A2:E2')
    ws['A3'] = "Claim Period: June 15 - July 31, 2024"
    ws.merge_cells('A3:E3')

    headers = ['Work Package', 'Contract Value', 'Claimed to Date', '% Complete', 'This Claim']
    ws.append([])
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=6)

    data = [
        ['Site Establishment', 8000, 8000, 100, 8000],
        ['Demolition', 15000, 15000, 100, 15000],
        ['Foundation Works', 45000, 45000, 100, 45000],
        ['Slab', 38000, 38000, 100, 38000],
        ['Framing', 66000, 19800, 30, 19800],
        ['Preliminaries', 25000, 7500, 30, 7500],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(7, ws.max_row + 1):
        for col in [2, 3, 5]:
            ws.cell(row, col).number_format = '$#,##0.00'
        ws.cell(row, 4).number_format = '0"%"'

    total_row = ws.max_row + 2
    ws.cell(total_row, 4, "SUBTOTAL:")
    ws.cell(total_row, 4).font = TOTAL_FONT
    ws.cell(total_row, 5, f"=SUM(E7:E{ws.max_row})")
    ws.cell(total_row, 5).number_format = '$#,##0.00'
    ws.cell(total_row, 5).fill = TOTAL_FILL

    ws.cell(total_row + 1, 4, "GST (10%):")
    ws.cell(total_row + 1, 5, f"=E{total_row}*0.1")
    ws.cell(total_row + 1, 5).number_format = '$#,##0.00'

    ws.cell(total_row + 2, 4, "CLAIM TOTAL:")
    ws.cell(total_row + 2, 4).font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 5, f"=E{total_row}+E{total_row+1}")
    ws.cell(total_row + 2, 5).font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 5).number_format = '$#,##0.00'
    ws.cell(total_row + 2, 5).fill = TOTAL_FILL

    apply_borders_to_range(ws, 6, ws.max_row, 1, 5)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "11_CLIENT_BILLING", "Progress_Claim_01_July_2024.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_progress_claim_2():
    """11_CLIENT_BILLING/Progress_Claim_02_August_2024.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress Claim 02"

    ws['A1'] = "PROGRESS CLAIM #02 - AUGUST 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A2'] = "Claim Period: Aug 1-31, 2024"
    ws.merge_cells('A2:F2')

    headers = ['Work Package', 'Contract Value', 'Previous Claims', 'This Claim', '% Complete', 'Claimed to Date']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=4)

    data = [
        ['Framing', 66000, 19800, 39600, 90, 59400],
        ['Roofing', 28500, 0, 21375, 75, 21375],
        ['Windows & Doors', 22915, 0, 11457.50, 50, 11457.50],
        ['Rough Plumbing', 20000, 0, 12000, 60, 12000],
        ['Rough Electrical', 18000, 0, 9000, 50, 9000],
        ['Preliminaries', 25000, 7500, 6250, 55, 13750],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(5, ws.max_row + 1):
        for col in [2, 3, 4, 6]:
            ws.cell(row, col).number_format = '$#,##0.00'
        ws.cell(row, 5).number_format = '0"%"'

    total_row = ws.max_row + 2
    ws.cell(total_row, 3, "SUBTOTAL:")
    ws.cell(total_row, 4, f"=SUM(D5:D{ws.max_row})")
    ws.cell(total_row, 4).number_format = '$#,##0.00'
    ws.cell(total_row, 4).fill = TOTAL_FILL

    ws.cell(total_row + 1, 3, "GST:")
    ws.cell(total_row + 1, 4, f"=D{total_row}*0.1")
    ws.cell(total_row + 1, 4).number_format = '$#,##0.00'

    ws.cell(total_row + 2, 3, "TOTAL CLAIM:")
    ws.cell(total_row + 2, 4, f"=D{total_row}+D{total_row+1}")
    ws.cell(total_row + 2, 4).number_format = '$#,##0.00'
    ws.cell(total_row + 2, 4).font = Font(bold=True, size=12)
    ws.cell(total_row + 2, 4).fill = TOTAL_FILL

    apply_borders_to_range(ws, 4, ws.max_row, 1, 6)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "11_CLIENT_BILLING", "Progress_Claim_02_August_2024.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_client_payment_record():
    """11_CLIENT_BILLING/Client_Payment_Record.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Record"

    headers = ['Invoice/Claim', 'Issue Date', 'Amount (Inc GST)', 'Due Date', 'Payment Date', 'Amount Paid', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Initial Deposit', '2024-06-15', 50000.00, '2024-06-20', '2024-06-18', 50000.00, 'PAID'],
        ['Progress Claim 01', '2024-08-01', 146630.00, '2024-08-15', '2024-08-12', 146630.00, 'PAID'],
        ['Progress Claim 02', '2024-09-01', 109682.50, '2024-09-15', '2024-09-10', 109682.50, 'PAID'],
        ['Variation VO-001', '2024-07-15', 3520.00, '2024-07-29', '2024-07-28', 3520.00, 'PAID'],
        ['Variation VO-002', '2024-07-25', 4950.00, '2024-08-08', '2024-08-05', 4950.00, 'PAID'],
        ['Variation VO-003', '2024-08-10', 935.00, '2024-08-24', 'PENDING', 0, 'OUTSTANDING'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(2, ws.max_row + 1):
        for col in [3, 6]:
            ws.cell(row, col).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 2, "TOTAL INVOICED:")
    ws.cell(total_row, 2).font = TOTAL_FONT
    ws.cell(total_row, 3, f"=SUM(C2:C{ws.max_row})")
    ws.cell(total_row, 3).number_format = '$#,##0.00'
    ws.cell(total_row, 3).fill = TOTAL_FILL

    ws.cell(total_row + 1, 2, "TOTAL RECEIVED:")
    ws.cell(total_row + 1, 2).font = TOTAL_FONT
    ws.cell(total_row + 1, 3, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row + 1, 3).number_format = '$#,##0.00'
    ws.cell(total_row + 1, 3).fill = TOTAL_FILL

    ws.cell(total_row + 2, 2, "OUTSTANDING:")
    ws.cell(total_row + 2, 2).font = Font(bold=True, color="FF0000")
    ws.cell(total_row + 2, 3, f"=C{total_row}-C{total_row+1}")
    ws.cell(total_row + 2, 3).number_format = '$#,##0.00'
    ws.cell(total_row + 2, 3).font = Font(bold=True, color="FF0000")

    apply_borders_to_range(ws, 1, ws.max_row, 1, 7)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "11_CLIENT_BILLING", "Client_Payment_Record.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_retention_money_schedule():
    """11_CLIENT_BILLING/Retention_Money_Schedule.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retention Schedule"

    ws['A1'] = "RETENTION MONEY SCHEDULE - 5% Retention"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:F1')

    headers = ['Claim/Invoice', 'Claim Amount', 'Retention Rate', 'Retention Held', 'Net Payment', 'Retention Status']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=3)

    data = [
        ['Progress Claim 01', 146630.00, 0.05, 7331.50, 139298.50, 'HELD'],
        ['Progress Claim 02', 109682.50, 0.05, 5484.13, 104198.37, 'HELD'],
        ['Variation VO-001', 3520.00, 0.05, 176.00, 3344.00, 'HELD'],
        ['Variation VO-002', 4950.00, 0.05, 247.50, 4702.50, 'HELD'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).number_format = '0.0%'
        for col in [2, 4, 5]:
            ws.cell(row, col).number_format = '$#,##0.00'

    total_row = ws.max_row + 2
    ws.cell(total_row, 3, "TOTAL RETENTION HELD:")
    ws.cell(total_row, 3).font = TOTAL_FONT
    ws.cell(total_row, 4, f"=SUM(D4:D{ws.max_row})")
    ws.cell(total_row, 4).font = TOTAL_FONT
    ws.cell(total_row, 4).fill = TOTAL_FILL
    ws.cell(total_row, 4).number_format = '$#,##0.00'

    ws.cell(total_row + 2, 1, "NOTE: Retention released upon Practical Completion + 3 months (or Defects Liability Period end)")
    ws.cell(total_row + 2, 1).font = Font(italic=True, size=10)
    ws.merge_cells(f'A{total_row+2}:F{total_row+2}')

    apply_borders_to_range(ws, 3, ws.max_row, 1, 6)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "11_CLIENT_BILLING", "Retention_Money_Schedule.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_invoice_aging_report():
    """11_CLIENT_BILLING/Invoice_Aging_Report.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Report"

    ws['A1'] = "INVOICE AGING REPORT - As at Sep 15, 2024"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:F1')

    headers = ['Invoice', 'Issue Date', 'Amount', 'Days Outstanding', 'Aging Category', 'Status']
    ws.append([])
    ws.append(headers)
    apply_header_style(ws, row=3)

    data = [
        ['Progress Claim 01', '2024-08-01', 146630.00, 0, 'PAID', 'Paid 08/12'],
        ['Progress Claim 02', '2024-09-01', 109682.50, 0, 'PAID', 'Paid 09/10'],
        ['VO-003', '2024-08-10', 935.00, 36, '31-60 days', 'OVERDUE'],
        ['VO-004', '2024-08-20', 9350.00, 26, '0-30 days', 'DUE SOON'],
    ]

    for row_data in data:
        ws.append(row_data)

    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).number_format = '$#,##0.00'

    # Highlight overdue
    for row in range(4, ws.max_row + 1):
        if 'OVERDUE' in str(ws.cell(row, 5).value):
            for col in range(1, 7):
                ws.cell(row, col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    apply_borders_to_range(ws, 3, ws.max_row, 1, 6)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "11_CLIENT_BILLING", "Invoice_Aging_Report.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("RECREATING EXCEL SPREADSHEETS - PART 3 (LABOUR, SITE REPORTS, VARIATIONS, CLIENT BILLING)")
    print("=" * 80)

    # 08_LABOUR_TIMESHEETS (5 files)
    print("\n[08_LABOUR_TIMESHEETS]")
    create_carpenter_timesheet_week1()
    create_labourer_timesheet()
    create_supervisor_timesheet()
    create_plumber_timesheet()
    create_electrician_timesheet()

    # 09_SITE_REPORTS_PHOTOS (3 files)
    print("\n[09_SITE_REPORTS_PHOTOS]")
    create_daily_site_report_july()
    create_weekly_site_report()
    create_safety_inspection_log()

    # 10_VARIATIONS_CHANGES (2 files)
    print("\n[10_VARIATIONS_CHANGES]")
    create_variation_order_register()
    create_change_order_log()

    # 11_CLIENT_BILLING (5 files)
    print("\n[11_CLIENT_BILLING]")
    create_progress_claim_1()
    create_progress_claim_2()
    create_client_payment_record()
    create_retention_money_schedule()
    create_invoice_aging_report()

    print("\n" + "=" * 80)
    print("PART 3 COMPLETE: 15 files created")
    print("CUMULATIVE TOTAL: 8 + 7 + 15 = 30 files")
    print("=" * 80)
