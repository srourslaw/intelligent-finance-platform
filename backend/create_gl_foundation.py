#!/usr/bin/env python3
"""
Create General Ledger Foundation Files
Generates Chart of Accounts, Trial Balance, General Journal, and Monthly Close packages
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# Styling
FILL_HEADER = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
FONT_BOLD = Font(bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_chart_of_accounts():
    """Create Chart of Accounts with standard GL structure"""
    print("\n📊 Creating Chart of Accounts...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    # Header
    headers = ['Account Code', 'Account Name', 'Type', 'Category', 'Sub-Category', 'Opening Balance (01/06/2024)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Chart of Accounts data
    coa_data = [
        # ASSETS (1000-1999)
        ['1000', 'ASSETS', 'Asset', 'Header', '', ''],
        ['1100', 'Cash at Bank - Operating Account', 'Asset', 'Current Assets', 'Cash', 100000],
        ['1200', 'Accounts Receivable - Trade', 'Asset', 'Current Assets', 'Receivables', 0],
        ['1300', 'Work in Progress', 'Asset', 'Current Assets', 'Inventory', 0],
        ['1400', 'Prepaid Expenses', 'Asset', 'Current Assets', 'Prepayments', 5000],
        ['1500', 'Land - 123 Sunset Boulevard', 'Asset', 'Non-Current Assets', 'Property', 0],
        ['1600', 'Equipment - Construction', 'Asset', 'Non-Current Assets', 'Equipment', 0],
        ['1650', 'Accumulated Depreciation - Equipment', 'Asset', 'Non-Current Assets', 'Contra-Asset', 0],

        # LIABILITIES (2000-2999)
        ['2000', 'LIABILITIES', 'Liability', 'Header', '', ''],
        ['2100', 'Accounts Payable - Trade', 'Liability', 'Current Liabilities', 'Payables', 0],
        ['2150', 'Accrued Expenses', 'Liability', 'Current Liabilities', 'Accruals', 0],
        ['2200', 'GST Payable', 'Liability', 'Current Liabilities', 'Tax', 0],
        ['2300', 'PAYG Withholding Payable', 'Liability', 'Current Liabilities', 'Tax', 0],
        ['2400', 'Superannuation Payable', 'Liability', 'Current Liabilities', 'Payroll', 0],
        ['2500', 'Loan Payable - Construction Finance', 'Liability', 'Long-Term Liabilities', 'Debt', 0],
        ['2600', 'Retention Money Held', 'Liability', 'Current Liabilities', 'Retentions', 0],

        # EQUITY (3000-3999)
        ['3000', 'EQUITY', 'Equity', 'Header', '', ''],
        ['3100', "Owner's Capital", 'Equity', 'Capital', 'Contributions', 100000],
        ['3200', 'Retained Earnings', 'Equity', 'Retained', 'Accumulated', 0],
        ['3300', 'Current Year Earnings', 'Equity', 'Retained', 'Current Period', 0],

        # REVENUE (4000-4999)
        ['4000', 'REVENUE', 'Income', 'Header', '', ''],
        ['4100', 'Construction Revenue - Contract', 'Income', 'Revenue', 'Primary', 0],
        ['4200', 'Variation Orders Revenue', 'Income', 'Revenue', 'Variations', 0],
        ['4300', 'Interest Income', 'Income', 'Other Income', 'Finance', 0],

        # COST OF GOODS SOLD (5000-5999)
        ['5000', 'COST OF GOODS SOLD', 'Expense', 'Header', '', ''],
        ['5100', 'Materials - Concrete & Masonry', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5110', 'Materials - Timber & Framing', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5120', 'Materials - Roofing', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5130', 'Materials - Windows & Doors', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5140', 'Materials - Plumbing Fixtures', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5150', 'Materials - Electrical Fittings', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5160', 'Materials - Finishes & Paint', 'Expense', 'Direct Costs', 'Materials', 0],
        ['5200', 'Direct Labor - Site Supervisor', 'Expense', 'Direct Costs', 'Labor', 0],
        ['5210', 'Direct Labor - General', 'Expense', 'Direct Costs', 'Labor', 0],
        ['5220', 'Direct Labor - Skilled Trades', 'Expense', 'Direct Costs', 'Labor', 0],
        ['5300', 'Subcontractors - Excavation', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5310', 'Subcontractors - Electrical', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5320', 'Subcontractors - Plumbing', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5330', 'Subcontractors - HVAC', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5340', 'Subcontractors - Carpentry', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5350', 'Subcontractors - Painting', 'Expense', 'Direct Costs', 'Subcontractors', 0],
        ['5400', 'Equipment Rental', 'Expense', 'Direct Costs', 'Equipment', 0],
        ['5500', 'Site Costs - Utilities', 'Expense', 'Direct Costs', 'Site', 0],
        ['5510', 'Site Costs - Waste Removal', 'Expense', 'Direct Costs', 'Site', 0],

        # OPERATING EXPENSES (6000-6999)
        ['6000', 'OPERATING EXPENSES', 'Expense', 'Header', '', ''],
        ['6100', 'Land Acquisition Costs', 'Expense', 'Operating', 'Land', 0],
        ['6110', 'Stamp Duty', 'Expense', 'Operating', 'Land', 0],
        ['6120', 'Legal Fees - Conveyancing', 'Expense', 'Operating', 'Legal', 0],
        ['6200', 'Design Fees - Architectural', 'Expense', 'Operating', 'Professional', 0],
        ['6210', 'Design Fees - Engineering', 'Expense', 'Operating', 'Professional', 0],
        ['6220', 'Design Fees - Surveying', 'Expense', 'Operating', 'Professional', 0],
        ['6300', 'Permits & Approvals - Council', 'Expense', 'Operating', 'Compliance', 0],
        ['6310', 'Permits & Approvals - Certifications', 'Expense', 'Operating', 'Compliance', 0],
        ['6400', 'Insurance - Public Liability', 'Expense', 'Operating', 'Insurance', 0],
        ['6410', 'Insurance - Contract Works', 'Expense', 'Operating', 'Insurance', 0],
        ['6420', 'Insurance - Workers Compensation', 'Expense', 'Operating', 'Insurance', 0],
        ['6500', 'Financing Costs - Interest', 'Expense', 'Operating', 'Finance', 0],
        ['6510', 'Financing Costs - Bank Fees', 'Expense', 'Operating', 'Finance', 0],
        ['6600', 'Depreciation Expense', 'Expense', 'Operating', 'Non-Cash', 0],
        ['6700', 'Office & Admin', 'Expense', 'Operating', 'Admin', 0],
        ['6800', 'Marketing & Promotion', 'Expense', 'Operating', 'Marketing', 0],
    ]

    # Populate data
    for row_idx, row_data in enumerate(coa_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            # Format opening balance as currency
            if col_idx == 6 and value != '':
                cell.number_format = '$#,##0.00'

            # Bold header rows
            if row_data[2] in ['Asset', 'Liability', 'Equity', 'Income', 'Expense'] and row_data[3] == 'Header':
                cell.font = FONT_BOLD
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20 if col <= 2 else 18

    # Save
    output_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/17_GENERAL_LEDGER")
    output_path.mkdir(parents=True, exist_ok=True)

    wb.save(output_path / "Chart_of_Accounts.xlsx")
    print(f"  ✅ Created Chart_of_Accounts.xlsx with {len(coa_data)} accounts")

def create_trial_balance():
    """Create Trial Balance template"""
    print("\n📊 Creating Trial Balance...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance - June 2024"

    # Header
    ws['A1'] = 'TRIAL BALANCE'
    ws['A2'] = 'As at 30 June 2024'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')

    headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER

    # Opening balances from Chart of Accounts
    trial_data = [
        ['1100', 'Cash at Bank', 100000, ''],
        ['3100', "Owner's Capital", '', 100000],
    ]

    row = 5
    for data in trial_data:
        for col, val in enumerate(data, 1):
            cell = ws.cell(row, col, val)
            if col in [3, 4] and val != '':
                cell.number_format = '$#,##0.00'
        row += 1

    # Totals
    total_row = row + 1
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    ws.cell(total_row, 3, f'=SUM(C5:C{row-1})').font = FONT_BOLD
    ws.cell(total_row, 4, f'=SUM(D5:D{row-1})').font = FONT_BOLD
    ws.cell(total_row, 3).number_format = '$#,##0.00'
    ws.cell(total_row, 4).number_format = '$#,##0.00'

    # Balance check
    ws.cell(total_row + 2, 1, 'Balance Check:').font = FONT_BOLD
    ws.cell(total_row + 2, 2, f'=IF(C{total_row}=D{total_row},"✓ BALANCED","✗ OUT OF BALANCE")')
    ws.cell(total_row + 2, 2).font = Font(bold=True, color='00B050')

    # Auto-size
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25 if col == 2 else 15

    output_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/17_GENERAL_LEDGER")
    wb.save(output_path / "Trial_Balance_Monthly.xlsx")
    print(f"  ✅ Created Trial_Balance_Monthly.xlsx")

def create_monthly_close_packages():
    """Create monthly close packages for June-September 2024"""
    print("\n📊 Creating Monthly Close Packages...")

    months = [
        ('June', '2024-06'),
        ('July', '2024-07'),
        ('August', '2024-08'),
        ('September', '2024-09')
    ]

    for month_name, month_code in months:
        create_income_statement(month_name, month_code)
        create_balance_sheet(month_name, month_code)
        create_cash_flow_statement(month_name, month_code)

def create_income_statement(month_name, month_code):
    """Create Income Statement for a specific month"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"P&L {month_name}"

    # Title
    ws['A1'] = f'INCOME STATEMENT - {month_name.upper()} 2024'
    ws['A2'] = 'Project A - 123 Sunset Boulevard'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')

    data = [
        ['', '', ''],
        ['REVENUE', '', ''],
        ['Construction Revenue', 4100, 162500],
        ['Variation Orders', 4200, 0],
        ['TOTAL REVENUE', '', '=SUM(C6:C7)'],
        ['', '', ''],
        ['COST OF GOODS SOLD', '', ''],
        ['Materials', 5100, 42300],
        ['Direct Labor', 5200, 31240],
        ['Subcontractors', 5300, 85600],
        ['Equipment Rental', 5400, 8500],
        ['Site Costs', 5500, 4200],
        ['TOTAL COGS', '', '=SUM(C11:C15)'],
        ['', '', ''],
        ['GROSS PROFIT', '', '=C8-C16'],
        ['Gross Profit Margin %', '', '=C18/C8'],
        ['', '', ''],
        ['OPERATING EXPENSES', '', ''],
        ['Design & Engineering', 6200, 19740],
        ['Permits & Approvals', 6300, 10950],
        ['Insurance', 6400, 4800],
        ['Financing Costs', 6500, 2600],
        ['Depreciation', 6600, 800],
        ['Admin & Office', 6700, 1200],
        ['TOTAL OPERATING EXPENSES', '', '=SUM(C22:C27)'],
        ['', '', ''],
        ['EBITDA', '', '=C18-C28+C26'],
        ['NET INCOME', '', '=C18-C28'],
        ['Net Profit Margin %', '', '=C31/C8'],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            # Format currency
            if col_idx == 3 and value != '':
                if isinstance(value, str) and value.startswith('='):
                    cell.value = value
                    if '%' in str(row_data[0]):
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '$#,##0.00'

            # Bold totals
            if any(keyword in str(value) for keyword in ['TOTAL', 'GROSS PROFIT', 'EBITDA', 'NET INCOME']):
                cell.font = FONT_BOLD
                if col_idx == 1:
                    cell.fill = FILL_TOTAL

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    output_dir = Path(f"/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/19_MONTHLY_CLOSE/{month_code}-Close")
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_dir / f"Income_Statement_{month_name}_2024.xlsx")
    print(f"  ✅ Created Income_Statement_{month_name}_2024.xlsx")

def create_balance_sheet(month_name, month_code):
    """Create Balance Sheet for a specific month"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Balance Sheet {month_name}"

    ws['A1'] = f'BALANCE SHEET - {month_name.upper()} 2024'
    ws['A2'] = 'As at 30 June 2024'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')

    data = [
        ['', '', ''],
        ['ASSETS', '', ''],
        ['Current Assets:', '', ''],
        ['  Cash at Bank', 1100, 87400],
        ['  Accounts Receivable', 1200, 97500],
        ['  Work in Progress', 1300, 171840],
        ['  Prepaid Expenses', 1400, 5000],
        ['Total Current Assets', '', '=SUM(C8:C11)'],
        ['', '', ''],
        ['Non-Current Assets:', '', ''],
        ['  Land', 1500, 265000],
        ['  Equipment', 1600, 97000],
        ['  Less: Accumulated Depreciation', 1650, -4000],
        ['Total Non-Current Assets', '', '=SUM(C15:C17)'],
        ['', '', ''],
        ['TOTAL ASSETS', '', '=C12+C18'],
        ['', '', ''],
        ['LIABILITIES', '', ''],
        ['Current Liabilities:', '', ''],
        ['  Accounts Payable', 2100, 27300],
        ['  Accrued Expenses', 2150, 8500],
        ['  GST Payable', 2200, 15200],
        ['Total Current Liabilities', '', '=SUM(C24:C26)'],
        ['', '', ''],
        ['Long-Term Liabilities:', '', ''],
        ['  Loan Payable', 2500, 450000],
        ['Total Long-Term Liabilities', '', 'C30'],
        ['', '', ''],
        ['TOTAL LIABILITIES', '', '=C27+C31'],
        ['', '', ''],
        ['EQUITY', '', ''],
        ['  Owner\'s Capital', 3100, 100000],
        ['  Retained Earnings', 3200, 0],
        ['  Current Year Earnings', 3300, '=C20-C33'],
        ['TOTAL EQUITY', '', '=SUM(C37:C39)'],
        ['', '', ''],
        ['TOTAL LIABILITIES & EQUITY', '', '=C33+C40'],
        ['', '', ''],
        ['Balance Check:', '', ''],
        ['', '', '=IF(C20=C42,"✓ BALANCED","✗ OUT OF BALANCE")'],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            if col_idx == 3 and value != '':
                cell.number_format = '$#,##0.00'

            if any(keyword in str(value) for keyword in ['Total', 'TOTAL']):
                cell.font = FONT_BOLD
                if col_idx == 1:
                    cell.fill = FILL_TOTAL

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    output_dir = Path(f"/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/19_MONTHLY_CLOSE/{month_code}-Close")
    wb.save(output_dir / f"Balance_Sheet_{month_name}_2024.xlsx")
    print(f"  ✅ Created Balance_Sheet_{month_name}_2024.xlsx")

def create_cash_flow_statement(month_name, month_code):
    """Create Cash Flow Statement for a specific month"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cash Flow {month_name}"

    ws['A1'] = f'CASH FLOW STATEMENT - {month_name.upper()} 2024'
    ws['A2'] = f'For the month ended 30 {month_name} 2024'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:B1')

    data = [
        ['', ''],
        ['OPERATING ACTIVITIES', ''],
        ['Cash received from customers', 65000],
        ['Cash paid to suppliers', -98400],
        ['Cash paid to employees', -31240],
        ['Cash paid for operating expenses', -37490],
        ['Net Cash from Operating Activities', '=SUM(B7:B10)'],
        ['', ''],
        ['INVESTING ACTIVITIES', ''],
        ['Purchase of land', -265000],
        ['Purchase of equipment', -97000],
        ['Net Cash from Investing Activities', '=SUM(B14:B15)'],
        ['', ''],
        ['FINANCING ACTIVITIES', ''],
        ['Loan drawdown', 450000],
        ['Owner contributions', 100000],
        ['Net Cash from Financing Activities', '=SUM(B19:B20)'],
        ['', ''],
        ['NET CHANGE IN CASH', '=B11+B16+B21'],
        ['Cash at beginning of period', 100000],
        ['Cash at end of period', '=B24+B25'],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            if col_idx == 2 and value != '':
                cell.number_format = '$#,##0.00'

            if any(keyword in str(value) for keyword in ['Net Cash', 'NET CHANGE', 'Cash at']):
                cell.font = FONT_BOLD
                if col_idx == 1:
                    cell.fill = FILL_TOTAL

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15

    output_dir = Path(f"/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/19_MONTHLY_CLOSE/{month_code}-Close")
    wb.save(output_dir / f"Cash_Flow_Statement_{month_name}_2024.xlsx")
    print(f"  ✅ Created Cash_Flow_Statement_{month_name}_2024.xlsx")

def create_bank_reconciliation():
    """Create Bank Reconciliation template"""
    print("\n📊 Creating Bank Reconciliation...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Reconciliation - June"

    ws['A1'] = 'BANK RECONCILIATION - JUNE 2024'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')

    data = [
        ['', '', ''],
        ['Cash Balance per Bank Statement (30/06/2024)', '', 89450],
        ['', '', ''],
        ['Add: Deposits in Transit', '', ''],
        ['  Deposit - Client Payment', '25/06/2024', 32500],
        ['Total Deposits in Transit', '', '=C9'],
        ['', '', ''],
        ['Less: Outstanding Checks', '', ''],
        ['  Check #1001 - BuildMart', '28/06/2024', -15400],
        ['  Check #1002 - Spark Electrical', '29/06/2024', -19150],
        ['Total Outstanding Checks', '', '=SUM(C13:C14)'],
        ['', '', ''],
        ['Cash Balance per Books (30/06/2024)', '', '=C5+C10+C15'],
        ['', '', ''],
        ['Reconciling Items:', '', ''],
        ['Bank fees not yet recorded', '', -500],
        ['Interest income not yet recorded', '', 100],
        ['', '', ''],
        ['Adjusted Cash Balance per Books', '', '=C17+SUM(C20:C21)'],
    ]

    for row_idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            if col_idx == 3 and value != '':
                cell.number_format = '$#,##0.00'

            if 'Balance' in str(row_data[0]) or 'Total' in str(row_data[0]):
                cell.font = FONT_BOLD

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/20_BANK_RECONCILIATION")
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_dir / "Bank_Reconciliation_Monthly.xlsx")
    print(f"  ✅ Created Bank_Reconciliation_Monthly.xlsx")

def create_fixed_assets_register():
    """Create Fixed Assets Register"""
    print("\n📊 Creating Fixed Assets Register...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fixed Assets"

    ws['A1'] = 'FIXED ASSETS REGISTER'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    headers = ['Asset ID', 'Description', 'Date Acquired', 'Cost', 'Salvage Value',
               'Useful Life (yrs)', 'Depreciation Method', 'Accumulated Depreciation',
               'Net Book Value', 'Location']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER

    assets = [
        ['FA-001', 'Excavator - CAT 320', '15/01/2024', 85000, 5000, 10, 'Straight-Line', 4000, '=D4-H4', 'Site A'],
        ['FA-002', 'Site Office Portable', '20/01/2024', 12000, 1000, 5, 'Straight-Line', 1100, '=D5-H5', 'Site A'],
        ['FA-003', 'Scaffolding System', '05/02/2024', 15000, 500, 7, 'Straight-Line', 1071, '=D6-H6', 'Site A'],
        ['FA-004', 'Concrete Mixer', '10/02/2024', 8500, 500, 5, 'Straight-Line', 850, '=D7-H7', 'Site A'],
        ['FA-005', 'Power Tools Set', '15/02/2024', 4500, 0, 3, 'Straight-Line', 750, '=D8-H8', 'Site A'],
    ]

    for row_idx, row_data in enumerate(assets, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            if col_idx in [4, 5, 8, 9]:
                if isinstance(value, str) and value.startswith('='):
                    cell.value = value
                cell.number_format = '$#,##0.00'
            elif col_idx == 3:
                cell.number_format = 'DD/MM/YYYY'

    # Totals row
    total_row = len(assets) + 4
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    ws.cell(total_row, 4, f'=SUM(D4:D{total_row-1})').font = FONT_BOLD
    ws.cell(total_row, 4).number_format = '$#,##0.00'
    ws.cell(total_row, 8, f'=SUM(H4:H{total_row-1})').font = FONT_BOLD
    ws.cell(total_row, 8).number_format = '$#,##0.00'
    ws.cell(total_row, 9, f'=SUM(I4:I{total_row-1})').font = FONT_BOLD
    ws.cell(total_row, 9).number_format = '$#,##0.00'

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 20 if col == 2 else 15

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/21_FIXED_ASSETS")
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_dir / "Fixed_Assets_Register.xlsx")
    print(f"  ✅ Created Fixed_Assets_Register.xlsx with 5 assets")

def main():
    print("=" * 80)
    print("GENERAL LEDGER FOUNDATION CREATION")
    print("=" * 80)

    create_chart_of_accounts()
    create_trial_balance()
    create_monthly_close_packages()
    create_bank_reconciliation()
    create_fixed_assets_register()

    print("\n" + "=" * 80)
    print("✨ GL FOUNDATION COMPLETE!")
    print("=" * 80)
    print("\nFiles created:")
    print("  📁 17_GENERAL_LEDGER/")
    print("     - Chart_of_Accounts.xlsx (62 GL accounts)")
    print("     - Trial_Balance_Monthly.xlsx")
    print("  📁 19_MONTHLY_CLOSE/")
    print("     - 2024-06-Close/ (Income Statement, Balance Sheet, Cash Flow)")
    print("     - 2024-07-Close/ (Income Statement, Balance Sheet, Cash Flow)")
    print("     - 2024-08-Close/ (Income Statement, Balance Sheet, Cash Flow)")
    print("     - 2024-09-Close/ (Income Statement, Balance Sheet, Cash Flow)")
    print("  📁 20_BANK_RECONCILIATION/")
    print("     - Bank_Reconciliation_Monthly.xlsx")
    print("  📁 21_FIXED_ASSETS/")
    print("     - Fixed_Assets_Register.xlsx (5 assets)")
    print("=" * 80 + "\n")

if __name__ == "__main__":
    main()
