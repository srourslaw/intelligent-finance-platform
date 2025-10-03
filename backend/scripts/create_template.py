"""
Create a default financial template Excel file.

This template includes:
- Balance Sheet
- Income Statement
- Cash Flow Statement
- Pre-formatted with headers and formulas
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path

def create_financial_template():
    """Create a comprehensive financial template"""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create Balance Sheet
    create_balance_sheet(wb)

    # Create Income Statement
    create_income_statement(wb)

    # Create Cash Flow Statement
    create_cash_flow(wb)

    # Save template
    output_dir = Path("data/templates")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "financial_template.xlsx"

    wb.save(output_path)
    print(f"✅ Template created: {output_path}")
    return str(output_path)


def create_balance_sheet(wb: Workbook):
    """Create Balance Sheet with proper structure"""
    ws = wb.create_sheet("Balance Sheet")

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # Title
    ws['A1'] = "BALANCE SHEET"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws.merge_cells('A1:C1')

    # Date header
    ws['B3'] = "Amount ($)"
    ws['B3'].font = Font(bold=True)
    ws['B3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ASSETS
    row = 4
    ws[f'A{row}'] = "ASSETS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Current Assets
    ws[f'A{row}'] = "Current Assets"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    current_assets = [
        ("Cash on Hand", "B5"),
        ("Cash in Bank - Operating", "B6"),
        ("Accounts Receivable", "B7"),
        ("Allowance for Doubtful Accounts", "B8"),
        ("Prepaid Expenses", "B9"),
        ("Inventory - Raw Materials", "B10"),
        ("Inventory - Work in Progress", "B11"),
        ("Inventory - Finished Goods", "B12"),
    ]

    for item, cell in current_assets:
        ws[f'A{row}'] = item
        ws[cell].number_format = '$#,##0.00'
        row += 1

    # Total Current Assets (formula)
    ws[f'A{row}'] = "Total Current Assets"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B5:B12)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # Non-Current Assets
    ws[f'A{row}'] = "Non-Current Assets"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    non_current_assets = [
        ("Land", "B20"),
        ("Buildings", "B21"),
        ("Machinery & Equipment", "B22"),
        ("Vehicles", "B23"),
        ("Furniture & Fixtures", "B24"),
        ("Accumulated Depreciation", "B25"),
    ]

    for item, cell in non_current_assets:
        ws[f'A{row}'] = item
        ws[cell].number_format = '$#,##0.00'
        row += 1

    # Total Non-Current Assets
    ws[f'A{row}'] = "Total Non-Current Assets"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B20:B25)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # TOTAL ASSETS
    ws[f'A{row}'] = "TOTAL ASSETS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws[f'B{row}'] = "=B13+B26"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 2

    # LIABILITIES
    ws[f'A{row}'] = "LIABILITIES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Current Liabilities
    ws[f'A{row}'] = "Current Liabilities"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    current_liabilities = [
        ("Accounts Payable", "B35"),
        ("Credit Card Debt", "B36"),
        ("Short-term Loans", "B37"),
        ("Accrued Expenses", "B38"),
        ("Taxes Payable", "B39"),
    ]

    for item, cell in current_liabilities:
        ws[f'A{row}'] = item
        ws[cell].number_format = '$#,##0.00'
        row += 1

    # Total Current Liabilities
    ws[f'A{row}'] = "Total Current Liabilities"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B35:B39)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # Long-term Liabilities
    ws[f'A{row}'] = "Long-term Liabilities"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Mortgage Payable"
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "Long-term Loans"
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1

    # Total Long-term Liabilities
    ws[f'A{row}'] = "Total Long-term Liabilities"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B45:B46)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1

    # TOTAL LIABILITIES
    ws[f'A{row}'] = "TOTAL LIABILITIES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=B40+B47"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # EQUITY
    ws[f'A{row}'] = "EQUITY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Share Capital"
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "Retained Earnings"
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1

    # Total Equity
    ws[f'A{row}'] = "TOTAL EQUITY"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B55:B56)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # TOTAL LIABILITIES + EQUITY
    ws[f'A{row}'] = "TOTAL LIABILITIES + EQUITY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws[f'B{row}'] = "=B48+B57"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18


def create_income_statement(wb: Workbook):
    """Create Income Statement"""
    ws = wb.create_sheet("Income Statement")

    # Header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "INCOME STATEMENT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font

    ws['C3'] = "Amount ($)"
    ws['C3'].font = Font(bold=True)
    ws['C3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Revenue
    row = 4
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Product Sales"
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "Service Revenue"
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "TOTAL REVENUE"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C5:C6)"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 2

    # COGS
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    cogs_items = [
        "Beginning Inventory",
        "Purchases",
        "Direct Labor",
        "Manufacturing Overhead",
    ]

    for item in cogs_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'].number_format = '$#,##0.00'
        row += 1

    ws[f'A{row}'] = "Less: Ending Inventory"
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "TOTAL COGS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C15:C18)-C19"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 2

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C7-C20"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 2

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    opex_items = [
        "Salaries and Wages",
        "Rent",
        "Utilities",
        "Insurance",
        "Depreciation",
        "Marketing",
        "Office Supplies",
        "Professional Fees",
    ]

    for item in opex_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'].number_format = '$#,##0.00'
        row += 1

    ws[f'A{row}'] = "TOTAL OPERATING EXPENSES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C25:C32)"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 2

    # Operating Income
    ws[f'A{row}'] = "OPERATING INCOME"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C21-C33"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 2

    # Other Income/Expenses
    ws[f'A{row}'] = "Interest Expense"
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "Other Income"
    ws[f'C{row}'].number_format = '$#,##0.00'
    row += 2

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True, size=12)
    ws[f'C{row}'] = "=C34-C37+C38"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'C{row}'].font = Font(color="FFFFFF", bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['C'].width = 18


def create_cash_flow(wb: Workbook):
    """Create Cash Flow Statement"""
    ws = wb.create_sheet("Cash Flow")

    # Header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "CASH FLOW STATEMENT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font

    ws['D3'] = "Amount ($)"
    ws['D3'].font = Font(bold=True)
    ws['D3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Operating Activities
    row = 4
    ws[f'A{row}'] = "OPERATING ACTIVITIES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Net Profit"
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 2

    ws[f'A{row}'] = "Adjustments:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    operating_items = [
        "Add: Depreciation",
        "Decrease in Accounts Receivable",
        "Increase in Inventory",
        "Increase in Accounts Payable",
    ]

    for item in operating_items:
        ws[f'A{row}'] = item
        ws[f'D{row}'].number_format = '$#,##0.00'
        row += 1

    ws[f'A{row}'] = "Net Cash from Operating Activities"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "=SUM(D5:D13)"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 2

    # Investing Activities
    ws[f'A{row}'] = "INVESTING ACTIVITIES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    investing_items = [
        "Purchase of Equipment",
        "Purchase of Property",
        "Sale of Assets",
    ]

    for item in investing_items:
        ws[f'A{row}'] = item
        ws[f'D{row}'].number_format = '$#,##0.00'
        row += 1

    ws[f'A{row}'] = "Net Cash from Investing Activities"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "=SUM(D25:D27)"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 2

    # Financing Activities
    ws[f'A{row}'] = "FINANCING ACTIVITIES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    financing_items = [
        "Proceeds from Loans",
        "Repayment of Debt",
        "Dividends Paid",
    ]

    for item in financing_items:
        ws[f'A{row}'] = item
        ws[f'D{row}'].number_format = '$#,##0.00'
        row += 1

    ws[f'A{row}'] = "Net Cash from Financing Activities"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "=SUM(D35:D37)"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 2

    # Net Change in Cash
    ws[f'A{row}'] = "NET CHANGE IN CASH"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "=D14+D28+D38"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 2

    # Beginning/Ending Cash
    ws[f'A{row}'] = "Beginning Cash Balance"
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 1

    ws[f'A{row}'] = "ENDING CASH BALANCE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True, size=12)
    ws[f'D{row}'] = "=D39+D41"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'D{row}'].font = Font(color="FFFFFF", bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['D'].width = 18


if __name__ == "__main__":
    create_financial_template()
