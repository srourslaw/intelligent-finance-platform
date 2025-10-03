"""
Mountain View Terrace - Realistic Document Generator
Creates Excel spreadsheets and PDF invoices for a luxury mountain construction project

All vendors, amounts, and styling completely different from other projects
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import random

# Project Configuration
PROJECT_NAME = "789 Mountain View Terrace"
PROJECT_ID = "C-789-MV"
CLIENT_NAME = "Michael & Emma Chen"
BUILDER = "Mountain Homes Builders Pty Ltd"
ABN = "78 945 612 345"

# Mountain-themed color scheme (different from project-a)
MOUNTAIN_BLUE = "2C5F7F"
MOUNTAIN_GREEN = "4A7C59"
MOUNTAIN_GREY = "6B7280"
ACCENT_ORANGE = "D97706"

class MountainViewDocumentGenerator:
    """Generate realistic construction documents for Mountain View project"""

    def __init__(self):
        self.base_path = Path("data")
        self.start_date = datetime(2024, 8, 1)

    def create_invoice_pdf(self, invoice_data, output_path):
        """Create a professional PDF invoice"""
        c = canvas.Canvas(str(output_path), pagesize=A4)
        width, height = A4

        # Header with mountain theme
        c.setFillColorRGB(0.17, 0.37, 0.50)  # Mountain blue
        c.rect(0, height - 120, width, 120, fill=True, stroke=False)

        # Company logo area (text-based)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 24)
        c.drawString(40, height - 60, "MOUNTAIN HOMES")
        c.setFont("Helvetica", 12)
        c.drawString(40, height - 80, "BUILDERS PTY LTD")
        c.setFont("Helvetica", 9)
        c.drawString(40, height - 95, "ABN: 78 945 612 345")

        # Invoice title
        c.setFont("Helvetica-Bold", 28)
        c.drawRightString(width - 40, height - 60, "INVOICE")
        c.setFont("Helvetica", 11)
        c.drawRightString(width - 40, height - 85, f"#{invoice_data['invoice_number']}")

        # Company details (left side)
        y = height - 150
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, "Mountain Homes Builders Pty Ltd")
        c.setFont("Helvetica", 9)
        y -= 15
        c.drawString(40, y, "123 Alpine Way, Katoomba NSW 2780")
        y -= 12
        c.drawString(40, y, "Phone: (02) 4782 5555")
        y -= 12
        c.drawString(40, y, "Email: accounts@mountainhomes.com.au")

        # Bill To (right side)
        y = height - 150
        c.setFont("Helvetica-Bold", 10)
        c.drawString(width/2 + 20, y, "BILL TO:")
        c.setFont("Helvetica", 9)
        y -= 15
        c.drawString(width/2 + 20, y, CLIENT_NAME)
        y -= 12
        c.drawString(width/2 + 20, y, PROJECT_NAME)
        y -= 12
        c.drawString(width/2 + 20, y, "Blue Mountains NSW 2780")

        # Invoice details box
        y = height - 240
        c.setFillColorRGB(0.42, 0.47, 0.50)  # Mountain grey
        c.rect(40, y - 5, width - 80, 60, fill=True, stroke=False)

        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(50, y + 40, "Invoice Date:")
        c.drawString(50, y + 25, "Due Date:")
        c.drawString(50, y + 10, "Project:")

        c.setFont("Helvetica", 9)
        c.drawString(150, y + 40, invoice_data['date'])
        c.drawString(150, y + 25, invoice_data['due_date'])
        c.drawString(150, y + 10, f"{PROJECT_ID}")

        c.setFont("Helvetica-Bold", 9)
        c.drawString(width/2 + 20, y + 40, "Terms:")
        c.drawString(width/2 + 20, y + 25, "Status:")

        c.setFont("Helvetica", 9)
        c.drawString(width/2 + 100, y + 40, "Net 30")
        status_color = (0.2, 0.7, 0.3) if invoice_data.get('paid') else (0.85, 0.46, 0.02)
        c.setFillColorRGB(*status_color)
        c.drawString(width/2 + 100, y + 25, "PAID" if invoice_data.get('paid') else "PENDING")

        # Line items table
        y -= 100
        c.setFillColorRGB(0, 0, 0)

        # Table header
        c.setFillColorRGB(0.29, 0.49, 0.35)  # Mountain green
        c.rect(40, y, width - 80, 25, fill=True, stroke=False)

        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y + 10, "Description")
        c.drawString(width - 250, y + 10, "Qty")
        c.drawString(width - 190, y + 10, "Unit Price")
        c.drawString(width - 120, y + 10, "Amount")

        # Line items
        y -= 5
        c.setFillColorRGB(0, 0, 0)
        for item in invoice_data['items']:
            y -= 20
            c.setFont("Helvetica", 9)
            c.drawString(50, y, item['description'][:60])
            c.drawString(width - 250, y, str(item.get('quantity', 1)))
            c.drawString(width - 190, y, f"${item.get('unit_price', item['amount']):,.2f}")
            c.drawString(width - 120, y, f"${item['amount']:,.2f}")

            # Draw line
            c.setStrokeColorRGB(0.8, 0.8, 0.8)
            c.line(40, y - 5, width - 40, y - 5)

        # Totals section
        y -= 40
        c.setFont("Helvetica-Bold", 10)
        c.drawString(width - 250, y, "Subtotal:")
        c.drawString(width - 120, y, f"${invoice_data['subtotal']:,.2f}")

        y -= 20
        c.drawString(width - 250, y, "GST (10%):")
        c.drawString(width - 120, y, f"${invoice_data['gst']:,.2f}")

        y -= 25
        c.setFillColorRGB(0.29, 0.49, 0.35)
        c.rect(width - 260, y - 5, 220, 25, fill=True, stroke=False)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(width - 250, y + 5, "TOTAL:")
        c.drawString(width - 120, y + 5, f"${invoice_data['total']:,.2f}")

        # Footer
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.setFont("Helvetica", 8)
        footer_y = 50
        c.drawCentredString(width/2, footer_y, "Mountain Homes Builders Pty Ltd | Licensed Builder NSW #298456")
        c.drawCentredString(width/2, footer_y - 12, "Payment Details: BSB 012-345 | Account 9876-5432 | Reference: " + invoice_data['invoice_number'])

        c.save()

    def create_purchase_orders_excel(self):
        """Create Purchase Orders Master Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Orders"

        # Different styling - Orange/Grey theme
        header_fill = PatternFill(start_color=ACCENT_ORANGE, end_color=ACCENT_ORANGE, fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        alt_row_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"{PROJECT_NAME} - Purchase Orders Register"
        title_cell.font = Font(bold=True, size=16, color=MOUNTAIN_BLUE)
        title_cell.alignment = Alignment(horizontal='center')

        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
        ws['A2'].font = Font(italic=True, size=9, color=MOUNTAIN_GREY)

        # Headers
        headers = ['PO Number', 'Date', 'Supplier', 'Category', 'Description', 'Amount', 'GST', 'Total', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Purchase orders data
        pos = [
            ['PO-MV-001', '2024-08-15', 'Mountain Earthmoving Services', 'Site Works', 'Site clearing and bulk earthworks', 35300, 3530, 38830, 'Completed'],
            ['PO-MV-002', '2024-08-20', 'Blue Mountains Concrete Solutions', 'Foundations', 'Strip footings and concrete', 41925, 4192.50, 46117.50, 'Completed'],
            ['PO-MV-003', '2024-09-10', 'Alpine Timber & Hardware', 'Framing', 'Timber framing materials - full house', 68800, 6880, 75680, 'In Progress'],
            ['PO-MV-004', '2024-09-20', 'Mountain Plumbing Supplies', 'Plumbing', 'Copper pipes, fixtures, hot water system', 22380, 2238, 24618, 'Ordered'],
            ['PO-MV-005', '2024-09-25', 'Highland Roofing Materials', 'Roofing', 'Colorbond Ultra + skylights', 45200, 4520, 49720, 'Quoted'],
            ['PO-MV-006', '2024-10-01', 'Mountain View Windows & Doors', 'Windows', 'Double-glazed windows bushfire rated', 68500, 6850, 75350, 'Quoted'],
        ]

        row = 5
        for po in pos:
            for col, value in enumerate(po, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                if col in [6, 7, 8]:  # Currency columns
                    cell.number_format = '$#,##0.00'
                if row % 2 == 0:
                    cell.fill = alt_row_fill
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 32
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

        # Summary
        summary_row = row + 2
        ws[f'E{summary_row}'] = 'TOTAL PURCHASE ORDERS:'
        ws[f'E{summary_row}'].font = Font(bold=True)
        ws[f'H{summary_row}'] = f'=SUM(H5:H{row-1})'
        ws[f'H{summary_row}'].number_format = '$#,##0.00'
        ws[f'H{summary_row}'].font = Font(bold=True, size=12, color=MOUNTAIN_BLUE)

        wb.save(self.base_path / "06_PURCHASE_ORDERS_INVOICES" / "Purchase_Orders_Master.xlsx")
        print("✅ Created Purchase Orders Master.xlsx")

    def create_invoices(self):
        """Create invoice PDFs"""
        invoice_folder = self.base_path / "06_PURCHASE_ORDERS_INVOICES"
        paid_folder = invoice_folder / "Invoices_Paid"
        pending_folder = invoice_folder / "Invoices_Pending"

        paid_folder.mkdir(exist_ok=True)
        pending_folder.mkdir(exist_ok=True)

        # Invoice 1 - Paid
        inv1 = {
            'invoice_number': 'MES-2024-1847',
            'date': '20 August 2024',
            'due_date': '10 September 2024',
            'paid': True,
            'items': [
                {'description': 'Site clearing and vegetation removal', 'quantity': 1, 'unit_price': 8500, 'amount': 8500},
                {'description': 'Bulk excavation - 180m³ @ $45/m³', 'quantity': 180, 'unit_price': 45, 'amount': 8100},
                {'description': 'Cut and fill earthworks - stepped design', 'quantity': 1, 'unit_price': 12500, 'amount': 12500},
                {'description': 'Site compaction and preparation', 'quantity': 1, 'unit_price': 6200, 'amount': 6200},
            ],
            'subtotal': 35300,
            'gst': 3530,
            'total': 38830
        }
        self.create_invoice_pdf(inv1, paid_folder / "MES-2024-1847.pdf")

        # Invoice 2 - Paid
        inv2 = {
            'invoice_number': 'BMCS-2024-5623',
            'date': '5 September 2024',
            'due_date': '25 September 2024',
            'paid': True,
            'items': [
                {'description': 'N32 Concrete - Strip footings 45m³ @ $285/m³', 'quantity': 45, 'unit_price': 285, 'amount': 12825},
                {'description': 'Concrete pump hire and placement', 'quantity': 1, 'unit_price': 850, 'amount': 850},
                {'description': 'Reinforcement steel mesh and bar', 'quantity': 1, 'unit_price': 8950, 'amount': 8950},
                {'description': 'Formwork, boxing and strip', 'quantity': 1, 'unit_price': 6500, 'amount': 6500},
                {'description': 'Labour - placement and finishing', 'quantity': 1, 'unit_price': 12800, 'amount': 12800},
            ],
            'subtotal': 41925,
            'gst': 4192.50,
            'total': 46117.50
        }
        self.create_invoice_pdf(inv2, paid_folder / "BMCS-2024-5623.pdf")

        # Invoice 3 - Pending
        inv3 = {
            'invoice_number': 'ATH-2024-7891',
            'date': '15 September 2024',
            'due_date': '5 October 2024',
            'paid': False,
            'items': [
                {'description': 'F27 Pine framing timber - wall frames', 'quantity': 1, 'unit_price': 28500, 'amount': 28500},
                {'description': 'LVL beams and lintels - engineered', 'quantity': 1, 'unit_price': 12400, 'amount': 12400},
                {'description': 'Roof trusses - engineered timber', 'quantity': 1, 'unit_price': 18900, 'amount': 18900},
                {'description': 'Treated pine - retaining & decking', 'quantity': 1, 'unit_price': 5800, 'amount': 5800},
                {'description': 'Fixings, nails, brackets, hangers', 'quantity': 1, 'unit_price': 3200, 'amount': 3200},
            ],
            'subtotal': 68800,
            'gst': 6880,
            'total': 75680
        }
        self.create_invoice_pdf(inv3, pending_folder / "ATH-2024-7891.pdf")

        # Invoice 4 - Pending
        inv4 = {
            'invoice_number': 'MPS-2024-3456',
            'date': '25 September 2024',
            'due_date': '15 October 2024',
            'paid': False,
            'items': [
                {'description': 'Copper piping and fittings - full house', 'quantity': 1, 'unit_price': 6800, 'amount': 6800},
                {'description': 'PVC drainage pipes 100mm/150mm', 'quantity': 1, 'unit_price': 2850, 'amount': 2850},
                {'description': 'Rheem 315L hot water system', 'quantity': 1, 'unit_price': 2950, 'amount': 2950},
                {'description': 'Bathroom fixtures - premium range', 'quantity': 1, 'unit_price': 8500, 'amount': 8500},
                {'description': 'Grohe kitchen tapware suite', 'quantity': 1, 'unit_price': 1280, 'amount': 1280},
            ],
            'subtotal': 22380,
            'gst': 2238,
            'total': 24618
        }
        self.create_invoice_pdf(inv4, pending_folder / "MPS-2024-3456.pdf")

        print("✅ Created 4 invoice PDFs (2 paid, 2 pending)")

    def create_budget_tracker(self):
        """Create comprehensive budget tracking spreadsheet"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget Tracker"

        # Mountain theme colors
        header_fill = PatternFill(start_color=MOUNTAIN_BLUE, end_color=MOUNTAIN_BLUE, fill_type="solid")
        category_fill = PatternFill(start_color=MOUNTAIN_GREEN, end_color=MOUNTAIN_GREEN, fill_type="solid")
        total_fill = PatternFill(start_color=ACCENT_ORANGE, end_color=ACCENT_ORANGE, fill_type="solid")

        # Title section
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{PROJECT_NAME} - Budget Tracker"
        ws['A1'].font = Font(bold=True, size=18, color=MOUNTAIN_BLUE)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Contract Value: $820,000 | As of: {datetime.now().strftime('%d %B %Y')}"
        ws['A2'].font = Font(size=11, color=MOUNTAIN_GREY)
        ws.merge_cells('A2:G2')

        # Headers
        headers = ['Category', 'Budgeted', 'Committed', 'Spent', 'Variance', 'Status', '% Used']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Budget categories
        categories = [
            ['Land Acquisition', 418215, 418215, 418215, 0, 'Complete', '=D5/B5'],
            ['Design & Consultants', 75000, 73550, 73550, 1450, 'Complete', '=D6/B6'],
            ['Permits & Approvals', 28000, 26950, 26950, 1050, 'Complete', '=D7/B7'],
            ['Site Works', 73500, 67330, 67330, 6170, 'Complete', '=D8/B8'],
            ['Foundations', 78000, 46118, 46118, 31882, 'Under Budget', '=D9/B9'],
            ['Framing & Structure', 125000, 75680, 0, 49320, 'In Progress', '=D10/B10'],
            ['Roofing', 52000, 0, 0, 52000, 'Not Started', '=D11/B11'],
            ['External Cladding', 68000, 0, 0, 68000, 'Not Started', '=D12/B12'],
            ['Windows & Doors', 85000, 0, 0, 85000, 'Not Started', '=D13/B13'],
            ['Plumbing', 68000, 24618, 0, 43382, 'Materials Ordered', '=D14/B14'],
            ['Electrical', 72000, 0, 0, 72000, 'Not Started', '=D15/B15'],
            ['Insulation', 18000, 0, 0, 18000, 'Not Started', '=D16/B16'],
            ['Plasterboard', 45000, 0, 0, 45000, 'Not Started', '=D17/B17'],
            ['Tiling', 48000, 0, 0, 48000, 'Not Started', '=D18/B18'],
            ['Kitchen', 55000, 0, 0, 55000, 'Not Started', '=D19/B19'],
            ['Bathrooms', 48000, 0, 0, 48000, 'Not Started', '=D20/B20'],
            ['Flooring', 42000, 0, 0, 42000, 'Not Started', '=D21/B21'],
            ['Painting', 35000, 0, 0, 35000, 'Not Started', '=D22/B22'],
            ['Cabinetry', 38000, 0, 0, 38000, 'Not Started', '=D23/B23'],
            ['Fixtures & Fittings', 28000, 0, 0, 28000, 'Not Started', '=D24/B24'],
        ]

        row = 5
        for cat in categories:
            for col, value in enumerate(cat, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                if col in [2, 3, 4, 5]:  # Currency
                    cell.number_format = '$#,##0'
                if col == 7:  # Percentage
                    cell.number_format = '0%'
            row += 1

        # Total row
        total_row = row
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{total_row}'].fill = total_fill

        for col in ['B', 'C', 'D', 'E']:
            cell = ws[f'{col}{total_row}']
            cell.value = f'=SUM({col}5:{col}{row-1})'
            cell.number_format = '$#,##0'
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = total_fill

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 10

        wb.save(self.base_path / "12_BUDGET_TRACKING" / "Budget_Tracker.xlsx")
        print("✅ Created Budget Tracker.xlsx")

    def create_subcontractor_register(self):
        """Create subcontractor register"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subcontractors"

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{PROJECT_NAME} - Subcontractor Register"
        ws['A1'].font = Font(bold=True, size=16, color=MOUNTAIN_BLUE)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Company', 'Trade', 'Contact', 'Phone', 'Contract Value', 'Start Date', 'Status', 'Insurance']
        header_fill = PatternFill(start_color=MOUNTAIN_GREEN, end_color=MOUNTAIN_GREEN, fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        # Subcontractors
        subs = [
            ['Elite Framing Solutions', 'Carpentry', 'James Robertson', '0428 765 432', 125000, '15-Sep-24', 'In Progress', '✓'],
            ['Alpine Plumbing Services', 'Plumbing', 'Robert Chen', '0412 543 876', 68000, '1-Oct-24', 'Scheduled', '✓'],
            ['PowerTech Electrical', 'Electrical', 'Sarah Mitchell', '0421 876 543', 72000, '15-Oct-24', 'Scheduled', '✓'],
            ['Mountain View Roofing', 'Roofing', 'Tom Wilson', '0438 234 567', 52000, '1-Nov-24', 'Scheduled', '✓'],
            ['Precision Tiling Co', 'Tiling', 'Maria Santos', '0434 567 890', 48000, '15-Jan-25', 'Not Started', '✓'],
        ]

        for row_idx, sub in enumerate(subs, 4):
            for col_idx, value in enumerate(sub, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                if col_idx == 5:
                    cell.number_format = '$#,##0'

        # Column widths
        for col, width in enumerate([30, 15, 20, 15, 15, 12, 15, 10], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(self.base_path / "07_SUBCONTRACTORS" / "Subcontractor_Register.xlsx")
        print("✅ Created Subcontractor Register.xlsx")

    def create_client_billing(self):
        """Create client billing/progress claims"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progress Claims"

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{PROJECT_NAME} - Client Progress Claims"
        ws['A1'].font = Font(bold=True, size=16, color=MOUNTAIN_BLUE)

        # Headers
        headers = ['Claim #', 'Date', 'Description', 'Work Value', 'Cumulative %', 'Total Claim']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=MOUNTAIN_BLUE, end_color=MOUNTAIN_BLUE, fill_type="solid")

        # Claims
        claims = [
            ['PC-001', '1-Aug-24', 'Contract deposit (10%)', 82000, 10, 90200],
            ['PC-002', '15-Sep-24', 'Site works and foundations complete', 113448, 23.8, 124792.80],
        ]

        for row_idx, claim in enumerate(claims, 4):
            for col_idx, value in enumerate(claim, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                if col_idx in [4, 6]:
                    cell.number_format = '$#,##0.00'
                if col_idx == 5:
                    cell.number_format = '0.0%'
                    cell.value = value / 100

        wb.save(self.base_path / "11_CLIENT_BILLING" / "Progress_Claims.xlsx")
        print("✅ Created Progress Claims.xlsx")

def main():
    """Generate all documents"""
    print("🏔️  Mountain View Terrace - Document Generator")
    print("=" * 60)

    generator = MountainViewDocumentGenerator()

    generator.create_purchase_orders_excel()
    generator.create_invoices()
    generator.create_budget_tracker()
    generator.create_subcontractor_register()
    generator.create_client_billing()

    print("\n" + "=" * 60)
    print("✅ All documents generated successfully!")
    print(f"📁 Location: {generator.base_path.absolute()}")
    print("\nDocuments created:")
    print("  • Purchase Orders Master (Excel)")
    print("  • 4 Invoice PDFs (2 paid, 2 pending)")
    print("  • Budget Tracker (Excel)")
    print("  • Subcontractor Register (Excel)")
    print("  • Progress Claims (Excel)")

if __name__ == "__main__":
    main()
