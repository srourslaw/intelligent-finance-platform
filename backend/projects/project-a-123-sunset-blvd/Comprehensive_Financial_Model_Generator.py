"""
Comprehensive Financial Model Generator
Generates a complete Excel workbook with Balance Sheet, Income Statement, 
Cash Flow, Equity Statement, Ratios Dashboard, and Assumptions sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar


class FinancialModelGenerator:
    """Generate comprehensive financial model Excel workbook"""
    
    def __init__(self, company_name="My Company", period_type="yearly", 
                 num_periods=5, start_month=1, start_year=2025):
        """
        Initialize the financial model generator
        
        Args:
            company_name: Name of company/project
            period_type: 'monthly', 'quarterly', or 'yearly'
            num_periods: Number of periods to generate
            start_month: Starting month (1-12)
            start_year: Starting year
        """
        self.company_name = company_name
        self.period_type = period_type
        self.num_periods = num_periods
        self.start_month = start_month
        self.start_year = start_year
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Define styles
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.section_font = Font(bold=True, size=11)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.total_font = Font(bold=True, size=10)
        self.total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def generate_period_headers(self):
        """Generate period headers based on configuration"""
        periods = ['Line Item']
        month_names = list(calendar.month_abbr)[1:]  # Jan-Dec
        
        current_month = self.start_month
        current_year = self.start_year
        current_quarter = (self.start_month - 1) // 3 + 1
        
        for i in range(self.num_periods):
            if self.period_type == 'monthly':
                periods.append(f"{month_names[current_month - 1]} {current_year}")
                current_month += 1
                if current_month > 12:
                    current_month = 1
                    current_year += 1
            elif self.period_type == 'quarterly':
                periods.append(f"Q{current_quarter} {current_year}")
                current_quarter += 1
                if current_quarter > 4:
                    current_quarter = 1
                    current_year += 1
            else:  # yearly
                periods.append(f"FY {current_year}")
                current_year += 1
        
        return periods
    
    def style_header_row(self, ws, row_num, num_cols):
        """Apply header styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def style_section_row(self, ws, row_num, num_cols):
        """Apply section styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.section_font
            cell.fill = self.section_fill
            cell.border = self.border
    
    def style_total_row(self, ws, row_num, num_cols):
        """Apply total styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.total_font
            cell.fill = self.total_fill
            cell.border = self.border
    
    def create_balance_sheet(self):
        """Create Balance Sheet with formulas"""
        ws = self.wb.create_sheet("Balance Sheet")
        periods = self.generate_period_headers()
        num_cols = len(periods)
        
        # Title
        ws['A1'] = f"{self.company_name} - BALANCE SHEET"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        
        ws['A2'] = "As at Period End"
        ws['A2'].font = Font(italic=True, size=11)
        ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
        
        # Period headers
        current_row = 4
        for col_idx, period in enumerate(periods, start=1):
            ws.cell(row=current_row, column=col_idx, value=period)
        self.style_header_row(ws, current_row, num_cols)
        
        # ASSETS
        current_row += 1
        ws.cell(row=current_row, column=1, value="ASSETS")
        self.style_section_row(ws, current_row, num_cols)
        
        # Current Assets
        current_row += 1
        ws.cell(row=current_row, column=1, value="Current Assets")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        current_assets = [
            'Cash on Hand', 'Cash in Bank - Operating', 'Cash in Bank - Savings',
            'Short-term Investments', 'Accounts Receivable', 'Allowance for Doubtful Debts',
            'Inventory - Raw Materials', 'Inventory - WIP', 'Inventory - Finished Goods',
            'Prepaid Rent', 'Prepaid Insurance', 'Prepaid Rates', 
            'Other Prepaid Expenses', 'Other Current Assets'
        ]
        
        start_row = current_row + 1
        for item in current_assets:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Current Assets
        current_row += 1
        total_current_row = current_row
        ws.cell(row=current_row, column=1, value="Total Current Assets")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col, 
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # Non-Current Assets
        current_row += 2
        ws.cell(row=current_row, column=1, value="Non-Current Assets")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        non_current_assets = [
            'Land', 'Buildings', 'Leasehold Improvements', 'Machinery and Equipment',
            'Motor Vehicles', 'Furniture and Fixtures', 'Computer Equipment',
            'Office Equipment', 'Accumulated Depreciation', 'Intangible Assets',
            'Goodwill', 'Patents and Trademarks', 'Software', 'Accumulated Amortization',
            'Long-term Investments', 'Other Non-Current Assets'
        ]
        
        start_row = current_row + 1
        for item in non_current_assets:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Non-Current Assets
        current_row += 1
        total_non_current_row = current_row
        ws.cell(row=current_row, column=1, value="Total Non-Current Assets")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # TOTAL ASSETS
        current_row += 2
        total_assets_row = current_row
        ws.cell(row=current_row, column=1, value="TOTAL ASSETS")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{total_current_row}+{col_letter}{total_non_current_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # LIABILITIES
        current_row += 2
        ws.cell(row=current_row, column=1, value="LIABILITIES")
        self.style_section_row(ws, current_row, num_cols)
        
        # Current Liabilities
        current_row += 1
        ws.cell(row=current_row, column=1, value="Current Liabilities")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        current_liabilities = [
            'Bank Overdraft', 'Accounts Payable', 'Credit Card Debt', 'GST/VAT Payable',
            'Income Tax Payable', 'PAYG Withholding', 'Superannuation Payable',
            'Workcover Payable', 'Accrued Expenses', 'Accrued Salaries',
            'Unearned Revenue', 'Current Portion of Long-term Debt', 'Other Current Liabilities'
        ]
        
        start_row = current_row + 1
        for item in current_liabilities:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Current Liabilities
        current_row += 1
        total_current_liab_row = current_row
        ws.cell(row=current_row, column=1, value="Total Current Liabilities")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # Long-term Liabilities
        current_row += 2
        ws.cell(row=current_row, column=1, value="Long-term Liabilities")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        long_term_liabilities = [
            'Mortgage Payable', 'Equipment Finance', 'Motor Vehicle Loans',
            'Long-term Bank Loans', 'Bonds Payable', 'Lease Liabilities',
            'Long-term Provisions', 'Deferred Tax Liabilities', 'Other Long-term Liabilities'
        ]
        
        start_row = current_row + 1
        for item in long_term_liabilities:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Long-term Liabilities
        current_row += 1
        total_long_term_liab_row = current_row
        ws.cell(row=current_row, column=1, value="Total Long-term Liabilities")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # TOTAL LIABILITIES
        current_row += 2
        total_liabilities_row = current_row
        ws.cell(row=current_row, column=1, value="TOTAL LIABILITIES")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{total_current_liab_row}+{col_letter}{total_long_term_liab_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # EQUITY
        current_row += 2
        ws.cell(row=current_row, column=1, value="EQUITY")
        self.style_section_row(ws, current_row, num_cols)
        
        current_row += 1
        equity_items = [
            'Share Capital', 'Retained Earnings', 'Current Year Profit', 'Reserves', 'Other Equity'
        ]
        
        start_row = current_row
        for item in equity_items:
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
            current_row += 1
        end_row = current_row - 1
        
        # TOTAL EQUITY
        total_equity_row = current_row
        ws.cell(row=current_row, column=1, value="TOTAL EQUITY")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # TOTAL LIABILITIES & EQUITY
        current_row += 2
        ws.cell(row=current_row, column=1, value="TOTAL LIABILITIES & EQUITY")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{total_liabilities_row}+{col_letter}{total_equity_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # Balance Check
        current_row += 2
        ws.cell(row=current_row, column=1, value="Balance Check (Should = 0)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{total_assets_row}-{col_letter}{current_row - 2}")
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        for col in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_income_statement(self):
        """Create Income Statement with formulas"""
        ws = self.wb.create_sheet("Income Statement")
        periods = self.generate_period_headers()
        num_cols = len(periods)
        
        # Title
        ws['A1'] = f"{self.company_name} - INCOME STATEMENT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        
        ws['A2'] = "For the Period Ended"
        ws['A2'].font = Font(italic=True, size=11)
        ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
        
        # Period headers
        current_row = 4
        for col_idx, period in enumerate(periods, start=1):
            ws.cell(row=current_row, column=col_idx, value=period)
        self.style_header_row(ws, current_row, num_cols)
        
        # REVENUE
        current_row += 1
        ws.cell(row=current_row, column=1, value="REVENUE")
        self.style_section_row(ws, current_row, num_cols)
        
        revenue_items = [
            'Product Sales', 'Service Revenue', 'Subscription Revenue',
            'Commission Income', 'Other Revenue', 'Sales Returns', 'Sales Discounts'
        ]
        
        start_row = current_row + 1
        for item in revenue_items:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Revenue
        current_row += 1
        total_revenue_row = current_row
        ws.cell(row=current_row, column=1, value="Total Revenue")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # COST OF GOODS SOLD
        current_row += 2
        ws.cell(row=current_row, column=1, value="COST OF GOODS SOLD")
        self.style_section_row(ws, current_row, num_cols)
        
        cogs_items = [
            'Opening Inventory', 'Purchases', 'Direct Labor',
            'Manufacturing Overhead', 'Closing Inventory'
        ]
        
        start_row = current_row + 1
        for item in cogs_items:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total COGS
        current_row += 1
        total_cogs_row = current_row
        ws.cell(row=current_row, column=1, value="Total COGS")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # GROSS PROFIT
        current_row += 2
        gross_profit_row = current_row
        ws.cell(row=current_row, column=1, value="GROSS PROFIT")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{total_revenue_row}-{col_letter}{total_cogs_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # OPERATING EXPENSES
        current_row += 2
        ws.cell(row=current_row, column=1, value="OPERATING EXPENSES")
        self.style_section_row(ws, current_row, num_cols)
        
        opex_items = [
            'Salaries and Wages', 'Employee Benefits', 'Rent', 'Utilities',
            'Insurance', 'Marketing and Advertising', 'Professional Fees',
            'Office Supplies', 'Depreciation', 'Amortization',
            'Repairs and Maintenance', 'Travel and Entertainment',
            'Bank Charges', 'Bad Debt Expense', 'Other Operating Expenses'
        ]
        
        start_row = current_row + 1
        for item in opex_items:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Total Operating Expenses
        current_row += 1
        total_opex_row = current_row
        ws.cell(row=current_row, column=1, value="Total Operating Expenses")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # OPERATING INCOME (EBIT)
        current_row += 2
        ebit_row = current_row
        ws.cell(row=current_row, column=1, value="OPERATING INCOME (EBIT)")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{gross_profit_row}-{col_letter}{total_opex_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # OTHER INCOME/(EXPENSE)
        current_row += 2
        ws.cell(row=current_row, column=1, value="OTHER INCOME/(EXPENSE)")
        self.style_section_row(ws, current_row, num_cols)
        
        other_items = [
            'Interest Income', 'Dividend Income', 'Gain on Sale of Assets',
            'Interest Expense', 'Loss on Sale of Assets', 'Foreign Exchange Loss',
            'Other Non-Operating Items'
        ]
        
        start_row = current_row + 1
        for item in other_items:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_row = current_row
        
        # Net Other Income/(Expense)
        current_row += 1
        net_other_row = current_row
        ws.cell(row=current_row, column=1, value="Net Other Income/(Expense)")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        self.style_total_row(ws, current_row, num_cols)
        
        # PROFIT BEFORE TAX
        current_row += 2
        ebt_row = current_row
        ws.cell(row=current_row, column=1, value="PROFIT BEFORE TAX (EBT)")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{ebit_row}+{col_letter}{net_other_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # Income Tax Expense
        current_row += 2
        ws.cell(row=current_row, column=1, value="Income Tax Expense")
        for col in range(2, num_cols + 1):
            ws.cell(row=current_row, column=col, value=0)
        tax_row = current_row
        
        # NET PROFIT AFTER TAX
        current_row += 2
        ws.cell(row=current_row, column=1, value="NET PROFIT AFTER TAX")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"={col_letter}{ebt_row}-{col_letter}{tax_row}")
        self.style_total_row(ws, current_row, num_cols)
        
        # KEY METRICS
        current_row += 3
        ws.cell(row=current_row, column=1, value="KEY METRICS")
        self.style_section_row(ws, current_row, num_cols)
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Gross Profit Margin %")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=IF({col_letter}{total_revenue_row}=0,0,{col_letter}{gross_profit_row}/{col_letter}{total_revenue_row}*100)")
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Operating Profit Margin %")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=IF({col_letter}{total_revenue_row}=0,0,{col_letter}{ebit_row}/{col_letter}{total_revenue_row}*100)")
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Net Profit Margin %")
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=current_row, column=col,
                   value=f"=IF({col_letter}{total_revenue_row}=0,0,{col_letter}{current_row - 1}/{col_letter}{total_revenue_row}*100)")
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        for col in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_cash_flow_statement(self):
        """Create Cash Flow Statement with formulas"""
        ws = self.wb.create_sheet("Cash Flow")
        periods = self.generate_period_headers()
        num_cols = len(periods)
        
        # Title
        ws['A1'] = f"{self.company_name} - CASH FLOW STATEMENT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        
        ws['A2'] = "For the Period Ended"
        ws['A2'].font = Font(italic=True, size=11)
        ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
        
        # Period headers
        current_row = 4
        for col_idx, period in enumerate(periods, start=1):
            ws.cell(row=current_row, column=col_idx, value=period)
        self.style_header_row(ws, current_row, num_cols)
        
        # OPERATING ACTIVITIES
        current_row += 1
        ws.cell(row=current_row, column=1, value="CASH FLOWS FROM OPERATING ACTIVITIES")
        self.style_section_row(ws, current_row, num_cols)
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Net Profit After Tax")
        for col in range(2, num_cols + 1):
            ws.cell(row=current_row, column=col, value=0)
        net_profit_row = current_row
        
        current_row += 2
        ws.cell(row=current_row, column=1, value="Adjustments for Non-Cash Items:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        adjustments = [
            'Add: Depreciation', 'Add: Amortization', 'Add: Loss on Sale of Assets',
            'Less: Gain on Sale of Assets', 'Add: Bad Debt Expense', 'Other Adjustments'
        ]
        
        start_adj_row = current_row + 1
        for item in adjustments:
            current_row += 1
            ws.cell(row=current_row, column=1, value=item)
            for col in range(2, num_cols + 1):
                ws.cell(row=current_row, column=col, value=0)
        end_adj_row = current_row
        
        current_row += 2
        ws.cell(row=current_row, column=1, value="Changes in Working Capital:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        working_capital = [
            '(Increase)/Decrease in Receivables',
            '(Increase)/Decrease in Inventory',
            '(Increase)/Decrease in Prepaid Expenses',
            'Increase/(Decrease) in Payables',
            'Increase/(Decrease) in Ac