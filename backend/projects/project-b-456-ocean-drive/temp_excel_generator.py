
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

# --- Data for Sheet 1: Budget Summary ---
budget_summary_data = {
    'Category': [
        'PRELIMINARIES', 'PRELIMINARIES', 'PRELIMINARIES', '',
        'DEMOLITION & SITE PREP', 'DEMOLITION & SITE PREP', '',
        'PARTITIONS & CEILINGS', 'PARTITIONS & CEILINGS', 'PARTITIONS & CEILINGS', '',
        'FLOORING', 'FLOORING', 'FLOORING', 'FLOORING', '',
        'JOINERY', 'JOINERY', 'JOINERY', 'JOINERY', 'JOINERY', '',
        'MECHANICAL (HVAC)', 'MECHANICAL (HVAC)', 'MECHANICAL (HVAC)', '',
        'ELECTRICAL & DATA', 'ELECTRICAL & DATA', 'ELECTRICAL & DATA', 'ELECTRICAL & DATA', '',
        'HYDRAULICS (PLUMBING)', 'HYDRAULICS (PLUMBING)', '',
        'FIRE SERVICES', 'FIRE SERVICES', '',
        'FINISHES', 'FINISHES'
    ],
    'Sub-Category': [
        'Site Establishment', 'Project Management Fees', 'Insurance & Permits', '',
        'Interior Demolition', 'Floor Grinding & Prep', '',
        'Steel Stud Framing', 'Plasterboard & Insulation', 'Suspended Ceilings', '',
        'Carpet Tiles - Office Area', 'Vinyl - Breakout Area', 'Epoxy - Server Room', 'Polished Concrete - Reception', '',
        'Reception Desk', 'Kitchen & Breakout Cabinets', 'Meeting Room Credenzas', 'Custom Wall Paneling', 'Storage Units', '',
        'AC Units & Ducting', 'Exhaust Fans', 'BMS & Controls', '',
        'Main Switchboard Upgrade', 'Lighting & Controls', 'Power Outlets', 'Data Cabling & Comms Rack', '',
        'Pipework Rough-in', 'Fixtures & Fit-off', '',
        'Sprinkler System', 'Smoke Detectors & Panel', '',
        'Painting', 'Glazing & Film'
    ],
    'Budget': [
        25000, 45000, 15000, '',
        35000, 10000, '',
        75000, 85000, 20000, '',
        45000, 25000, 15000, 10000, '',
        18000, 45000, 32000, 65000, 60000, '',
        90000, 25000, 35000, '',
        40000, 85000, 65000, 60000, '',
        35000, 40000, '',
        40000, 20000, '',
        55000, 35000
    ],
    'Actual': [
        22500, 37500, 16500, '',
        38500, 9500, '',
        60000, 70000, 18500, '',
        0, 5000, 0, 10000, '',
        5000, 10000, 0, 0, 15000, '',
        40000, 10000, 10000, '',
        42500, 30000, 25000, 20000, '',
        28000, 0, '',
        10000, 8000, '',
        0, 0
    ],
    'Committed': [
        2500, 7500, 0, '',
        0, 0, '',
        15000, 15000, 1500, '',
        45000, 20000, 15000, 0, '',
        13000, 35000, 32000, 65000, 45000, '',
        50000, 15000, 25000, '',
        0, 55000, 40000, 40000, '',
        7000, 40000, '',
        30000, 12000, '',
        55000, 35000
    ],
    'Forecast': [
        25000, 45000, 16500, '',
        38500, 9500, '',
        75000, 85000, 20000, '',
        45000, 25000, 15000, 10000, '',
        18000, 45000, 32000, 65000, 60000, '',
        95000, 25000, 35000, '',
        42500, 90000, 65000, 60000, '',
        35000, 40000, '',
        42000, 20000, '',
        55000, 35000
    ],
    'Variance': [0, 0, -1500, '', -3500, 500, '', 0, 0, 0, '', 0, 0, 0, 0, '', 0, 0, 0, 0, np.nan, '', -5000, 0, 0, '', -2500, -5000, 0, 0, '', 0, 0, '', -2000, 0, '', 0, 0],
    '% Complete': [90, 83, 100, '', 100, 100, '', 80, 82, 93, '', 0, 20, 0, 100, '', 28, 22, 0, 0, 25, '', 44, 40, 29, '', 100, 35, 38, 33, '', 80, 0, '', 25, 40, '', 0, 0],
    'Notes': [
        '', '', 'Council fees higher than expected', '',
        'Asbestos found - extra cost', '', '',
        '', 'check invoice #INV-DS-998', '', '',
        'Deposit paid', '', '', 'Complete', '',
        '', '', '', 'Long lead time item', 'Budget ???', '',
        'URGENT - Variation approved', '', '', '',
        '', 'Smart lighting VO approved', '', '', '',
        '', '', '',
        'Compliance issue found', '', '',
        '', ''
    ]
}
df_summary = pd.DataFrame(budget_summary_data)
df_summary.loc[20, 'Budget'] = '#REF!'

# --- Data for Sheet 2: Detailed Costs ---
detailed_costs_data = {
    'Date': [(datetime(2024, 7, 20) + timedelta(days=np.random.randint(0, 90))).strftime(np.random.choice(['%d/%m/%Y', '%m-%d-%y', '%Y/%m/%d'])) for _ in range(105)],
    'Supplier': np.random.choice(['Miami Demo Group', 'Drywall Solutions Inc.', 'Coastal Flooring', 'Millwork Masters', 'ACME Mechanical', 'Spark Electrical', 'Hydro Plumbers', 'FireSafe Systems', 'Pro Painters', 'Glass & Glazing Co.'], 105),
    'Invoice#': [f'INV-{np.random.randint(1000, 9999)}' for _ in range(105)],
    'PO#': [f'PO-{np.random.randint(100, 999)}' for _ in range(105)],
    'Description': ['' for _ in range(105)],
    'Category': np.random.choice(['PRELIMINARIES', 'DEMOLITION', 'PARTITIONS', 'FLOORING', 'JOINERY', 'MECHANICAL', 'ELECTRICAL', 'HYDRAULICS', 'FIRE', 'FINISHES'], 105),
    'Amount': [np.random.uniform(500, 5000) for _ in range(105)],
    'GST': [0 for _ in range(105)],
    'Total': [0 for _ in range(105)],
    'Paid?': np.random.choice(['Yes', 'No', 'PARTIAL'], 105),
    'Payment Date': [(datetime(2024, 8, 1) + timedelta(days=np.random.randint(0, 90))).strftime('%d/%m/%Y') if np.random.rand() > 0.3 else '' for _ in range(105)]
}
df_costs = pd.DataFrame(detailed_costs_data)
df_costs['GST'] = df_costs['Amount'] * 0.1
df_costs['Total'] = df_costs['Amount'] + df_costs['GST']
df_costs.loc[np.random.choice(df_costs.index, 20), 'Supplier'] = '' # Make some data missing

# --- Data for Sheet 3: Weekly Cashflow ---
cashflow_data = {
    'Week Ending': pd.to_datetime([datetime(2024, 7, 21) + timedelta(weeks=i) for i in range(20)]),
    'Opening Balance': [0]*20,
    'Money In': np.random.randint(0, 200000, 20),
    'Money Out': np.random.randint(10000, 150000, 20),
    'Closing Balance': [0]*20,
    'Notes': ['' for _ in range(20)]
}
df_cashflow = pd.DataFrame(cashflow_data)
df_cashflow.loc[5, 'Money In'] = 0 # Missing data
df_cashflow.loc[10, 'Notes'] = 'Client payment delayed'
df_cashflow.loc[15, 'Money Out'] = 250000 # Large payment
opening = 50000
for i, row in df_cashflow.iterrows():
    df_cashflow.loc[i, 'Opening Balance'] = opening
    closing = opening + row['Money In'] - row['Money Out']
    df_cashflow.loc[i, 'Closing Balance'] = closing
    opening = closing
df_cashflow.loc[8, 'Closing Balance'] = df_cashflow.loc[8, 'Closing Balance'] + 1000 # Manual error

# --- Data for Sheet 4: Notes & Calculations ---
notes_data = {
    'A': ['Notes & Calculations', '', 'Old Budget v1 (DO NOT USE)', 'Concrete Calculation', 'Contingency Check'],
    'B': ['', '', '~~Framing: $80,000~~', 'Slab Area (m2)', 850],
    'C': ['', '', '~~Joinery: $200,000~~', 'Rate ($/m3)', 250],
    'D': ['', '', '', 'Total Cost', '=B5*C5'],
    'E': ['', '', '', '', '']
}
df_notes = pd.DataFrame(notes_data)


# --- Create Excel File ---
file_path = "/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-b-456-ocean-drive/data/10_Budget_and_Cost_Control/MASTER_PROJECT_BUDGET.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='Budget Summary', index=False)
    df_costs.to_excel(writer, sheet_name='Detailed Costs', index=False)
    df_cashflow.to_excel(writer, sheet_name='Weekly Cashflow', index=False)
    df_notes.to_excel(writer, sheet_name='Notes & Calculations', index=False, header=False)

    # --- Apply Messy Formatting ---
    workbook = writer.book
    
    # Summary Sheet
    summary_sheet = writer.sheets['Budget Summary']
    summary_sheet.merge_cells('A1:I1')
    summary_sheet['A1'] = 'MASTER PROJECT BUDGET - Project B - 456 Ocean Drive'
    from openpyxl.styles import PatternFill, Font
    summary_sheet['A1'].font = Font(bold=True, size=16)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    summary_sheet['I5'].fill = red_fill
    summary_sheet['I6'].fill = green_fill
    summary_sheet.row_dimensions.group(10, 12, hidden=True) # Hidden rows
    summary_sheet.append(['','','','','','','','','Hidden rows 10-12'])

    # Notes Sheet
    notes_sheet = writer.sheets['Notes & Calculations']
    notes_sheet['C4'].font = Font(strike=True)
    notes_sheet['C5'].font = Font(strike=True)

print(f"Successfully created {file_path}")
