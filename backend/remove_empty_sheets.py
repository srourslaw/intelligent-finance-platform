#!/usr/bin/env python3
"""
Script to remove empty sheets from all Excel files in project-a-123-sunset-blvd
"""
import os
from pathlib import Path
import openpyxl

def is_sheet_empty(sheet):
    """Check if a worksheet is empty (no data in any cell)"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False
    return True

def remove_empty_sheets(file_path):
    """Remove empty sheets from an Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path)
        initial_sheet_count = len(wb.sheetnames)
        sheets_removed = []

        # Don't remove all sheets - keep at least one
        non_empty_sheets = [name for name in wb.sheetnames if not is_sheet_empty(wb[name])]

        if len(non_empty_sheets) == 0:
            print(f"  ⚠️  All sheets empty in {file_path.name}, keeping first sheet")
            return 0

        # Remove empty sheets
        for sheet_name in wb.sheetnames[:]:  # Use slice to avoid modifying list during iteration
            if is_sheet_empty(wb[sheet_name]) and len(wb.sheetnames) > 1:
                wb.remove(wb[sheet_name])
                sheets_removed.append(sheet_name)

        if sheets_removed:
            wb.save(file_path)
            print(f"  ✅ {file_path.name}")
            for sheet in sheets_removed:
                print(f"     - Removed: '{sheet}'")
            return len(sheets_removed)
        else:
            print(f"  ✓  {file_path.name} - No empty sheets")
            return 0

    except Exception as e:
        print(f"  ❌ Error processing {file_path.name}: {e}")
        return 0

def main():
    project_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd")

    # Find all Excel files
    excel_files = list(project_dir.rglob("*.xlsx"))

    print(f"Found {len(excel_files)} Excel files\n")
    print("Processing files...\n")

    total_removed = 0
    for file_path in sorted(excel_files):
        removed = remove_empty_sheets(file_path)
        total_removed += removed

    print(f"\n{'='*60}")
    print(f"Total empty sheets removed: {total_removed}")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
