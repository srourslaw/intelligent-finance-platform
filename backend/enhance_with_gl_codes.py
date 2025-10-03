#!/usr/bin/env python3
"""
Phase 2: Enhance Existing Files with GL Codes and Additional Data
Adds GL account mapping, AR/AP aging, revenue recognition, and opening balances
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
import random

FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
FILL_HEADER = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
FONT_BOLD = Font(bold=True, size=11)

def enhance_master_budget():
    """Add GL Account column to MASTER_PROJECT_BUDGET.xlsx"""
    print("\n📊 Enhancing MASTER_PROJECT_BUDGET.xlsx...")

    file_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/12_BUDGET_TRACKING/MASTER_PROJECT_BUDGET.xlsx")

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Budget Summary']

    # Add GL Account column after Category (column A)
    ws.insert_cols(2)
    ws.cell(4, 2, 'GL Account').font = FONT_HEADER
    ws.cell(4, 2).fill = FILL_HEADER

    # GL Account mapping based on categories
    gl_mapping = {
        'LAND': '6100',
        'DESIGN': '6200',
        'PERMITS': '6300',
        'FINANCE': '6500',
        'MATERIALS': '5100',
        'LABOR': '5200',
        'SUBCONTRACTORS': '5300',
        'EQUIPMENT': '5400',
        'SITE': '5500',
        'INSURANCE': '6400',
    }

    # Apply GL codes to rows
    for row in range(5, ws.max_row + 1):
        category = str(ws.cell(row, 1).value or '').upper()

        for key, gl_code in gl_mapping.items():
            if key in category:
                ws.cell(row, 2, gl_code)
                break

    wb.save(file_path)
    print("  ✅ Added GL Account column with mappings")

def enhance_client_payment_tracker():
    """Add revenue recognition columns to Client_Payment_Tracker.xlsx"""
    print("\n📊 Enhancing Client_Payment_Tracker.xlsx...")

    file_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/11_CLIENT_BILLING/Client_Payment_Tracker.xlsx")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find the last column
    max_col = ws.max_column

    # Add new columns
    new_headers = ['GL Account', '% Complete', 'Revenue Recognized', 'Deferred Revenue']
    for idx, header in enumerate(new_headers, max_col + 1):
        ws.cell(1, idx, header).font = FONT_HEADER
        ws.cell(1, idx).fill = FILL_HEADER

    # Populate GL Account and calculate revenue recognition
    for row in range(2, ws.max_row + 1):
        ws.cell(row, max_col + 1, '4100')  # GL Account: Construction Revenue

        # % Complete (random realistic values)
        pct_complete = random.choice([1.0, 1.0, 1.0, 0.8, 0.5])
        ws.cell(row, max_col + 2, pct_complete)
        ws.cell(row, max_col + 2).number_format = '0%'

        # Revenue Recognized = Milestone Value * % Complete
        milestone_col = 4  # Assuming column D has milestone value
        ws.cell(row, max_col + 3, f'=D{row}*{get_column_letter(max_col + 2)}{row}')
        ws.cell(row, max_col + 3).number_format = '$#,##0.00'

        # Deferred Revenue = Milestone Value - Revenue Recognized
        ws.cell(row, max_col + 4, f'=D{row}-{get_column_letter(max_col + 3)}{row}')
        ws.cell(row, max_col + 4).number_format = '$#,##0.00'

    wb.save(file_path)
    print("  ✅ Added revenue recognition columns (GL Account, % Complete, Revenue Recognized, Deferred Revenue)")

def enhance_invoices():
    """Add GL Account and aging to invoice files"""
    print("\n📊 Enhancing Invoice Files...")

    # Paid Invoices
    paid_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/06_PURCHASE_ORDERS_INVOICES/Invoices_Paid/Paid_Invoices_Register.xlsx")

    if paid_path.exists():
        wb = openpyxl.load_workbook(paid_path)
        ws = wb.active

        max_col = ws.max_column

        # Add columns
        ws.cell(1, max_col + 1, 'GL Account').font = FONT_HEADER
        ws.cell(1, max_col + 1).fill = FILL_HEADER
        ws.cell(1, max_col + 2, 'Category').font = FONT_HEADER
        ws.cell(1, max_col + 2).fill = FILL_HEADER

        categories = ['Materials', 'Subcontractors', 'Equipment', 'Services']
        gl_codes = {'Materials': '5100', 'Subcontractors': '5300', 'Equipment': '5400', 'Services': '6700'}

        for row in range(2, ws.max_row + 1):
            category = random.choice(categories)
            ws.cell(row, max_col + 2, category)
            ws.cell(row, max_col + 1, gl_codes[category])

        wb.save(paid_path)
        print("  ✅ Enhanced Paid_Invoices_Register.xlsx")

    # Pending Invoices
    pending_path = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/06_PURCHASE_ORDERS_INVOICES/Invoices_Pending/Pending_Invoices.xlsx")

    if pending_path.exists():
        wb = openpyxl.load_workbook(pending_path)
        ws = wb.active

        max_col = ws.max_column

        # Add columns
        new_cols = ['GL Account', 'Category', 'Payment Terms', 'Aging (Days)']
        for idx, header in enumerate(new_cols, max_col + 1):
            ws.cell(1, idx, header).font = FONT_HEADER
            ws.cell(1, idx).fill = FILL_HEADER

        for row in range(2, ws.max_row + 1):
            category = random.choice(categories)
            ws.cell(row, max_col + 2, category)
            ws.cell(row, max_col + 1, gl_codes[category])
            ws.cell(row, max_col + 3, random.choice(['Net 30', 'Net 60', 'Net 14', 'COD']))

            # Aging calculation (days since invoice date)
            # Assuming column C has invoice date
            ws.cell(row, max_col + 4, f'=TODAY()-C{row}')

        wb.save(pending_path)
        print("  ✅ Enhanced Pending_Invoices.xlsx")

def create_ar_aging_report():
    """Create Accounts Receivable Aging Report"""
    print("\n📊 Creating AR Aging Report...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Aging"

    # Title
    ws['A1'] = 'ACCOUNTS RECEIVABLE AGING REPORT'
    ws['A2'] = f'As at {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Headers
    headers = ['Customer', 'Invoice #', 'Invoice Date', 'Total', 'Current', '30 Days', '60 Days', '90+ Days']
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header).font = FONT_HEADER
        ws.cell(4, col).fill = FILL_HEADER

    # Sample AR data
    ar_data = [
        ['ABC Construction Ltd', 'INV-001', datetime(2024, 8, 15), 97500, 0, 97500, 0, 0],
        ['XYZ Developments', 'INV-002', datetime(2024, 9, 1), 65000, 65000, 0, 0, 0],
        ['Smith Property Group', 'INV-003', datetime(2024, 6, 20), 32500, 0, 0, 0, 32500],
        ['Metro Builders', 'INV-004', datetime(2024, 9, 15), 48750, 48750, 0, 0, 0],
    ]

    row = 5
    for data in ar_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row, col, value)
            if col == 3:  # Date
                cell.number_format = 'DD/MM/YYYY'
            elif col >= 4:  # Currency
                cell.number_format = '$#,##0.00'
        row += 1

    # Totals
    total_row = row
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    for col in range(4, 9):
        ws.cell(total_row, col, f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row-1})').font = FONT_BOLD
        ws.cell(total_row, col).number_format = '$#,##0.00'

    # Auto-size
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/11_CLIENT_BILLING")
    wb.save(output_dir / "AR_Aging_Report.xlsx")
    print("  ✅ Created AR_Aging_Report.xlsx")

def create_ap_aging_report():
    """Create Accounts Payable Aging Report"""
    print("\n📊 Creating AP Aging Report...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AP Aging"

    # Title
    ws['A1'] = 'ACCOUNTS PAYABLE AGING REPORT'
    ws['A2'] = f'As at {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:I1')

    # Headers
    headers = ['Vendor', 'Invoice #', 'Invoice Date', 'Total', 'Current', '30 Days', '60 Days', '90+ Days', 'Priority']
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header).font = FONT_HEADER
        ws.cell(4, col).fill = FILL_HEADER

    # Sample AP data
    ap_data = [
        ['BuildMart Supplies', 'BM-1234', datetime(2024, 9, 20), 15400, 15400, 0, 0, 0, 'NORMAL'],
        ['Spark Electrical', 'SE-5678', datetime(2024, 9, 15), 22080, 22080, 0, 0, 0, 'NORMAL'],
        ['Premium Plumbing', 'PP-9012', datetime(2024, 8, 25), 18900, 0, 18900, 0, 0, 'HIGH'],
        ['Concrete Co', 'CC-3456', datetime(2024, 9, 10), 31500, 31500, 0, 0, 0, 'NORMAL'],
        ['Sydney Tiles', 'ST-7890', datetime(2024, 7, 15), 12400, 0, 0, 12400, 0, 'URGENT'],
    ]

    row = 5
    for data in ap_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row, col, value)
            if col == 3:  # Date
                cell.number_format = 'DD/MM/YYYY'
            elif col >= 4 and col <= 8:  # Currency
                cell.number_format = '$#,##0.00'
        row += 1

    # Totals
    total_row = row
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    for col in range(4, 9):
        ws.cell(total_row, col, f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row-1})').font = FONT_BOLD
        ws.cell(total_row, col).number_format = '$#,##0.00'

    # Auto-size
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/06_PURCHASE_ORDERS_INVOICES")
    wb.save(output_dir / "AP_Aging_Report.xlsx")
    print("  ✅ Created AP_Aging_Report.xlsx")

def create_revenue_recognition_schedule():
    """Create Revenue Recognition Schedule with % of completion method"""
    print("\n📊 Creating Revenue Recognition Schedule...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Recognition"

    # Title
    ws['A1'] = 'REVENUE RECOGNITION SCHEDULE'
    ws['A2'] = 'Percentage of Completion Method'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Headers
    headers = ['Milestone', 'Contract Value', '% Complete (Physical)', '% Complete (Cost)',
               'Revenue Earned', 'Revenue Billed', 'Deferred Revenue', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header).font = FONT_HEADER
        ws.cell(4, col).fill = FILL_HEADER

    # Milestone data
    milestones = [
        ['Site Preparation', 32500, 1.0, 1.0, 'COMPLETE'],
        ['Foundation & Slab', 130000, 1.0, 1.0, 'COMPLETE'],
        ['Framing & Structure', 195000, 0.75, 0.72, 'IN PROGRESS'],
        ['Roof & Exterior', 97500, 0.40, 0.38, 'IN PROGRESS'],
        ['Internal Fit-out', 130000, 0.15, 0.12, 'STARTED'],
        ['Final Finishes', 65000, 0.0, 0.0, 'NOT STARTED'],
    ]

    row = 5
    for data in milestones:
        ws.cell(row, 1, data[0])  # Milestone
        ws.cell(row, 2, data[1])  # Contract Value
        ws.cell(row, 2).number_format = '$#,##0.00'

        ws.cell(row, 3, data[2])  # % Complete Physical
        ws.cell(row, 3).number_format = '0%'

        ws.cell(row, 4, data[3])  # % Complete Cost
        ws.cell(row, 4).number_format = '0%'

        # Revenue Earned = Contract Value * % Complete (Cost)
        ws.cell(row, 5, f'=B{row}*D{row}')
        ws.cell(row, 5).number_format = '$#,##0.00'

        # Revenue Billed (fully billed if complete, otherwise partial)
        if data[2] == 1.0:
            ws.cell(row, 6, data[1])
        else:
            ws.cell(row, 6, data[1] * data[2])
        ws.cell(row, 6).number_format = '$#,##0.00'

        # Deferred Revenue = Revenue Billed - Revenue Earned
        ws.cell(row, 7, f'=F{row}-E{row}')
        ws.cell(row, 7).number_format = '$#,##0.00'

        ws.cell(row, 8, data[4])  # Status

        row += 1

    # Totals
    total_row = row
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    for col in [2, 5, 6, 7]:
        ws.cell(total_row, col, f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{row-1})').font = FONT_BOLD
        ws.cell(total_row, col).number_format = '$#,##0.00'

    # Auto-size
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/11_CLIENT_BILLING")
    wb.save(output_dir / "Revenue_Recognition_Schedule.xlsx")
    print("  ✅ Created Revenue_Recognition_Schedule.xlsx")

def create_opening_balances():
    """Create Opening Balances document"""
    print("\n📊 Creating Opening Balances...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Balances"

    # Title
    ws['A1'] = 'OPENING BALANCES'
    ws['A2'] = 'As at 1 June 2024 (Start of Project)'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')

    # Headers
    headers = ['GL Account', 'Account Name', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header).font = FONT_HEADER
        ws.cell(4, col).fill = FILL_HEADER

    # Opening balances
    balances = [
        ['1100', 'Cash at Bank', 100000, ''],
        ['1400', 'Prepaid Expenses', 5000, ''],
        ['3100', "Owner's Capital", '', 100000],
        ['3200', 'Retained Earnings', '', 5000],
    ]

    row = 5
    for data in balances:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row, col, value)
            if col in [3, 4] and value != '':
                cell.number_format = '$#,##0.00'
        row += 1

    # Totals
    total_row = row + 1
    ws.cell(total_row, 1, 'TOTAL').font = FONT_BOLD
    ws.cell(total_row, 3, f'=SUM(C5:C{row-1})').font = FONT_BOLD
    ws.cell(total_row, 3).number_format = '$#,##0.00'
    ws.cell(total_row, 4, f'=SUM(D5:D{row-1})').font = FONT_BOLD
    ws.cell(total_row, 4).number_format = '$#,##0.00'

    # Balance check
    ws.cell(total_row + 2, 1, 'Balance Check:').font = FONT_BOLD
    ws.cell(total_row + 2, 2, f'=IF(C{total_row}=D{total_row},"✓ BALANCED","✗ OUT OF BALANCE")')
    ws.cell(total_row + 2, 2).font = Font(bold=True, color='00B050')

    # Auto-size
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25

    output_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/17_GENERAL_LEDGER")
    wb.save(output_dir / "Opening_Balances_June_2024.xlsx")
    print("  ✅ Created Opening_Balances_June_2024.xlsx")

def main():
    print("=" * 80)
    print("PHASE 2: ENHANCE EXISTING FILES WITH GL CODES")
    print("=" * 80)

    enhance_master_budget()
    enhance_client_payment_tracker()
    enhance_invoices()
    create_ar_aging_report()
    create_ap_aging_report()
    create_revenue_recognition_schedule()
    create_opening_balances()

    print("\n" + "=" * 80)
    print("✨ PHASE 2 COMPLETE!")
    print("=" * 80)
    print("\nEnhancements Applied:")
    print("  ✅ MASTER_PROJECT_BUDGET.xlsx - Added GL Account column")
    print("  ✅ Client_Payment_Tracker.xlsx - Added revenue recognition columns")
    print("  ✅ Paid_Invoices_Register.xlsx - Added GL Account and Category")
    print("  ✅ Pending_Invoices.xlsx - Added GL Account, Category, Payment Terms, Aging")
    print("\nNew Files Created:")
    print("  ✅ AR_Aging_Report.xlsx")
    print("  ✅ AP_Aging_Report.xlsx")
    print("  ✅ Revenue_Recognition_Schedule.xlsx")
    print("  ✅ Opening_Balances_June_2024.xlsx")
    print("=" * 80 + "\n")

if __name__ == "__main__":
    main()
