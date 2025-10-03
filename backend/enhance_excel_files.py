#!/usr/bin/env python3
"""
Comprehensive Excel Enhancement Script
Adds realistic construction project data, charts, and formulas to all Excel files
"""
import openpyxl
from openpyxl.chart import (
    PieChart, BarChart, LineChart, Reference,
    Series, DoughnutChart, ScatterChart
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
from pathlib import Path

# Color schemes
COLORS_PRIMARY = ['1F77B4', 'FF7F0E', '2CA02C', 'D62728', '9467BD', '8C564B']
FILL_HEADER = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
FONT_TOTAL = Font(bold=True, size=11)

def enhance_master_budget(file_path):
    """Enhance MASTER_PROJECT_BUDGET.xlsx with dashboard, charts, and additional data"""
    print(f"\n📊 Enhancing: {file_path.name}")

    wb = openpyxl.load_workbook(file_path)

    # Add Dashboard sheet at the beginning
    if 'Dashboard' not in wb.sheetnames:
        dashboard = wb.create_sheet('Dashboard', 0)

        # Dashboard title
        dashboard['A1'] = 'PROJECT BUDGET DASHBOARD'
        dashboard['A1'].font = Font(size=18, bold=True, color='1F4E78')
        dashboard.merge_cells('A1:F1')

        # KPI Cards
        dashboard['A3'] = 'Total Contract Value'
        dashboard['B3'] = 650000
        dashboard['B3'].number_format = '$#,##0'

        dashboard['D3'] = 'Total Spent'
        dashboard['E3'] = '=\'Budget Summary\'!E85'
        dashboard['E3'].number_format = '$#,##0'

        dashboard['A4'] = 'Budget Remaining'
        dashboard['B4'] = '=B3-E3'
        dashboard['B4'].number_format = '$#,##0'

        dashboard['D4'] = '% Complete'
        dashboard['E4'] = '=E3/B3'
        dashboard['E4'].number_format = '0%'

        # Add pie chart for budget allocation
        pie = PieChart()
        pie.title = "Budget Allocation by Category"
        pie.style = 10
        pie.height = 12
        pie.width = 20

        # Reference data from Budget Summary sheet
        labels = Reference(wb['Budget Summary'], min_col=1, min_row=5, max_row=15)
        data = Reference(wb['Budget Summary'], min_col=3, min_row=4, max_row=15)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        dashboard.add_chart(pie, "A6")

        # Add bar chart for budget vs actual
        bar = BarChart()
        bar.type = "col"
        bar.title = "Budget vs Actual Spending"
        bar.y_axis.title = 'Amount ($)'
        bar.x_axis.title = 'Category'
        bar.height = 12
        bar.width = 20

        cats = Reference(wb['Budget Summary'], min_col=1, min_row=5, max_row=15)
        budget_data = Reference(wb['Budget Summary'], min_col=3, min_row=4, max_row=15)
        actual_data = Reference(wb['Budget Summary'], min_col=4, min_row=4, max_row=15)

        bar.add_data(budget_data, titles_from_data=True)
        bar.add_data(actual_data)
        bar.set_categories(cats)

        dashboard.add_chart(bar, "L6")

        print("  ✅ Added Dashboard with KPIs and 2 charts")

    # Enhance Cashflow sheet with more data
    if 'Cashflow' in wb.sheetnames:
        cf_sheet = wb['Cashflow']
        start_date = datetime(2024, 1, 22)

        # Add more weekly data (extend to 52 weeks)
        current_row = 3
        balance = 100000

        for week in range(52):
            week_date = start_date + timedelta(weeks=week)
            cash_in = random.randint(50000, 150000) if week > 0 else 0
            cash_out = random.randint(30000, 80000)
            net = cash_in - cash_out
            balance += net

            cf_sheet[f'A{current_row}'] = week_date
            cf_sheet[f'B{current_row}'] = cash_in
            cf_sheet[f'C{current_row}'] = cash_out
            cf_sheet[f'D{current_row}'] = f'=B{current_row}-C{current_row}'
            cf_sheet[f'E{current_row}'] = f'=E{current_row-1}+D{current_row}' if current_row > 3 else balance

            # Format as currency
            for col in ['B', 'C', 'D', 'E']:
                cf_sheet[f'{col}{current_row}'].number_format = '$#,##0'

            current_row += 1

        # Add line chart for cashflow
        line = LineChart()
        line.title = "Cashflow Projection"
        line.style = 12
        line.y_axis.title = 'Balance ($)'
        line.x_axis.title = 'Week'
        line.height = 12
        line.width = 20

        dates = Reference(cf_sheet, min_col=1, min_row=3, max_row=current_row-1)
        balance_data = Reference(cf_sheet, min_col=5, min_row=2, max_row=current_row-1)

        line.add_data(balance_data, titles_from_data=True)
        line.set_categories(dates)

        cf_sheet.add_chart(line, "G3")

        print(f"  ✅ Extended Cashflow to 52 weeks + added line chart")

    # Add Cost Detail transactions
    if 'Cost Detail' in wb.sheetnames:
        detail_sheet = wb['Cost Detail']
        current_row = 13

        # Add 50 more realistic transactions
        suppliers = ['Smith Family Trust', 'BuildMart', 'Spark Electrical', 'Premium Plumbing',
                    'Ace Carpentry', 'Pro Paint', 'Sydney Tiles', 'Concrete Co', 'Steel Supplies']
        categories = ['Land', 'Materials', 'Labour', 'Subcontractors', 'Equipment']

        start_date = datetime(2024, 1, 15)

        for i in range(50):
            trans_date = start_date + timedelta(days=random.randint(1, 180))
            supplier = random.choice(suppliers)
            invoice = f"INV-{2024000 + i}"
            desc = f"Construction materials and services - {supplier}"
            category = random.choice(categories)
            amount = random.randint(1000, 50000)
            gst = amount * 0.1
            total = amount + gst
            status = random.choice(['PAID', 'PAID', 'PAID', 'OUTSTANDING'])

            detail_sheet[f'A{current_row}'] = trans_date
            detail_sheet[f'B{current_row}'] = supplier
            detail_sheet[f'C{current_row}'] = invoice
            detail_sheet[f'D{current_row}'] = desc
            detail_sheet[f'E{current_row}'] = category
            detail_sheet[f'F{current_row}'] = amount
            detail_sheet[f'G{current_row}'] = f'=F{current_row}*0.1'
            detail_sheet[f'H{current_row}'] = f'=F{current_row}+G{current_row}'
            detail_sheet[f'I{current_row}'] = status

            # Format
            detail_sheet[f'A{current_row}'].number_format = 'DD/MM/YYYY'
            for col in ['F', 'G', 'H']:
                detail_sheet[f'{col}{current_row}'].number_format = '$#,##0.00'

            current_row += 1

        print(f"  ✅ Added 50 transactions to Cost Detail")

    wb.save(file_path)
    print(f"  💾 Saved {file_path.name}\n")

def enhance_purchase_orders(file_path):
    """Enhance Purchase_Orders_Master.xlsx"""
    print(f"\n📦 Enhancing: {file_path.name}")

    wb = openpyxl.load_workbook(file_path)

    if 'PO Register' in wb.sheetnames:
        po_sheet = wb['PO Register']
        current_row = 33

        # Add 70 more POs
        suppliers = ['Spark Electrical', 'Premium Plumbing', 'Ace Carpentry', 'BuildMart',
                    'Pro Paint', 'Sydney Tiles', 'Concrete Co', 'Steel Supplies', 'Timber Traders']
        statuses = ['COMPLETE', 'COMPLETE', 'COMPLETE', 'PENDING', 'CANCELLED']

        start_date = datetime(2024, 1, 10)

        for i in range(70):
            po_num = f'PO-2024-{1031 + i}'
            po_date = start_date + timedelta(days=random.randint(1, 200))
            supplier = random.choice(suppliers)
            desc = f"Materials and services for construction - Phase {random.randint(1,5)}"
            amount = random.randint(5000, 80000)
            delivery = po_date + timedelta(days=random.randint(7, 30))
            invoice_recv = random.choice(['YES', 'YES', 'YES', 'NO'])
            status = random.choice(statuses)

            po_sheet[f'A{current_row}'] = po_num
            po_sheet[f'B{current_row}'] = po_date
            po_sheet[f'C{current_row}'] = supplier
            po_sheet[f'D{current_row}'] = desc
            po_sheet[f'E{current_row}'] = amount
            po_sheet[f'F{current_row}'] = delivery
            po_sheet[f'G{current_row}'] = invoice_recv
            po_sheet[f'H{current_row}'] = status

            # Format
            po_sheet[f'B{current_row}'].number_format = 'DD/MM/YYYY'
            po_sheet[f'E{current_row}'].number_format = '$#,##0'
            po_sheet[f'F{current_row}'].number_format = 'DD/MM/YYYY'

            current_row += 1

        # Add summary chart
        pie = PieChart()
        pie.title = "PO Status Distribution"
        pie.style = 10
        pie.height = 12
        pie.width = 16

        # Create summary data
        po_sheet['J2'] = 'Status'
        po_sheet['K2'] = 'Count'
        po_sheet['J3'] = 'COMPLETE'
        po_sheet['J4'] = 'PENDING'
        po_sheet['J5'] = 'CANCELLED'
        po_sheet['K3'] = f'=COUNTIF(H:H,"COMPLETE")'
        po_sheet['K4'] = f'=COUNTIF(H:H,"PENDING")'
        po_sheet['K5'] = f'=COUNTIF(H:H,"CANCELLED")'

        labels = Reference(po_sheet, min_col=10, min_row=3, max_row=5)
        data = Reference(po_sheet, min_col=11, min_row=2, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        po_sheet.add_chart(pie, "J7")

        print(f"  ✅ Added 70 POs + pie chart")

    wb.save(file_path)
    print(f"  💾 Saved {file_path.name}\n")

def enhance_subcontractor_register(file_path):
    """Enhance Subcontractor_Register.xlsx"""
    print(f"\n👷 Enhancing: {file_path.name}")

    wb = openpyxl.load_workbook(file_path)

    if 'Active Subbies' in wb.sheetnames:
        sub_sheet = wb['Active Subbies']
        current_row = 18

        # Add 20 more subcontractors
        companies = ['Elite Bricklaying', 'Master Tilers', 'Precision Welding', 'Quality Roofing',
                    'Apex Scaffolding', 'Summit Concreting', 'Peak Glazing', 'Prime Landscaping',
                    'Expert Waterproofing', 'Advanced HVAC', 'Superior Flooring']

        start_date = datetime(2024, 2, 1)

        for i in range(20):
            sub_id = f'SUB-{17 + i:03d}'
            company = random.choice(companies)
            contact = random.choice(['John S', 'Sarah M', 'Mike D', 'Lisa T', 'Tom R'])
            phone = f'04{random.randint(10000000, 99999999)}'
            email = f'{company.lower().replace(" ", "")}@example.com'
            abn = f'{random.randint(10000000000, 99999999999)}'
            license = f'LIC-{random.randint(100000, 999999)}'
            insurance_exp = start_date + timedelta(days=random.randint(180, 720))
            contract_val = random.randint(10000, 150000)
            status = random.choice(['COMPLETE', 'COMPLETE', 'IN PROGRESS', 'PENDING'])

            sub_sheet[f'A{current_row}'] = sub_id
            sub_sheet[f'B{current_row}'] = company
            sub_sheet[f'C{current_row}'] = contact
            sub_sheet[f'D{current_row}'] = phone
            sub_sheet[f'E{current_row}'] = email
            sub_sheet[f'F{current_row}'] = abn
            sub_sheet[f'G{current_row}'] = license
            sub_sheet[f'H{current_row}'] = insurance_exp
            sub_sheet[f'I{current_row}'] = contract_val
            sub_sheet[f'J{current_row}'] = status

            # Format
            sub_sheet[f'H{current_row}'].number_format = 'DD/MM/YYYY'
            sub_sheet[f'I{current_row}'].number_format = '$#,##0'

            current_row += 1

        # Add bar chart for contract values
        bar = BarChart()
        bar.type = "col"
        bar.title = "Subcontractor Contract Values"
        bar.y_axis.title = 'Contract Value ($)'
        bar.height = 15
        bar.width = 24

        companies = Reference(sub_sheet, min_col=2, min_row=2, max_row=current_row-1)
        values = Reference(sub_sheet, min_col=9, min_row=1, max_row=current_row-1)

        bar.add_data(values, titles_from_data=True)
        bar.set_categories(companies)

        sub_sheet.add_chart(bar, "L2")

        print(f"  ✅ Added 20 subcontractors + bar chart")

    wb.save(file_path)
    print(f"  💾 Saved {file_path.name}\n")

def enhance_generic_excel(file_path):
    """Generic enhancement for other Excel files - add data and basic chart"""
    print(f"\n📄 Enhancing: {file_path.name}")

    try:
        wb = openpyxl.load_workbook(file_path)
        first_sheet = wb.worksheets[0]

        # Find last row with data
        last_row = first_sheet.max_row

        # Only enhance if less than 30 rows
        if last_row < 30:
            # Add 20-40 more rows of similar data
            rows_to_add = random.randint(20, 40)

            # Get column types from existing data
            for i in range(rows_to_add):
                new_row = last_row + 1 + i

                # Copy formulas from row above
                for col in range(1, first_sheet.max_column + 1):
                    source_cell = first_sheet.cell(last_row, col)
                    target_cell = first_sheet.cell(new_row, col)

                    if source_cell.data_type == 'f':  # Formula
                        # Adjust formula for new row
                        target_cell.value = source_cell.value
                    elif source_cell.data_type == 'd':  # Date
                        target_cell.value = datetime.now() + timedelta(days=random.randint(1, 180))
                        target_cell.number_format = source_cell.number_format
                    elif source_cell.data_type == 'n':  # Number
                        target_cell.value = random.randint(1000, 50000)
                        target_cell.number_format = source_cell.number_format
                    else:  # String
                        target_cell.value = f"Item {new_row - 1}"

            print(f"  ✅ Added {rows_to_add} rows")

        wb.save(file_path)
        print(f"  💾 Saved {file_path.name}\n")

    except Exception as e:
        print(f"  ⚠️  Error enhancing {file_path.name}: {e}\n")

def main():
    project_dir = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data")

    print("=" * 80)
    print("EXCEL ENHANCEMENT SCRIPT")
    print("=" * 80)

    # Priority files - full enhancement
    priority_files = {
        '12_BUDGET_TRACKING/MASTER_PROJECT_BUDGET.xlsx': enhance_master_budget,
        '06_PURCHASE_ORDERS_INVOICES/Purchase_Orders_Master.xlsx': enhance_purchase_orders,
        '07_SUBCONTRACTORS/Subcontractor_Register.xlsx': enhance_subcontractor_register,
    }

    print("\n🎯 PHASE 1: Enhancing Priority Files")
    print("-" * 80)

    for file_rel_path, enhance_func in priority_files.items():
        file_path = project_dir / file_rel_path
        if file_path.exists():
            enhance_func(file_path)

    print("\n🎯 PHASE 2: Enhancing All Other Files")
    print("-" * 80)

    # Get all Excel files except priority ones
    all_excel_files = list(project_dir.rglob("*.xlsx"))
    priority_names = [Path(p).name for p in priority_files.keys()]
    other_files = [f for f in all_excel_files if f.name not in priority_names]

    for file_path in sorted(other_files):
        enhance_generic_excel(file_path)

    print("\n" + "=" * 80)
    print("✨ ENHANCEMENT COMPLETE!")
    print("=" * 80)
    print(f"Total files processed: {len(all_excel_files)}")
    print("=" * 80 + "\n")

if __name__ == "__main__":
    main()
