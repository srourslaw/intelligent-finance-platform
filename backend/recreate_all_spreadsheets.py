"""
Recreate ALL 74 Excel spreadsheets with logical structure and meaningful data.
This script processes each folder and creates realistic, dashboard-ready Excel files.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os

# Color schemes
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

BASE_PATH = "/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data"

def apply_header_style(ws, row=1):
    """Apply professional header styling"""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def apply_borders_to_range(ws, start_row, end_row, start_col, end_col):
    """Apply borders to a range of cells"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = THIN_BORDER

def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def random_date(start_month=6, end_month=10):
    """Generate random date in 2024"""
    month = random.randint(start_month, end_month)
    day = random.randint(1, 28)
    return f"2024-{month:02d}-{day:02d}"

# ============================================================================
# 02_PERMITS_APPROVALS
# ============================================================================

def create_development_permit_tracking():
    """02_PERMITS_APPROVALS/Development_Permit_Tracking.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Permit Tracking"

    headers = ['Permit Type', 'Authority', 'Application Date', 'Approval Date',
               'Permit Number', 'Fee (Ex GST)', 'GST', 'Total Fee', 'Status', 'Expiry Date']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Development Application', 'City of Sydney Council', '2024-03-15', '2024-05-20', 'DA-2024-1234', 8500, 850, 9350, 'APPROVED', '2025-05-20'],
        ['Construction Certificate', 'Private Certifier - BuildCert', '2024-05-25', '2024-06-10', 'CC-2024-5678', 3200, 320, 3520, 'APPROVED', '2025-06-10'],
        ['Demolition Permit', 'City of Sydney Council', '2024-06-01', '2024-06-05', 'DEM-2024-0912', 850, 85, 935, 'APPROVED', '2024-12-05'],
        ['Road Opening Permit', 'Sydney Water', '2024-06-15', '2024-06-20', 'ROP-2024-3456', 1200, 120, 1320, 'APPROVED', '2024-09-20'],
        ['Hoarding Permit', 'City of Sydney Council', '2024-06-10', '2024-06-18', 'HP-2024-7890', 650, 65, 715, 'APPROVED', '2024-12-18'],
        ['Asbestos Removal License', 'SafeWork NSW', '2024-05-15', '2024-05-22', 'ARL-2024-1122', 1100, 110, 1210, 'APPROVED', '2027-05-22'],
        ['Crane Permit', 'City of Sydney Council', '2024-07-01', '2024-07-05', 'CP-2024-3344', 2400, 240, 2640, 'APPROVED', '2024-11-05'],
        ['Occupation Certificate', 'Private Certifier - BuildCert', '2024-09-15', 'PENDING', 'OC-2024-PEND', 2800, 280, 3080, 'IN PROGRESS', 'TBD'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [6, 7, 8]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL FEES:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 7, f"=SUM(G2:G{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H2:H{ws.max_row})")
    for col in [6, 7, 8]:
        ws.cell(total_row, col).font = TOTAL_FONT
        ws.cell(total_row, col).fill = TOTAL_FILL
        ws.cell(total_row, col).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 10)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "02_PERMITS_APPROVALS", "Development_Permit_Tracking.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 03_DESIGN_DRAWINGS
# ============================================================================

def create_architect_fees_breakdown():
    """03_DESIGN_DRAWINGS/Architect_Fees_Breakdown.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Architect Fees"

    headers = ['Design Phase', 'Consultant', 'Description', 'Hours', 'Rate per Hour',
               'Subtotal', 'GST', 'Total', 'Invoice Date', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Concept Design', 'Smith Architecture', 'Initial concept and sketches', 40, 180, 7200, 720, 7920, '2024-03-30', 'PAID'],
        ['Schematic Design', 'Smith Architecture', 'Detailed floor plans and elevations', 65, 180, 11700, 1170, 12870, '2024-04-25', 'PAID'],
        ['Design Development', 'Smith Architecture', 'Material selection and specifications', 55, 180, 9900, 990, 10890, '2024-05-15', 'PAID'],
        ['Construction Documentation', 'Smith Architecture', 'Final construction drawings', 80, 180, 14400, 1440, 15840, '2024-06-05', 'PAID'],
        ['Engineering Coordination', 'StructEng Partners', 'Structural engineering review', 30, 220, 6600, 660, 7260, '2024-05-20', 'PAID'],
        ['3D Renderings', 'Viz Studio', 'Photorealistic renders for marketing', 20, 150, 3000, 300, 3300, '2024-04-10', 'PAID'],
        ['Site Inspections', 'Smith Architecture', 'Construction phase inspections (6 visits)', 18, 180, 3240, 324, 3564, '2024-08-15', 'INVOICED'],
        ['As-Built Documentation', 'Smith Architecture', 'Final as-built drawings', 25, 180, 4500, 450, 4950, '2024-10-01', 'PENDING'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Add formulas for calculated columns
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 6, f"=D{row}*E{row}")  # Subtotal
        ws.cell(row, 7, f"=F{row}*0.1")      # GST
        ws.cell(row, 8, f"=F{row}+G{row}")   # Total

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [5, 6, 7, 8]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL DESIGN FEES:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 7, f"=SUM(G2:G{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H2:H{ws.max_row})")
    for col in [6, 7, 8]:
        ws.cell(total_row, col).font = TOTAL_FONT
        ws.cell(total_row, col).fill = TOTAL_FILL
        ws.cell(total_row, col).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 10)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "03_DESIGN_DRAWINGS", "Architect_Fees_Breakdown.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_structural_engineer_invoices():
    """03_DESIGN_DRAWINGS/Structural_Engineer_Invoices.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structural Engineering"

    headers = ['Service Type', 'Engineer', 'Invoice Number', 'Date', 'Description',
               'Amount (Ex GST)', 'GST', 'Total', 'Payment Date', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Structural Analysis', 'StructEng Partners', 'SE-2024-001', '2024-04-15', 'Load bearing analysis and calculations', 8500, 850, 9350, '2024-05-01', 'PAID'],
        ['Foundation Design', 'StructEng Partners', 'SE-2024-002', '2024-05-10', 'Footing and slab design', 6200, 620, 6820, '2024-05-25', 'PAID'],
        ['Steel Frame Design', 'StructEng Partners', 'SE-2024-003', '2024-05-20', 'Structural steel specifications', 7800, 780, 8580, '2024-06-05', 'PAID'],
        ['Site Inspections', 'StructEng Partners', 'SE-2024-004', '2024-07-15', 'Foundation inspection (3 visits)', 2400, 240, 2640, '2024-07-30', 'PAID'],
        ['Certification', 'StructEng Partners', 'SE-2024-005', '2024-09-20', 'Structural certification for handover', 3500, 350, 3850, '2024-10-05', 'INVOICED'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [6, 7, 8]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 5, "TOTAL ENGINEERING FEES:")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 6, f"=SUM(F2:F{ws.max_row})")
    ws.cell(total_row, 7, f"=SUM(G2:G{ws.max_row})")
    ws.cell(total_row, 8, f"=SUM(H2:H{ws.max_row})")
    for col in [6, 7, 8]:
        ws.cell(total_row, col).font = TOTAL_FONT
        ws.cell(total_row, col).fill = TOTAL_FILL
        ws.cell(total_row, col).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 10)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "03_DESIGN_DRAWINGS", "Structural_Engineer_Invoices.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 04_FINANCE_INSURANCE
# ============================================================================

def create_construction_loan_details():
    """04_FINANCE_INSURANCE/Construction_Loan_Details.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Details"

    # Loan Summary
    ws['A1'] = "CONSTRUCTION LOAN SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    summary_data = [
        ['Lender:', 'Commonwealth Bank of Australia'],
        ['Loan Amount:', 600000],
        ['Interest Rate:', '6.25%'],
        ['Loan Term:', '12 months (construction phase)'],
        ['Facility Fee:', 1500],
        ['Application Fee:', 850],
        ['Loan Start Date:', '2024-06-15'],
        ['First Drawdown:', '2024-06-20'],
        ['Monthly Interest (approx):', 3125],
    ]

    row = 3
    for item in summary_data:
        ws.cell(row, 1, item[0]).font = Font(bold=True)
        ws.cell(row, 2, item[1])
        if isinstance(item[1], (int, float)) and item[0] not in ['Loan Term:', 'Loan Start Date:', 'First Drawdown:']:
            if 'Rate' in item[0] or '%' in str(item[1]):
                ws.cell(row, 2).number_format = '0.00%'
            else:
                ws.cell(row, 2).number_format = '$#,##0.00'
        row += 1

    # Drawdown Schedule
    ws.cell(row + 1, 1, "DRAWDOWN SCHEDULE").font = Font(bold=True, size=12)
    row += 2

    headers = ['Drawdown #', 'Date', 'Purpose', 'Amount Drawn', 'Cumulative Drawn', 'Remaining Available']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    drawdowns = [
        [1, '2024-06-20', 'Land purchase and stamp duty', 260000, 260000, 340000],
        [2, '2024-07-05', 'Demolition and site preparation', 45000, 305000, 295000],
        [3, '2024-07-20', 'Foundation and slab', 85000, 390000, 210000],
        [4, '2024-08-10', 'Framing and structural steel', 95000, 485000, 115000],
        [5, '2024-09-01', 'Roofing and external cladding', 65000, 550000, 50000],
        [6, '2024-09-20', 'Final payment (projected)', 50000, 600000, 0],
    ]

    row += 1
    start_data_row = row
    for dd in drawdowns:
        for col, value in enumerate(dd, 1):
            cell = ws.cell(row, col, value)
            cell.border = THIN_BORDER
            if col > 3:
                cell.number_format = '$#,##0.00'
        row += 1

    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "04_FINANCE_INSURANCE", "Construction_Loan_Details.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_insurance_policies():
    """04_FINANCE_INSURANCE/Insurance_Policies.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insurance Policies"

    headers = ['Policy Type', 'Insurer', 'Policy Number', 'Coverage Amount',
               'Premium (Annual)', 'Start Date', 'End Date', 'Status', 'Broker']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Contract Works Insurance', 'QBE Insurance', 'CW-2024-789456', 800000, 4800, '2024-06-15', '2025-06-15', 'ACTIVE', 'Metro Insurance Brokers'],
        ['Public Liability', 'Allianz', 'PL-2024-123789', 20000000, 3200, '2024-06-15', '2025-06-15', 'ACTIVE', 'Metro Insurance Brokers'],
        ['Professional Indemnity', 'Chubb', 'PI-2024-456123', 5000000, 2800, '2024-06-15', '2025-06-15', 'ACTIVE', 'Metro Insurance Brokers'],
        ['Home Building Compensation', 'icare NSW', 'HBC-2024-987654', 600000, 1100, '2024-06-01', '2030-06-01', 'ACTIVE', 'Direct'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = '$#,##0'
        ws.cell(row, 5).number_format = '$#,##0.00'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 4, "TOTAL ANNUAL PREMIUMS:")
    ws.cell(total_row, 4).font = TOTAL_FONT
    ws.cell(total_row, 5, f"=SUM(E2:E{ws.max_row})")
    ws.cell(total_row, 5).font = TOTAL_FONT
    ws.cell(total_row, 5).fill = TOTAL_FILL
    ws.cell(total_row, 5).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 9)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "04_FINANCE_INSURANCE", "Insurance_Policies.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_interest_expense_tracking():
    """04_FINANCE_INSURANCE/Interest_Expense_Tracking.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interest Expense"

    headers = ['Month', 'Opening Balance', 'Drawdowns', 'Closing Balance',
               'Interest Rate', 'Days in Month', 'Interest Charged', 'Payment Date', 'Status']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['June 2024', 0, 260000, 260000, 0.0625, 10, 445.21, '2024-07-05', 'PAID'],
        ['July 2024', 260000, 130000, 390000, 0.0625, 31, 2066.78, '2024-08-05', 'PAID'],
        ['August 2024', 390000, 95000, 485000, 0.0625, 31, 2538.36, '2024-09-05', 'PAID'],
        ['September 2024', 485000, 115000, 600000, 0.0625, 30, 3062.50, '2024-10-05', 'INVOICED'],
        ['October 2024 (proj)', 600000, 0, 600000, 0.0625, 31, 3184.93, '2024-11-05', 'PENDING'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [2, 3, 4, 7]:
            ws.cell(row, col).number_format = '$#,##0.00'
        ws.cell(row, 5).number_format = '0.00%'

    # Add total row
    total_row = ws.max_row + 2
    ws.cell(total_row, 6, "TOTAL INTEREST:")
    ws.cell(total_row, 6).font = TOTAL_FONT
    ws.cell(total_row, 7, f"=SUM(G2:G{ws.max_row})")
    ws.cell(total_row, 7).font = TOTAL_FONT
    ws.cell(total_row, 7).fill = TOTAL_FILL
    ws.cell(total_row, 7).number_format = '$#,##0.00'

    apply_borders_to_range(ws, 1, ws.max_row, 1, 9)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "04_FINANCE_INSURANCE", "Interest_Expense_Tracking.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# 05_QUOTES_ESTIMATES
# ============================================================================

def create_concrete_supplier_quotes():
    """05_QUOTES_ESTIMATES/Concrete_Supplier_Quotes.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concrete Quotes"

    headers = ['Supplier', 'Quote Number', 'Concrete Type', 'Strength (MPa)',
               'Volume (m³)', 'Price per m³', 'Subtotal', 'GST', 'Total', 'Valid Until', 'Selected']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['BetaMix Concrete', 'BM-Q-2024-1234', 'N32 Standard Mix', 32, 45, 185, 8325, 832.50, 9157.50, '2024-07-31', 'YES'],
        ['Hanson Concrete', 'HAN-Q-2024-5678', 'N32 Standard Mix', 32, 45, 192, 8640, 864.00, 9504.00, '2024-07-31', 'NO'],
        ['Boral Concrete', 'BOR-Q-2024-9012', 'N32 Standard Mix', 32, 45, 188, 8460, 846.00, 9306.00, '2024-07-31', 'NO'],
        ['BetaMix Concrete', 'BM-Q-2024-1235', 'N40 High Strength', 40, 12, 210, 2520, 252.00, 2772.00, '2024-08-15', 'YES'],
        ['Hanson Concrete', 'HAN-Q-2024-5679', 'N40 High Strength', 40, 12, 218, 2616, 261.60, 2877.60, '2024-08-15', 'NO'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [6, 7, 8, 9]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Highlight selected rows
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 11).value == 'YES':
            for col in range(1, 12):
                ws.cell(row, col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    apply_borders_to_range(ws, 1, ws.max_row, 1, 11)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "05_QUOTES_ESTIMATES", "Concrete_Supplier_Quotes.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_framing_contractor_quotes():
    """05_QUOTES_ESTIMATES/Framing_Contractor_Quotes.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Framing Quotes"

    headers = ['Contractor', 'Quote Number', 'Scope of Work', 'Labour Cost',
               'Materials Cost', 'Subtotal', 'GST', 'Total', 'Timeline (weeks)', 'Selected']
    ws.append(headers)
    apply_header_style(ws)

    data = [
        ['Timberland Frames', 'TF-Q-2024-001', 'Complete timber framing inc walls, roof, trusses', 28000, 32000, 60000, 6000, 66000, 4, 'YES'],
        ['Ace Framing Solutions', 'AFS-Q-2024-045', 'Complete timber framing inc walls, roof, trusses', 25000, 34000, 59000, 5900, 64900, 5, 'NO'],
        ['Premium Carpentry', 'PC-Q-2024-112', 'Complete timber framing inc walls, roof, trusses', 30000, 31000, 61000, 6100, 67100, 4, 'NO'],
    ]

    for row_data in data:
        ws.append(row_data)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in [4, 5, 6, 7, 8]:
            ws.cell(row, col).number_format = '$#,##0.00'

    # Highlight selected row
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 10).value == 'YES':
            for col in range(1, 11):
                ws.cell(row, col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    apply_borders_to_range(ws, 1, ws.max_row, 1, 10)
    auto_adjust_column_width(ws)

    file_path = os.path.join(BASE_PATH, "05_QUOTES_ESTIMATES", "Framing_Contractor_Quotes.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("RECREATING ALL EXCEL SPREADSHEETS - PART 1")
    print("=" * 80)

    # 02_PERMITS_APPROVALS (1 file)
    print("\n[02_PERMITS_APPROVALS]")
    create_development_permit_tracking()

    # 03_DESIGN_DRAWINGS (2 files)
    print("\n[03_DESIGN_DRAWINGS]")
    create_architect_fees_breakdown()
    create_structural_engineer_invoices()

    # 04_FINANCE_INSURANCE (3 files)
    print("\n[04_FINANCE_INSURANCE]")
    create_construction_loan_details()
    create_insurance_policies()
    create_interest_expense_tracking()

    # 05_QUOTES_ESTIMATES (2 files)
    print("\n[05_QUOTES_ESTIMATES]")
    create_concrete_supplier_quotes()
    create_framing_contractor_quotes()

    print("\n" + "=" * 80)
    print("PART 1 COMPLETE: 8 files created")
    print("=" * 80)
