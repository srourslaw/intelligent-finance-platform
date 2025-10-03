"""
Main Orchestrator
Runs the complete financial consolidation pipeline
"""

import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .extractors.file_scanner import FileScanner
from .extractors.data_extractor import DataExtractor
from .extractors.excel_extractor import ExcelExtractor
from .classifiers.line_item_classifier import LineItemClassifier
from .consolidators.financial_consolidator import FinancialConsolidator
from .config.settings import OUTPUT_DIR, OUTPUT_FILENAME


class FinancialConsolidationPipeline:
    """Main pipeline orchestrator"""

    def __init__(self, project_data_dir: str = None):
        """
        Initialize pipeline

        Args:
            project_data_dir: Path to project data directory
        """
        self.project_data_dir = project_data_dir
        self.scanner = FileScanner(project_data_dir)
        self.data_extractor = DataExtractor()
        self.classifier = LineItemClassifier()
        self.consolidator = FinancialConsolidator()

        self.file_inventory = []
        self.extracted_data = []
        self.all_line_items = []
        self.classified_items = []
        self.consolidated_data = {}
        self.totals = {}

    def run(self, excel_only: bool = True, save_intermediates: bool = True):
        """
        Run complete consolidation pipeline

        Args:
            excel_only: Process only Excel files (faster)
            save_intermediates: Save intermediate JSON files
        """
        print("\n" + "=" * 80)
        print("FINANCIAL CONSOLIDATION PIPELINE")
        print("=" * 80)
        print(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 80 + "\n")

        # Step 1: Scan files
        print("STEP 1: Scanning files...")
        self.file_inventory = self.scanner.scan(recursive=True)

        if save_intermediates:
            self.scanner.save_inventory()

        # Step 2: Extract data
        print("\nSTEP 2: Extracting data from files...")
        if excel_only:
            excel_files = [f for f in self.file_inventory if f['file_type'] in ['.xlsx', '.xls']]
            print(f"Processing {len(excel_files)} Excel files only...")
            self.extracted_data = self.data_extractor.extract_from_files(excel_files)
        else:
            self.extracted_data = self.data_extractor.extract_from_files(self.file_inventory)

        if save_intermediates:
            self.data_extractor.save_extracted_data()

        # Step 3: Extract line items from all sheets
        print("\nSTEP 3: Extracting financial line items...")
        self.all_line_items = self._extract_all_line_items()
        print(f"Extracted {len(self.all_line_items)} line items")

        if save_intermediates:
            self._save_line_items()

        # Step 4: Classify line items
        print("\nSTEP 4: Classifying line items...")
        self.classified_items = self.classifier.classify_batch(self.all_line_items)

        stats = self.classifier.get_classification_stats(self.classified_items)
        print(f"Classification complete:")
        print(f"  Total items: {stats['total']}")
        print(f"  Classified: {stats['classified']} ({stats['classification_rate']:.1f}%)")
        print(f"  High confidence: {stats['high_confidence']}")

        if save_intermediates:
            self._save_classified_items()

        # Step 5: Consolidate data
        print("\nSTEP 5: Consolidating financial data...")
        self.consolidated_data = self.consolidator.consolidate(self.classified_items)
        self.totals = self.consolidator.get_totals(self.consolidated_data)

        # Print summary
        summary = self.consolidator.generate_summary(self.consolidated_data, self.totals)
        print(summary)

        if save_intermediates:
            self._save_consolidated_data()

        # Step 6: Generate Excel output
        print("\nSTEP 6: Generating consolidated Excel file...")
        self._generate_excel_output()

        print("\n" + "=" * 80)
        print("PIPELINE COMPLETE")
        print("=" * 80)
        print(f"End time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 80 + "\n")

    def _extract_all_line_items(self) -> list:
        """Extract all line items from extracted data"""
        line_items = []

        for file_data in self.extracted_data:
            if file_data.get('error'):
                continue

            for sheet in file_data.get('sheets', []):
                for row in sheet.get('data', []):
                    # Find description and amount columns
                    description = None
                    amount = None

                    for key, value in row.items():
                        if value is None or key == '_row_number':
                            continue

                        key_lower = key.lower()

                        # Look for description
                        if not description and any(kw in key_lower for kw in
                                                   ['description', 'item', 'category', 'account', 'line item', 'name']):
                            if isinstance(value, str) and len(value) > 2:
                                description = value

                        # Look for amount
                        if not amount and any(kw in key_lower for kw in
                                            ['amount', 'total', 'cost', 'price', 'value', 'budget', 'actual', 'balance']):
                            if isinstance(value, (int, float)):
                                amount = float(value)
                            elif isinstance(value, str):
                                # Try to parse string amount
                                try:
                                    cleaned = value.replace('$', '').replace(',', '').strip()
                                    amount = float(cleaned)
                                except:
                                    pass

                    # Add line item if both description and amount found
                    if description and amount and amount != 0:
                        line_items.append({
                            'description': description,
                            'amount': amount,
                            'source_file': file_data['filename'],
                            'source_sheet': sheet['sheet_name'],
                            'row_number': row.get('_row_number')
                        })

        return line_items

    def _save_line_items(self):
        """Save extracted line items to JSON"""
        output_path = OUTPUT_DIR / 'line_items.json'
        with open(output_path, 'w') as f:
            json.dump({
                'total_items': len(self.all_line_items),
                'line_items': self.all_line_items
            }, f, indent=2)
        print(f"  ✅ Line items saved to: {output_path}")

    def _save_classified_items(self):
        """Save classified items to JSON"""
        output_path = OUTPUT_DIR / 'classified_items.json'
        with open(output_path, 'w') as f:
            json.dump({
                'total_items': len(self.classified_items),
                'classified_items': self.classified_items
            }, f, indent=2, default=str)
        print(f"  ✅ Classified items saved to: {output_path}")

    def _save_consolidated_data(self):
        """Save consolidated data to JSON"""
        output_path = OUTPUT_DIR / 'consolidated_data.json'
        with open(output_path, 'w') as f:
            json.dump({
                'consolidated_data': self.consolidated_data,
                'totals': self.totals
            }, f, indent=2)
        print(f"  ✅ Consolidated data saved to: {output_path}")

    def _generate_excel_output(self):
        """Generate final consolidated Excel file"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create Balance Sheet
            self._create_balance_sheet(wb)

            # Create Income Statement
            self._create_income_statement(wb)

            # Create Summary Sheet
            self._create_summary_sheet(wb)

            # Save workbook
            output_path = OUTPUT_DIR / OUTPUT_FILENAME
            wb.save(output_path)

            print(f"  ✅ Excel file generated: {output_path}")

        except Exception as e:
            print(f"  ❌ Error generating Excel: {e}")

    def _create_balance_sheet(self, wb):
        """Create Balance Sheet worksheet"""
        ws = wb.create_sheet("Balance Sheet")

        # Header
        ws['A1'] = "BALANCE SHEET"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"As at {datetime.now().strftime('%B %d, %Y')}"

        row = 4

        # Assets
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        # Current Assets
        ws[f'A{row}'] = "Current Assets"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for line_item, amount in self.consolidated_data.get('current_assets', {}).items():
            ws[f'A{row}'] = f"  {line_item}"
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1

        ws[f'A{row}'] = "Total Current Assets"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = self.totals.get('current_assets', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2

        # Non-Current Assets
        ws[f'A{row}'] = "Non-Current Assets"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for line_item, amount in self.consolidated_data.get('non_current_assets', {}).items():
            ws[f'A{row}'] = f"  {line_item}"
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1

        ws[f'A{row}'] = "Total Non-Current Assets"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = self.totals.get('non_current_assets', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2

        # Total Assets
        ws[f'A{row}'] = "TOTAL ASSETS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = self.totals['total_assets']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        row += 3

        # Liabilities & Equity (similar pattern)
        ws[f'A{row}'] = "LIABILITIES & EQUITY"
        ws[f'A{row}'].font = Font(bold=True, size=14)

        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20

    def _create_income_statement(self, wb):
        """Create Income Statement worksheet"""
        ws = wb.create_sheet("Income Statement")

        ws['A1'] = "INCOME STATEMENT"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Period Ended {datetime.now().strftime('%B %d, %Y')}"

        row = 4

        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        for line_item, amount in self.consolidated_data.get('revenue', {}).items():
            ws[f'A{row}'] = f"  {line_item}"
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1

        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = self.totals['total_revenue']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2

        # COGS, Expenses, etc. (similar pattern)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20

    def _create_summary_sheet(self, wb):
        """Create Summary worksheet"""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet

        ws['A1'] = "FINANCIAL CONSOLIDATION SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 5
        ws[f'A{row}'] = "Key Metrics:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        metrics = [
            ('Total Assets', self.totals['total_assets']),
            ('Total Liabilities', self.totals['total_liabilities']),
            ('Total Equity', self.totals['total_equity']),
            ('Total Revenue', self.totals['total_revenue']),
            ('Gross Profit', self.totals['gross_profit']),
            ('Operating Income', self.totals['operating_income']),
            ('Net Income', self.totals['net_income']),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20


def main():
    """Main execution"""
    pipeline = FinancialConsolidationPipeline()
    pipeline.run(excel_only=True, save_intermediates=True)


if __name__ == "__main__":
    main()
