"""
Excel Extractor Module
Extracts financial data from Excel files (.xlsx, .xls)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Any, Optional
import re

from ..config.settings import EXCEL_MAX_ROWS_TO_SCAN, EXCEL_HEADER_ROW_SEARCH_LIMIT


class ExcelExtractor:
    """Extract data from Excel files"""

    def __init__(self, file_path: str):
        """
        Initialize Excel extractor

        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.extracted_data = {}

    def extract(self) -> Dict[str, Any]:
        """
        Extract all data from Excel file

        Returns:
            Dictionary with extracted data from all sheets
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

            self.extracted_data = {
                'file_path': str(self.file_path),
                'filename': self.file_path.name,
                'sheets': [],
                'total_sheets': len(self.workbook.sheetnames)
            }

            for sheet_name in self.workbook.sheetnames:
                sheet_data = self._extract_sheet(sheet_name)
                if sheet_data:
                    self.extracted_data['sheets'].append(sheet_data)

            self.workbook.close()
            return self.extracted_data

        except Exception as e:
            return {
                'file_path': str(self.file_path),
                'filename': self.file_path.name,
                'error': str(e),
                'sheets': []
            }

    def _extract_sheet(self, sheet_name: str) -> Optional[Dict]:
        """
        Extract data from a single sheet

        Args:
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet data or None if empty
        """
        ws = self.workbook[sheet_name]

        # Check if sheet is empty
        if ws.max_row == 0 or ws.max_column == 0:
            return None

        # Find header row
        header_row_idx = self._find_header_row(ws)

        if header_row_idx is None:
            # No clear header found, treat first row as header
            header_row_idx = 1

        # Extract headers
        headers = self._extract_headers(ws, header_row_idx)

        # Extract data rows
        data_rows = self._extract_data_rows(ws, header_row_idx, headers)

        # Only return sheet data if there are actual data rows
        if not data_rows:
            return None

        return {
            'sheet_name': sheet_name,
            'header_row': header_row_idx,
            'headers': headers,
            'data': data_rows,
            'row_count': len(data_rows),
            'column_count': len(headers)
        }

    def _find_header_row(self, ws) -> Optional[int]:
        """
        Find the header row by looking for rows with text values

        Args:
            ws: Worksheet object

        Returns:
            Row index of header row or None
        """
        for row_idx in range(1, min(EXCEL_HEADER_ROW_SEARCH_LIMIT + 1, ws.max_row + 1)):
            row = ws[row_idx]

            # Count non-empty text cells
            text_cells = 0
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    text_cells += 1

            # If row has multiple text cells, likely a header
            if text_cells >= 2:
                return row_idx

        return None

    def _extract_headers(self, ws, header_row_idx: int) -> List[str]:
        """
        Extract column headers

        Args:
            ws: Worksheet object
            header_row_idx: Index of header row

        Returns:
            List of header names
        """
        headers = []
        row = ws[header_row_idx]

        for col_idx, cell in enumerate(row, start=1):
            if cell.value:
                # Clean header name
                header = str(cell.value).strip()
                # Remove extra whitespace
                header = re.sub(r'\s+', ' ', header)
                headers.append(header)
            else:
                # Use column letter for empty headers
                headers.append(f"Column_{get_column_letter(col_idx)}")

        return headers

    def _extract_data_rows(self, ws, header_row_idx: int, headers: List[str]) -> List[Dict]:
        """
        Extract data rows

        Args:
            ws: Worksheet object
            header_row_idx: Index of header row
            headers: List of column headers

        Returns:
            List of dictionaries (one per row)
        """
        data_rows = []
        max_row = min(ws.max_row, header_row_idx + EXCEL_MAX_ROWS_TO_SCAN)

        for row_idx in range(header_row_idx + 1, max_row + 1):
            row = ws[row_idx]

            # Check if row is empty
            if all(cell.value is None for cell in row):
                continue

            # Create row dictionary
            row_data = {}
            for col_idx, cell in enumerate(row[:len(headers)]):
                header = headers[col_idx]
                value = cell.value

                # Clean and standardize value
                if value is not None:
                    if isinstance(value, str):
                        value = value.strip()
                        # Remove extra whitespace
                        value = re.sub(r'\s+', ' ', value)
                    row_data[header] = value
                else:
                    row_data[header] = None

            # Only add row if it has at least one non-None value
            if any(v is not None for v in row_data.values()):
                row_data['_row_number'] = row_idx
                data_rows.append(row_data)

        return data_rows

    def extract_simple(self) -> List[List[Any]]:
        """
        Extract raw data as 2D list (simpler format)

        Returns:
            List of lists (rows and columns)
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = self.workbook.active

            data = []
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    data.append(list(row))

            self.workbook.close()
            return data

        except Exception as e:
            print(f"Error extracting {self.file_path}: {e}")
            return []

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names"""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"Error getting sheet names from {self.file_path}: {e}")
            return []

    def extract_financial_data(self) -> Dict[str, List[Dict]]:
        """
        Extract financial data with smart detection

        Returns:
            Dictionary mapping data types to extracted records
        """
        extracted = self.extract()
        financial_data = {
            'line_items': [],
            'transactions': [],
            'balances': []
        }

        for sheet in extracted.get('sheets', []):
            # Analyze sheet structure
            sheet_type = self._detect_sheet_type(sheet)

            if sheet_type == 'line_items':
                # Extract line items (e.g., budget categories, GL accounts)
                financial_data['line_items'].extend(
                    self._extract_line_items(sheet)
                )
            elif sheet_type == 'transactions':
                # Extract transactions (e.g., invoices, payments)
                financial_data['transactions'].extend(
                    self._extract_transactions(sheet)
                )
            elif sheet_type == 'balances':
                # Extract balances (e.g., account balances)
                financial_data['balances'].extend(
                    self._extract_balances(sheet)
                )

        return financial_data

    def _detect_sheet_type(self, sheet: Dict) -> str:
        """Detect what type of financial data is in the sheet"""
        headers_lower = [h.lower() for h in sheet.get('headers', [])]

        # Check for transaction indicators
        transaction_keywords = ['invoice', 'date', 'amount', 'vendor', 'payment', 'transaction']
        if sum(any(kw in h for kw in transaction_keywords) for h in headers_lower) >= 3:
            return 'transactions'

        # Check for balance indicators
        balance_keywords = ['balance', 'opening', 'closing', 'debit', 'credit']
        if sum(any(kw in h for kw in balance_keywords) for h in headers_lower) >= 2:
            return 'balances'

        # Default to line items
        return 'line_items'

    def _extract_line_items(self, sheet: Dict) -> List[Dict]:
        """Extract line items from sheet"""
        line_items = []

        for row in sheet.get('data', []):
            # Look for description and amount columns
            description = None
            amount = None

            for key, value in row.items():
                key_lower = key.lower()

                if value is not None:
                    # Find description
                    if any(kw in key_lower for kw in ['description', 'item', 'category', 'account', 'line item']):
                        if isinstance(value, str):
                            description = value

                    # Find amount
                    if any(kw in key_lower for kw in ['amount', 'total', 'cost', 'price', 'value', 'budget', 'actual']):
                        if isinstance(value, (int, float)):
                            amount = float(value)

            if description and amount:
                line_items.append({
                    'description': description,
                    'amount': amount,
                    'source_sheet': sheet['sheet_name'],
                    'row_number': row.get('_row_number'),
                    'raw_data': row
                })

        return line_items

    def _extract_transactions(self, sheet: Dict) -> List[Dict]:
        """Extract transactions from sheet"""
        transactions = []

        for row in sheet.get('data', []):
            transaction = {
                'source_sheet': sheet['sheet_name'],
                'row_number': row.get('_row_number')
            }

            for key, value in row.items():
                if value is not None and key != '_row_number':
                    key_lower = key.lower()

                    # Map common fields
                    if 'date' in key_lower:
                        transaction['date'] = value
                    elif any(kw in key_lower for kw in ['invoice', 'reference', 'ref', 'number']):
                        transaction['reference'] = str(value)
                    elif any(kw in key_lower for kw in ['vendor', 'supplier', 'customer', 'client']):
                        transaction['party'] = str(value)
                    elif any(kw in key_lower for kw in ['amount', 'total']):
                        if isinstance(value, (int, float)):
                            transaction['amount'] = float(value)
                    elif 'description' in key_lower:
                        transaction['description'] = str(value)

            # Only add if has essential fields
            if transaction.get('amount') is not None:
                transactions.append(transaction)

        return transactions

    def _extract_balances(self, sheet: Dict) -> List[Dict]:
        """Extract account balances from sheet"""
        balances = []

        for row in sheet.get('data', []):
            balance = {
                'source_sheet': sheet['sheet_name'],
                'row_number': row.get('_row_number')
            }

            for key, value in row.items():
                if value is not None and key != '_row_number':
                    key_lower = key.lower()

                    if any(kw in key_lower for kw in ['account', 'description', 'line item']):
                        if isinstance(value, str):
                            balance['account'] = value
                    elif 'debit' in key_lower:
                        if isinstance(value, (int, float)):
                            balance['debit'] = float(value)
                    elif 'credit' in key_lower:
                        if isinstance(value, (int, float)):
                            balance['credit'] = float(value)
                    elif any(kw in key_lower for kw in ['balance', 'amount', 'total']):
                        if isinstance(value, (int, float)):
                            balance['balance'] = float(value)

            if balance.get('account'):
                balances.append(balance)

        return balances


def main():
    """Test function"""
    test_file = "/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data/12_BUDGET_TRACKING/MASTER_PROJECT_BUDGET.xlsx"

    extractor = ExcelExtractor(test_file)
    data = extractor.extract()

    print(f"File: {data['filename']}")
    print(f"Total sheets: {data['total_sheets']}")

    for sheet in data['sheets']:
        print(f"\nSheet: {sheet['sheet_name']}")
        print(f"  Rows: {sheet['row_count']}, Columns: {sheet['column_count']}")
        print(f"  Headers: {sheet['headers'][:5]}...")
        if sheet['data']:
            print(f"  Sample row: {list(sheet['data'][0].keys())[:3]}")


if __name__ == "__main__":
    main()
