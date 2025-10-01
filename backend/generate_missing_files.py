"""
Generate all missing dummy data files for construction project
Creates realistic, messy Excel and PDF files with inconsistencies
"""
import os
import random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from PIL import Image, ImageDraw, ImageFont

# Base directory
BASE_DIR = Path(__file__).parent / "dummy_data"

# Helper functions
def random_date(start_date, end_date):
    """Generate random date between two dates"""
    time_between = end_date - start_date
    days_between = time_between.days
    random_days = random.randint(0, days_between)
    return start_date + timedelta(days=random_days)

def create_messy_excel(file_path, sheet_data):
    """Create Excel file with messy, realistic data"""
    wb = openpyxl.Workbook()

    for sheet_name, data in sheet_data.items():
        if sheet_name == 'Sheet':
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        headers = data['headers']
        rows = data['rows']

        # Write headers with formatting
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write data rows with random messiness
        for row_idx, row_data in enumerate(rows, 2):
            # Randomly skip rows to create gaps
            if random.random() < 0.05:  # 5% chance of empty row
                row_idx += 1

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Add random formatting issues
                if random.random() < 0.02:  # 2% chance of formula error
                    cell.value = "#REF!"
                elif random.random() < 0.03:  # 3% chance of #N/A
                    cell.value = "#N/A"

    wb.save(file_path)
    print(f"✓ Created: {file_path}")

def create_pdf_document(file_path, title, content_lines):
    """Create simple PDF document"""
    c = canvas.Canvas(str(file_path), pagesize=letter)
    width, height = letter

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1*inch, height - 1*inch, title)

    # Content
    c.setFont("Helvetica", 10)
    y_position = height - 1.5*inch
    for line in content_lines:
        c.drawString(1*inch, y_position, line)
        y_position -= 0.3*inch
        if y_position < 1*inch:
            c.showPage()
            y_position = height - 1*inch

    c.save()
    print(f"✓ Created: {file_path}")

def create_placeholder_image(file_path, text):
    """Create placeholder image with text"""
    img = Image.new('RGB', (800, 600), color=(200, 200, 200))
    draw = ImageDraw.Draw(img)

    # Draw text in center
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 40)
    except:
        font = ImageFont.load_default()

    # Get text bounding box
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    position = ((800 - text_width) // 2, (600 - text_height) // 2)
    draw.text(position, text, fill=(50, 50, 50), font=font)

    img.save(file_path)
    print(f"✓ Created: {file_path}")

print("=" * 60)
print("GENERATING MISSING FILES - PHASE 1: CRITICAL EXCEL FILES")
print("=" * 60)

# 01_LAND_PURCHASE
print("\n[01_LAND_PURCHASE]")
folder = BASE_DIR / "01_LAND_PURCHASE"
folder.mkdir(parents=True, exist_ok=True)

# Land_Costs.xlsx
create_messy_excel(folder / "Land_Costs.xlsx", {
    "Land Costs": {
        "headers": ["Item", "Description", "Amount", "GST", "Total", "Date Paid", "Status"],
        "rows": [
            ["Land Purchase", "123 Sunset Boulevard", 250000, 0, 250000, "2024-06-15", "PAID"],
            ["Stamp Duty", "NSW Stamp Duty", 9970, 0, 9970, "2024-06-15", "PAID"],
            ["Legal Fees", "Johnson & Partners Solicitors", 2800, 280, 3080, "2024-06-20", "PAID"],
            ["Title Transfer", "Land Registry", 150, 0, 150, "2024-06-20", "PAID"],
            ["Soil Test", "GeoTech Surveyors", 1850, 185, 2035, "2024-06-25", "PAID"],
            ["Survey Report", "Land Survey Services", 1200, 120, 1320, "2024-06-22", "PAID"],
            ["", "", "", "", "=SUM(E2:E7)", "", ""],  # Formula
        ]
    }
})

# Stamp_Duty_Calculation.xlsx
create_messy_excel(folder / "Stamp_Duty_Calculation.xlsx", {
    "Calculation": {
        "headers": ["Property Value", "Stamp Duty Rate", "Calculated Amount", "Actual Paid", "Difference"],
        "rows": [
            [250000, "3.99%", 9970, 9970, 0],
            ["", "", "", "", ""],
            ["Notes:", "NSW First Home Buyer", "", "", ""],
            ["", "Concession NOT applied", "", "", ""],
        ]
    }
})

print("\n[02_PERMITS_APPROVALS]")
folder = BASE_DIR / "02_PERMITS_APPROVALS"
folder.mkdir(parents=True, exist_ok=True)

# Permits_Costs_Tracker.xlsx
create_messy_excel(folder / "Permits_Costs_Tracker.xlsx", {
    "Permits": {
        "headers": ["Permit Type", "Authority", "Application Fee", "Inspection Fees", "Total", "Status", "Date Approved"],
        "rows": [
            ["Development Application", "Sydney Council", 2850, 0, 2850, "Approved", "2024-07-10"],
            ["Building Permit", "Certified Builder", 4200, 850, 5050, "Approved", "2024-07-25"],
            ["Plumbing Approval", "Sydney Water", 340, 200, 540, "Approved", "2024-08-05"],
            ["Electrical Approval", "Level 2 Electrician", 280, 150, 430, "Approved", "2024-08-10"],
            ["Water Connection", "Sydney Water", 1200, 0, 1200, "Pending", ""],
            ["Occupancy Certificate", "Council", 680, 0, 680, "Not Yet", ""],
        ]
    }
})

print("\n[03_DESIGN_DRAWINGS]")
folder = BASE_DIR / "03_DESIGN_DRAWINGS" / "Architectural"
folder.mkdir(parents=True, exist_ok=True)

# Design_Fees.xlsx
create_messy_excel(folder / "Design_Fees.xlsx", {
    "Fees": {
        "headers": ["Service", "Provider", "Quote", "Actual", "Variance", "Paid?"],
        "rows": [
            ["Architectural Design", "Smith & Associates", 12000, 12800, -800, "YES"],
            ["Structural Engineering", "BuildSafe Eng.", 4500, 4890, -390, "YES"],
            ["Energy Rating", "EnergyCert", 650, 650, 0, "Yes"],
            ["Council Liaison", "Smith & Associates", 800, 1200, -400, "YES"],
            ["Revisions (x3)", "Extra charges", 0, 2400, -2400, "yes"],
            ["", "", "=SUM(C2:C6)", "=SUM(D2:D6)", "=SUM(E2:E6)", ""],
        ]
    }
})

folder = BASE_DIR / "03_DESIGN_DRAWINGS" / "Engineering"
folder.mkdir(parents=True, exist_ok=True)

# engineering_costs.xlsx
create_messy_excel(folder / "engineering_costs.xlsx", {
    "Costs": {
        "headers": ["Engineering Service", "Cost", "Date", "Notes"],
        "rows": [
            ["Structural Plans", 3200, "2024-06-28", "Initial design"],
            ["Structural Plans REV A", 890, "2024-07-15", "Revisions"],
            ["Electrical Design", 1450, "2024-07-05", ""],
            ["Plumbing Design", 1200, "2024-07-08", ""],
            ["BASIX Certificate", 450, "2024-07-12", "Energy compliance"],
        ]
    }
})

print(f"\n✓ Phase 1 Complete - Created {len(list((BASE_DIR / '01_LAND_PURCHASE').glob('*.xlsx')))} Excel files in folder 01")
print(f"✓ Created {len(list((BASE_DIR / '02_PERMITS_APPROVALS').glob('*.xlsx')))} Excel files in folder 02")
print(f"✓ Created {len(list((BASE_DIR / '03_DESIGN_DRAWINGS').rglob('*.xlsx')))} Excel files in folder 03")
