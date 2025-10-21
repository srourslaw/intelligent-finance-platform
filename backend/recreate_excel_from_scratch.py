#!/usr/bin/env python3
"""
RECREATE ALL EXCEL FILES FROM SCRATCH
Create logical, meaningful spreadsheets for construction project management
Remove all nonsense data and create dashboard-ready information
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
import random

BASE_DIR = Path("/Users/husseinsrour/Downloads/intelligent-finance-platform/backend/projects/project-a-123-sunset-blvd/data")

# Styling
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row=1, max_col=10):
    """Apply professional styling to header row"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def auto_size_columns(ws):
    """Auto-size all columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

print("="*80)
print("RECREATING ALL EXCEL FILES FROM SCRATCH")
print("Creating logical, dashboard-ready construction project data")
print("="*80)

# ============================================================================
# 01_LAND_PURCHASE - Land Acquisition Costs
# ============================================================================
print("\n[01_LAND_PURCHASE]")
folder = BASE_DIR / "01_LAND_PURCHASE"

# Land Costs
filepath = folder / "Land_Costs.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Land Acquisition"

headers = ['Cost Category', 'Vendor/Authority', 'Amount (Ex GST)', 'GST', 'Total (Inc GST)', 'Payment Date', 'Payment Method', 'Status']
ws.append(headers)

data = [
    ['Land Purchase Price', 'Smith Family Trust', 250000, 0, 250000, '2024-06-15', 'Bank Transfer', 'PAID'],
    ['Stamp Duty', 'NSW Revenue', 9970, 0, 9970, '2024-06-15', 'BPAY', 'PAID'],
    ['Legal Fees - Conveyancing', 'Johnson & Partners Solicitors', 2800, 280, 3080, '2024-06-20', 'Cheque', 'PAID'],
    ['Title Transfer Fee', 'NSW Land Registry', 150, 0, 150, '2024-06-20', 'Credit Card', 'PAID'],
    ['Building & Pest Inspection', 'Premium Building Inspections', 550, 55, 605, '2024-05-28', 'Direct Debit', 'PAID'],
    ['Soil Testing', 'GeoTech Australia Pty Ltd', 1850, 185, 2035, '2024-06-25', 'Bank Transfer', 'PAID'],
    ['Survey & Boundary Marking', 'Land Survey Services', 980, 98, 1078, '2024-06-28', 'Credit Card', 'PAID'],
    ['Section 10.7 Certificate', 'Sydney Council', 180, 0, 180, '2024-05-20', 'BPAY', 'PAID'],
]

for row in data:
    ws.append(row)

# Add total row
total_row = ws.max_row + 2
ws.cell(total_row, 1, "TOTAL LAND ACQUISITION COSTS")
ws.cell(total_row, 3, f"=SUM(C2:C{ws.max_row - 1})")
ws.cell(total_row, 4, f"=SUM(D2:D{ws.max_row - 1})")
ws.cell(total_row, 5, f"=SUM(E2:E{ws.max_row - 1})")
for col in range(1, 6):
    ws.cell(total_row, col).fill = TOTAL_FILL
    ws.cell(total_row, col).font = TOTAL_FONT

style_header_row(ws, 1, 8)
auto_size_columns(ws)

# Format currency columns
for row in range(2, ws.max_row + 1):
    for col in [3, 4, 5]:
        ws.cell(row, col).number_format = '$#,##0.00'

wb.save(filepath)
print(f"  ✅ {filepath.name}")

# Stamp Duty Calculation
filepath = folder / "Stamp_Duty_Calculation.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Stamp Duty Breakdown"

ws['A1'] = "NSW STAMP DUTY CALCULATION"
ws['A1'].font = Font(size=14, bold=True)
ws.merge_cells('A1:C1')

ws.append(['', '', ''])
ws.append(['Property Purchase Price', '', 250000])
ws.append(['Property Type', '', 'Vacant Land - Residential'])
ws.append(['Buyer Status', '', 'First Home Buyer'])
ws.append(['', '', ''])

ws.append(['Calculation Breakdown', 'Rate', 'Amount'])
style_header_row(ws, 7, 3)

ws.append(['First $14,000', '1.25%', 175])
ws.append(['$14,001 - $31,000', '1.50%', 255])
ws.append(['$31,001 - $83,000', '1.75%', 910])
ws.append(['$83,001 - $310,000', '3.50%', 7945])
ws.append(['Over $310,000', '0%', 0])
ws.append(['', '', ''])
ws.append(['TOTAL STAMP DUTY', '', 9970])

ws['C14'].fill = TOTAL_FILL
ws['C14'].font = TOTAL_FONT

for row in range(8, 14):
    ws.cell(row, 3).number_format = '$#,##0.00'
ws['C3'].number_format = '$#,##0'

auto_size_columns(ws)
wb.save(filepath)
print(f"  ✅ {filepath.name}")

# ============================================================================
# 12_BUDGET_TRACKING - Master Project Budget
# ============================================================================
print("\n[12_BUDGET_TRACKING]")
folder = BASE_DIR / "12_BUDGET_TRACKING"
folder.mkdir(parents=True, exist_ok=True)

filepath = folder / "MASTER_PROJECT_BUDGET.xlsx"
wb = openpyxl.Workbook()

# Budget Summary Sheet
ws = wb.active
ws.title = "Budget Summary"

ws['A1'] = "SUNSET BOULEVARD PROJECT - MASTER BUDGET"
ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
ws.merge_cells('A1:G1')

ws['A2'] = f"Project: 123 Sunset Boulevard, Sydney NSW 2000"
ws['A3'] = f"Budget Date: {datetime.now().strftime('%d %B %Y')}"
ws.append(['', '', '', '', '', '', ''])

headers = ['Cost Category', 'Budget Amount', 'Actual Spent', 'Variance', '% Spent', 'Status', 'Notes']
ws.append(headers)
style_header_row(ws, 5, 7)

budget_data = [
    ['Land & Acquisition', 266480, 266480, '=B6-C6', '=C6/B6', 'Complete', 'All settlement costs paid'],
    ['Design & Engineering', 28500, 24850, '=B7-C7', '=C7/B7', 'Complete', 'Final invoices received'],
    ['Council Permits & Approvals', 8950, 8950, '=B8-C8', '=C8/B8', 'Complete', 'All permits approved'],
    ['Site Establishment', 12400, 11280, '=B9-C9', '=C9/B9', 'In Progress', 'Scaffolding ongoing'],
    ['Foundation & Slab', 45200, 42850, '=B10-C10', '=C10/B10', 'Complete', 'Slab poured and cured'],
    ['Frame & Structure', 85600, 78200, '=B11-C11', '=C11/B11', 'In Progress', 'Roof trusses installed'],
    ['Windows & External Doors', 24800, 24800, '=B12-C12', '=C12/B12', 'Complete', 'All units installed'],
    ['Roofing & Gutters', 18200, 16500, '=B13-C13', '=C13/B13', 'In Progress', '80% complete'],
    ['External Cladding', 32400, 15200, '=B14-C14', '=C14/B14', 'In Progress', 'Brickwork underway'],
    ['Plumbing (Rough-in)', 22800, 18400, '=B15-C15', '=C15/B15', 'In Progress', 'First fix complete'],
    ['Electrical (Rough-in)', 28600, 22100, '=B16-C16', '=C16/B16', 'In Progress', 'First fix underway'],
    ['Insulation & Sarking', 8900, 0, '=B17-C17', '=C17/B17', 'Not Started', 'Scheduled next month'],
    ['Plasterboard & Lining', 24500, 0, '=B18-C18', '=C18/B18', 'Not Started', 'Post electrical rough-in'],
    ['Tiling (Bathrooms & Kitchen)', 18400, 0, '=B19-C19', '=C19/B19', 'Not Started', 'Awaiting plastering'],
    ['Kitchen & Cabinetry', 38500, 19250, '=B20-C20', '=C20/B20', 'In Progress', '50% deposit paid'],
    ['Bathroom Fit-out', 16200, 0, '=B21-C21', '=C21/B21', 'Not Started', 'Materials ordered'],
    ['Flooring', 19800, 0, '=B22-C22', '=C22/B22', 'Not Started', 'Timber on order'],
    ['Painting (Internal & External)', 22600, 0, '=B23-C23', '=C23/B23', 'Not Started', 'Scheduled month 8'],
    ['Final Electrical (Fit-off)', 12400, 0, '=B24-C24', '=C24/B24', 'Not Started', 'Post painting'],
    ['Final Plumbing (Fit-off)', 10200, 0, '=B25-C25', '=C25/B25', 'Not Started', 'Post tiling'],
    ['Landscaping & Driveway', 28400, 0, '=B26-C26', '=C26/B26', 'Not Started', 'Final stage'],
    ['Contingency (5%)', 32500, 8450, '=B27-C27', '=C27/B27', 'Partial', 'Variation orders'],
]

for row_data in budget_data:
    ws.append(row_data)

# Total row
total_row = ws.max_row + 1
ws.cell(total_row, 1, "TOTAL PROJECT BUDGET")
ws.cell(total_row, 2, f"=SUM(B6:B{ws.max_row})")
ws.cell(total_row, 3, f"=SUM(C6:C{ws.max_row})")
ws.cell(total_row, 4, f"=B{total_row}-C{total_row}")
ws.cell(total_row, 5, f"=C{total_row}/B{total_row}")

for col in range(1, 6):
    ws.cell(total_row, col).fill = TOTAL_FILL
    ws.cell(total_row, col).font = TOTAL_FONT

# Format currency and percentages
for row in range(6, total_row + 1):
    for col in [2, 3, 4]:
        ws.cell(row, col).number_format = '$#,##0'
    ws.cell(row, 5).number_format = '0%'

auto_size_columns(ws)

# Cash Flow Sheet
ws = wb.create_sheet("Cash Flow Forecast")
ws['A1'] = "PROJECT CASH FLOW FORECAST"
ws['A1'].font = Font(size=14, bold=True)
ws.merge_cells('A1:E1')

ws.append(['', '', '', '', ''])
headers = ['Month', 'Planned Spend', 'Actual Spend', 'Variance', 'Cumulative Spent']
ws.append(headers)
style_header_row(ws, 3, 5)

months = ['June 2024', 'July 2024', 'August 2024', 'September 2024', 'October 2024',
          'November 2024', 'December 2024', 'January 2025', 'February 2025', 'March 2025']
planned = [266000, 35000, 58000, 72000, 48000, 45000, 38000, 28000, 24000, 36000]
actual = [266480, 33850, 55120, 68200, 45850, 0, 0, 0, 0, 0]

for i, month in enumerate(months):
    row_num = i + 4
    ws.append([month, planned[i], actual[i] if i < 5 else 0,
               f'=B{row_num}-C{row_num}',
               f'=SUM(C$4:C{row_num})'])

for row in range(4, 14):
    for col in [2, 3, 4, 5]:
        ws.cell(row, col).number_format = '$#,##0'

auto_size_columns(ws)

wb.save(filepath)
print(f"  ✅ {filepath.name}")

# ============================================================================
# 06_PURCHASE_ORDERS_INVOICES - Invoices and POs
# ============================================================================
print("\n[06_PURCHASE_ORDERS_INVOICES]")
folder = BASE_DIR / "06_PURCHASE_ORDERS_INVOICES"

# Materials Purchase Summary
filepath = folder / "Materials_Purchases_Summary.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Materials Summary"

headers = ['Date', 'Supplier', 'Material Type', 'Quantity', 'Unit', 'Unit Price', 'Subtotal', 'GST', 'Total', 'Invoice #', 'Status']
ws.append(headers)

materials_data = [
    ['2024-07-22', 'BetaMix Concrete', 'Concrete 25MPa', 12, 'm³', 285, '=D2*F2', '=G2*0.1', '=G2+H2', 'BMX-8821', 'PAID'],
    ['2024-07-22', 'BetaMix Concrete', 'Concrete 32MPa', 18, 'm³', 315, '=D3*F3', '=G3*0.1', '=G3+H3', 'BMX-8821', 'PAID'],
    ['2024-07-18', 'Ace Steel Reinforcing', 'N12 Rebar', 650, 'kg', 1.85, '=D4*F4', '=G4*0.1', '=G4+H4', 'ACE-5589', 'PAID'],
    ['2024-07-18', 'Ace Steel Reinforcing', 'F82 Mesh', 180, 'm²', 8.50, '=D5*F5', '=G5*0.1', '=G5+H5', 'ACE-5589', 'PAID'],
    ['2024-08-01', 'Timberland Frames', 'Pine Framing 90x45', 285, 'LM', 12.80, '=D6*F6', '=G6*0.1', '=G6+H6', 'TFT-2914', 'PAID'],
    ['2024-08-01', 'Timberland Frames', 'Roof Trusses', 1, 'set', 8850, '=D7*F7', '=G7*0.1', '=G7+H7', 'TFT-2914', 'PAID'],
    ['2024-08-20', 'Premium Windows & Doors', 'Aluminum Windows', 8, 'units', 685, '=D8*F8', '=G8*0.1', '=G8+H8', 'PWD-7733', 'PAID'],
    ['2024-08-20', 'Premium Windows & Doors', 'Entry Door', 1, 'unit', 2850, '=D9*F9', '=G9*0.1', '=G9+H9', 'PWD-7733', 'PAID'],
    ['2024-09-05', 'BuildMart Supplies', 'Roof Tiles - Terracotta', 520, 'm²', 24.50, '=D10*F10', '=G10*0.1', '=G10+H10', 'BM-4521', 'DUE'],
    ['2024-09-05', 'BuildMart Supplies', 'Colorbond Guttering', 45, 'm', 32, '=D11*F11', '=G11*0.1', '=G11+H11', 'BM-4521', 'DUE'],
]

for row in materials_data:
    ws.append(row)

style_header_row(ws, 1, 11)

for row in range(2, ws.max_row + 1):
    for col in [6, 7, 8, 9]:
        ws.cell(row, col).number_format = '$#,##0.00'

auto_size_columns(ws)
wb.save(filepath)
print(f"  ✅ {filepath.name}")

print("\n" + "="*80)
print("✅ EXCEL FILES RECREATED FROM SCRATCH")
print("="*80)
print("\nAll files now contain:")
print("  ✓ Logical, meaningful column names")
print("  ✓ Realistic construction project data")
print("  ✓ Dashboard-ready information")
print("  ✓ Proper formulas and calculations")
print("  ✓ Professional formatting")
print("  ✓ No empty rows or nonsense data")
print("="*80 + "\n")
