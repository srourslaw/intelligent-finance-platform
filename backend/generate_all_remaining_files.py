"""
Generate ALL remaining missing files - comprehensive script
Phase 2: All critical Excel, PDF, and image files
"""
import os
import random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont

BASE_DIR = Path(__file__).parent / "dummy_data"

def create_messy_excel(file_path, sheet_data):
    """Create Excel with messy data"""
    wb = openpyxl.Workbook()
    for sheet_name, data in sheet_data.items():
        if sheet_name == 'Sheet':
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        headers = data['headers']
        rows = data['rows']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for row_idx, row_data in enumerate(rows, 2):
            if random.random() < 0.03:
                row_idx += 1
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)
    print(f"✓ {file_path.name}")

def create_pdf(file_path, title, lines):
    """Create PDF"""
    c = canvas.Canvas(str(file_path), pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1*inch, height - 1*inch, title)
    c.setFont("Helvetica", 10)
    y = height - 1.5*inch
    for line in lines:
        c.drawString(1*inch, y, line)
        y -= 0.25*inch
        if y < 1*inch:
            c.showPage()
            y = height - 1*inch
    c.save()
    print(f"✓ {file_path.name}")

def create_image(file_path, text):
    """Create placeholder image"""
    img = Image.new('RGB', (800, 600), color=(220, 220, 220))
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 36)
    except:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), text, font=font)
    w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(((800-w)//2, (600-h)//2), text, fill=(60, 60, 60), font=font)
    img.save(file_path)
    print(f"✓ {file_path.name}")

print("\n" + "="*60)
print("PHASE 2: GENERATING ALL REMAINING FILES")
print("="*60)

# 04_FINANCE_INSURANCE
print("\n[04_FINANCE_INSURANCE]")
folder = BASE_DIR / "04_FINANCE_INSURANCE"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "Loan_Drawdown_Schedule.xlsx", {
    "Drawdowns": {
        "headers": ["Stage", "Description", "Amount Approved", "Amount Drawn", "Date", "Balance Remaining"],
        "rows": [
            ["Deposit", "Land Purchase", 250000, 250000, "2024-06-15", 400000],
            ["Slab", "Foundation Complete", 80000, 80000, "2024-08-20", 320000],
            ["Frame", "Frame & Roof", 120000, 120000, "2024-09-25", 200000],
            ["Lockup", "Windows/Doors", 100000, 0, "", 200000],
            ["Fixing", "Internal Fit-out", 80000, 0, "", 200000],
            ["Completion", "Final Payment", 120000, 0, "", 200000],
        ]
    }
})

create_messy_excel(folder / "Interest_Calculations.xlsx", {
    "Interest": {
        "headers": ["Month", "Principal Drawn", "Interest Rate", "Monthly Interest", "Cumulative Interest"],
        "rows": [
            ["June 2024", 250000, "6.5%", 1354, 1354],
            ["July 2024", 250000, "6.5%", 1354, 2708],
            ["August 2024", 330000, "6.5%", 1787, 4495],
            ["Sept 2024", 450000, "6.5%", 2437, 6932],
            ["Oct 2024", 450000, "6.7%", 2513, 9445],
        ]
    }
})

create_messy_excel(folder / "Financing_Costs_Summary.xlsx", {
    "Summary": {
        "headers": ["Item", "Amount", "Paid", "Status"],
        "rows": [
            ["Loan Application Fee", 800, "Yes", "PAID"],
            ["Valuation Fee", 450, "Yes", "PAID"],
            ["Legal Fees - Loan", 1200, "Yes", "PAID"],
            ["Builder's Insurance", 3850, "Yes", "PAID"],
            ["Public Liability", 2100, "Yes", "PAID"],
            ["Interest Paid YTD", 9445, "Yes", "PAID"],
        ]
    }
})

# 05_QUOTES_ESTIMATES
print("\n[05_QUOTES_ESTIMATES]")
folder = BASE_DIR / "05_QUOTES_ESTIMATES"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "Initial_Project_Budget_Estimate.xlsx", {
    "Budget Estimate": {
        "headers": ["Category", "Estimated Cost", "Notes"],
        "rows": [
            ["Land & Legals", 265000, "Including stamp duty"],
            ["Design & Engineering", 18000, "Architect + Engineer"],
            ["Permits & Approvals", 12000, "Council + certifier"],
            ["Site Preparation", 25000, "Clear + excavate"],
            ["Foundation", 45000, "Footings + slab"],
            ["Frame & Roof", 65000, "Timber frame"],
            ["External Walls", 55000, "Brick veneer"],
            ["Windows & Doors", 28000, ""],
            ["Plumbing", 32000, "Rough-in + fixtures"],
            ["Electrical", 28000, ""],
            ["Insulation", 8000, ""],
            ["Plasterboard", 22000, ""],
            ["Kitchen", 25000, "Mid-range"],
            ["Bathrooms", 18000, "2 bathrooms"],
            ["Flooring", 15000, "Tiles + carpet"],
            ["Painting", 12000, "Internal + external"],
            ["Landscaping", 8000, "Basic"],
            ["Contingency", 35000, "5% buffer"],
            ["TOTAL", "=SUM(B2:B19)", ""],
        ]
    }
})

create_messy_excel(folder / "Quotes_Comparison.xlsx", {
    "Excavation": {
        "headers": ["Company", "Quote", "Days", "Notes", "Selected?"],
        "rows": [
            ["DirtMovers Pty Ltd", 8500, 3, "Includes disposal", "NO"],
            ["BigDig Excavation", 7800, 4, "Cheaper but slower", "YES"],
            ["QuickDig", 9200, 2, "Fast but expensive", "NO"],
        ]
    },
    "Concrete": {
        "headers": ["Company", "Quote", "m3", "Notes", "Selected?"],
        "rows": [
            ["ReadyMix Concrete", 12500, 45, "Standard mix", "YES"],
            ["Premium Concrete Co", 14200, 45, "High strength", "NO"],
            ["BudgetMix", 11800, 45, "Concerns re quality", "NO"],
        ]
    }
})

# 06_PURCHASE_ORDERS_INVOICES
print("\n[06_PURCHASE_ORDERS_INVOICES]")
folder = BASE_DIR / "06_PURCHASE_ORDERS_INVOICES"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "Invoices_Paid" / "Paid_Invoices_Register.xlsx", {
    "Register": {
        "headers": ["Invoice #", "Supplier", "Amount", "GST", "Total", "Date Paid", "Payment Method"],
        "rows": [
            ["RM-2024-8845", "ReadyMix Concrete", 12500, 1250, 13750, "2024-08-22", "EFT"],
            ["TSC-INV-4421", "Timber Supply Co", 18200, 1820, 20020, "2024-09-05", "EFT"],
            ["SES-2024-3421", "Spark Electric Supply", 8450, 845, 9295, "2024-09-12", "Credit Card"],
            ["BH-2024-0847", "Brick House Supplies", 15600, 1560, 17160, "2024-09-18", "EFT"],
            ["APS-2024-8912", "Aqua Plumbing Supply", 9200, 920, 10120, "2024-09-20", "EFT"],
            ["TR-2024-156", "TopRoof Materials", 12800, 1280, 14080, "2024-09-22", "EFT"],
            ["SF-PC-002", "SolidFoundations", 24500, 2450, 26950, "2024-08-25", "EFT"],
            ["TB-PC-001", "TimberBuilders", 18900, 1890, 20790, "2024-09-10", "EFT"],
            ["BR-PC-003", "BuildRight Framing", 32100, 3210, 35310, "2024-09-20", "EFT"],
            ["PPS-8834", "Premier Paint Supply", 4200, 420, 4620, "2024-09-28", "Credit Card"],
        ]
    }
})

create_messy_excel(folder / "Invoices_Pending" / "Pending_Invoices.xlsx", {
    "Pending": {
        "headers": ["Invoice #", "Supplier", "Amount", "Due Date", "Days Overdue", "Priority"],
        "rows": [
            ["INV-092024-PP", "Premier Painters", 11200, "2024-10-05", 0, "URGENT"],
            ["INV-TILES-001", "Premium Tiles", 8900, "2024-10-10", 0, "High"],
            ["LAND-092024", "GreenScapes Landscaping", 7200, "2024-10-15", 0, "Medium"],
        ]
    }
})

create_messy_excel(folder / "Materials_Purchases_Summary.xlsx", {
    "Summary": {
        "headers": ["Category", "Budget", "Actual Spent", "Committed", "Remaining"],
        "rows": [
            ["Concrete & Cement", 15000, 13750, 0, 1250],
            ["Timber & Framing", 22000, 20020, 0, 1980],
            ["Roofing Materials", 15000, 14080, 0, 920],
            ["Bricks & Blocks", 18000, 17160, 0, 840],
            ["Plumbing Fixtures", 12000, 10120, 0, 1880],
            ["Electrical Supplies", 10000, 9295, 0, 705],
            ["Paint & Finishes", 6000, 4620, 11200, -9820],
            ["Tiles & Flooring", 14000, 0, 8900, 5100],
            ["Landscaping", 8000, 0, 7200, 800],
        ]
    }
})

# 07_SUBCONTRACTORS
print("\n[07_SUBCONTRACTORS]")
folder = BASE_DIR / "07_SUBCONTRACTORS"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "SubcontractorPaymentTracker.xlsx", {
    "Payments": {
        "headers": ["Subcontractor", "Contract Value", "Paid to Date", "Retention", "Balance Due", "Status"],
        "rows": [
            ["Joe's Excavation", 7800, 7800, 0, 0, "Complete"],
            ["SolidFoundations", 26950, 26950, 0, 0, "Complete"],
            ["BuildRight Framing", 35310, 35310, 0, 0, "Complete"],
            ["Spark Electric", 28400, 18200, 2840, 7360, "In Progress"],
            ["AquaFlow Plumbing", 31200, 19800, 3120, 8280, "In Progress"],
            ["TopRoof", 14080, 14080, 0, 0, "Complete"],
            ["Premier Bricklaying", 42000, 32000, 4200, 5800, "In Progress"],
            ["ColorPro Painters", 11200, 0, 1120, 10080, "Starting"],
        ]
    }
})

# 08_LABOUR_TIMESHEETS
print("\n[08_LABOUR_TIMESHEETS]")
folder = BASE_DIR / "08_LABOUR_TIMESHEETS"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "Timesheets_Aug_2024.xlsx", {
    "Site Supervisor": {
        "headers": ["Date", "Employee", "Hours", "Rate", "Cost", "Task"],
        "rows": [
            ["2024-08-01", "John Smith", 8, 55, 440, "Site setup"],
            ["2024-08-02", "John Smith", 9, 55, 495, "Excavation supervision"],
            ["2024-08-05", "John Smith", 8, 55, 440, "Footings inspection"],
            ["2024-08-12", "John Smith", 10, 55, 550, "Concrete pour"],
            ["2024-08-19", "John Smith", 8, 55, 440, "Slab inspection"],
            ["2024-08-26", "John Smith", 8, 55, 440, "Frame prep"],
        ]
    },
    "Labour": {
        "headers": ["Date", "Worker", "Role", "Hours", "Rate", "Cost"],
        "rows": [
            ["2024-08-01", "Mike Johnson", "General Labour", 8, 32, 256],
            ["2024-08-01", "Dave Wilson", "General Labour", 8, 32, 256],
            ["2024-08-02", "Mike Johnson", "General Labour", 8, 32, 256],
            ["2024-08-05", "Steve Brown", "Skilled Labour", 8, 42, 336],
        ]
    }
})

create_messy_excel(folder / "Site_Supervisor_Hours.xlsx", {
    "Summary": {
        "headers": ["Month", "Total Hours", "Rate", "Total Cost", "Notes"],
        "rows": [
            ["June 2024", 40, 55, 2200, "Part month - setup"],
            ["July 2024", 168, 55, 9240, "Full month"],
            ["August 2024", 176, 55, 9680, "Full month"],
            ["Sept 2024", 184, 55, 10120, "Overtime"],
        ]
    }
})

create_messy_excel(folder / "Labour_Costs_Summary.xlsx", {
    "Summary": {
        "headers": ["Category", "Budget", "Actual", "Variance"],
        "rows": [
            ["Site Supervisor", 32000, 31240, 760],
            ["General Labour", 28000, 24890, 3110],
            ["Skilled Labour", 18000, 16200, 1800],
            ["TOTAL", 78000, 72330, 5670],
        ]
    }
})

create_messy_excel(folder / "Employee_Wage_Rates.xlsx", {
    "Rates": {
        "headers": ["Employee", "Role", "Hourly Rate", "Superannuation", "Insurance"],
        "rows": [
            ["John Smith", "Site Supervisor", 55, "11%", "WorkCover"],
            ["Mike Johnson", "General Labour", 32, "11%", "WorkCover"],
            ["Dave Wilson", "General Labour", 32, "11%", "WorkCover"],
            ["Steve Brown", "Skilled Labour", 42, "11%", "WorkCover"],
        ]
    }
})

# 09_SITE_REPORTS_PHOTOS
print("\n[09_SITE_REPORTS_PHOTOS]")
folder = BASE_DIR / "09_SITE_REPORTS_PHOTOS"
folder.mkdir(parents=True, exist_ok=True)

create_messy_excel(folder / "Site_Diary_August.xlsx", {
    "Daily Log": {
        "headers": ["Date", "Weather", "Trades On Site", "Work Completed", "Issues"],
        "rows": [
            ["2024-08-01", "Sunny", "Excavation", "Site cleared", ""],
            ["2024-08-02", "Cloudy", "Excavation", "Trenches dug", "Minor delay - rocks"],
            ["2024-08-05", "Rain", "None", "NO WORK", "Weather delay"],
            ["2024-08-12", "Sunny", "Concreter", "Footings poured", ""],
            ["2024-08-19", "Sunny", "Concreter", "Slab poured", ""],
            ["2024-08-26", "Sunny", "Framing", "Frame started", ""],
        ]
    }
})

create_messy_excel(folder / "Site_Diary_September.xlsx", {
    "Daily Log": {
        "headers": ["Date", "Weather", "Trades On Site", "Work Completed", "Issues"],
        "rows": [
            ["2024-09-02", "Sunny", "Framing", "Walls up", ""],
            ["2024-09-09", "Windy", "Framing", "Roof trusses", ""],
            ["2024-09-16", "Sunny", "Roofer", "Roof complete", ""],
            ["2024-09-23", "Rain", "Electrician", "Rough-in started", "Material delay"],
            ["2024-09-30", "Sunny", "Plumber", "Rough-in", ""],
        ]
    }
})

create_messy_excel(folder / "Weather_Delays_Log.xlsx", {
    "Delays": {
        "headers": ["Date", "Weather", "Hours Lost", "Impact", "Cost Impact"],
        "rows": [
            ["2024-08-05", "Heavy Rain", 8, "No work", 880],
            ["2024-08-06", "Rain", 4, "Half day", 440],
            ["2024-09-11", "Strong Wind", 6, "Roof work stopped", 660],
        ]
    }
})

# Create images
print("\n[09_SITE_REPORTS_PHOTOS - Images]")
photo_folder = folder / "Site_Photos"
photo_folder.mkdir(parents=True, exist_ok=True)

create_image(photo_folder / "20240801_Site_Cleared.jpg", "Site Cleared\n2024-08-01")
create_image(photo_folder / "20240815_Excavation_Complete.jpg", "Excavation Complete\n2024-08-15")
create_image(photo_folder / "20240820_Footings_Poured.jpg", "Footings Poured\n2024-08-20")
create_image(photo_folder / "20240825_Slab_Down.jpg", "Slab Down\n2024-08-25")
create_image(photo_folder / "20240905_Frame_Started.jpg", "Frame Started\n2024-09-05")
create_image(photo_folder / "20240920_Frame_Complete.jpg", "Frame Complete\n2024-09-20")
create_image(photo_folder / "20241001_Roof_On.jpg", "Roof On\n2024-10-01")

print("\n" + "="*60)
print("PHASE 2 COMPLETE!")
print("="*60)
print(f"\n✓ Folders 04-09 populated with Excel files and images")
print(f"✓ Ready for Phase 3 (remaining folders 10-18)")
